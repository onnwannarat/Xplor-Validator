"""
Payment Plan Import Checker
============================
Validates and prepares data for Payment Plan Import into the Onboarding Tool.

Validation rules aligned with the Onboarding Tool Error List:
  ERROR_MISSING_GUARDIAN           → No parent name provided
  ERROR_MISSING_PAYMENT_DAY        → Weekday is missing
  ERROR_INVALID_PAYMENT_DAY        → Weekday is Saturday or Sunday
  ERROR_INVALID_FREQUENCY          → Billing Cycle is not Weekly / Fortnightly / Monthly
  ERROR_MISSING_BOOKING_START_DATE → Start Date is missing
  ERROR_NEGATIVE_DIRECT_DEBIT_LIMIT → Direct Debit Limit is negative
  ERROR_NEGATIVE_FIXED_LIMIT       → Fixed Amount is negative
  ERROR_ONLY_ONE_AMOUNT_ALLOWED    → Both Limit and Fixed Amount are greater than zero

Additional checks:
  - Manual (Paused) plan Start Date must be a Monday
  - Service ID is required
  - Trailing spaces in names are stripped
  - Date format converted to DD/MM/YYYY
  - Weekday converted to 3-letter abbreviation
"""

import csv
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ─────────────────────────── Constants ──────────────────────────────────────

DEFAULT_COLUMNS = {
    "service_id": "Service ID",
    "service": "Service Name",
    "parent_id": "Parent Legacy ID",
    "parent_fn": "Primary Guardian First Name",
    "parent_ln": "Primary Guardian Last Name",
    "child_id": "Child Legacy ID",
    "child_fn": "Child First Name",
    "child_ln": "Child Last Name",
    "date": "Start Date",
    "weekday": "Weekday",
    "manual": "Manual",
    "cycle": "Cycle",
    "limit": "Limit",
    "fixed_amount": "Fixed Amount",
    "gateway": "Gateway Reference",
}

# Label shown in UI for each column key
COLUMN_LABELS = {
    "service_id": "Service ID  *required",
    "service": "Service Name",
    "parent_id": "Parent Legacy ID",
    "parent_fn": "Parent First Name",
    "parent_ln": "Parent Last Name",
    "child_id": "Child Legacy ID",
    "child_fn": "Child First Name",
    "child_ln": "Child Last Name",
    "date": "Start Date",
    "weekday": "Weekday",
    "manual": "Manual (Yes/No)",
    "cycle": "Billing Cycle",
    "limit": "Direct Debit Limit",
    "fixed_amount": "Fixed Amount",
    "gateway": "Gateway Reference",
}

WEEKDAY_MAP = {
    "monday": "Mon", "tuesday": "Tue", "wednesday": "Wed",
    "thursday": "Thu", "friday": "Fri", "saturday": "Sat",
    "sunday": "Sun",
    "mon": "Mon", "tue": "Tue", "wed": "Wed", "thu": "Thu",
    "fri": "Fri", "sat": "Sat", "sun": "Sun",
}

VALID_CYCLES = {"weekly", "fortnightly", "monthly"}
WEEKEND_DAYS = {"Sat", "Sun"}

# Output CSV template — must match payment_plan_onboarding_tools.csv
TEMPLATE_COLUMNS = [
    "Service_ID", "Service_Name", "Parent_Legacy_Id",
    "Parent_First_Name", "Parent_Last_Name", "Child_Legacy_Id",
    "Child_First_Name", "Child_Last_Name", "Direct_Debit_Start_Date",
    "Direct_Debit_Day", "Manual", "Billing_Cycle",
    "Direct_Debit_Limit", "Fixed_Amount", "Gateway_Reference",
]

DATE_FORMATS = [
    "%d/%m/%Y %I:%M:%S %p", "%d/%m/%Y %I:%M %p",       # 12-hour with AM/PM
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    "%Y-%m-%d %I:%M:%S %p", "%Y-%m-%d %I:%M %p",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
    "%d/%m/%y",
]


# ─────────────────────────── Helpers ────────────────────────────────────────

def parse_date(raw: str):
    raw = raw.strip()
    if not raw:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def is_numeric(val: str) -> bool:
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False


def _entry(row, col, key, default=""):
    col_name = col.get(key, "")
    return row.get(col_name, default).strip() if col_name else default


# ─────────────────────────── Core processor ─────────────────────────────────

def process_csv(filepath: str, col: dict) -> dict:
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    # Error buckets — order here controls display priority
    errors = {
        "weekend": [],   # ERROR_INVALID_PAYMENT_DAY (Sat/Sun)
        "missing_date": [],   # ERROR_MISSING_BOOKING_START_DATE
        "missing_weekday": [],   # ERROR_MISSING_PAYMENT_DAY
        "missing_parent": [],   # ERROR_MISSING_GUARDIAN
        "missing_service_id": [],   # Service ID required
        "invalid_cycle": [],   # ERROR_INVALID_FREQUENCY
        "manual_not_monday": [],   # Manual plan must start on Monday
        "negative_limit": [],   # ERROR_NEGATIVE_DIRECT_DEBIT_LIMIT
        "negative_fixed": [],   # ERROR_NEGATIVE_FIXED_LIMIT
        "both_amounts": [],   # ERROR_ONLY_ONE_AMOUNT_ALLOWED
        "unparseable_date": [],   # Date string unrecognised
        "unknown_weekday": [],   # Weekday value not in map
    }

    stats = {
        "total": len(rows),
        "date_fixed": 0,
        "weekday_fixed": 0,
        "spaces_fixed": 0,
    }

    processed = []
    for i, raw_row in enumerate(rows, start=2):
        row = dict(raw_row)

        # ── Strip all whitespace ──────────────────────────────────────────
        for key in row:
            stripped = row[key].strip()
            if stripped != row[key]:
                stats["spaces_fixed"] += 1
            row[key] = stripped

        # Shorthand helpers for this row
        def g(key): return _entry(row, col, key)

        parent_fn = g("parent_fn")
        parent_ln = g("parent_ln")
        parent_name = f"{parent_fn} {parent_ln}".strip()
        service = g("service")
        parent_id = g("parent_id")
        child_id = g("child_id")
        service_id = g("service_id")

        ctx = dict(row=i, parent_id=parent_id, child_id=child_id,
                   parent_name=parent_name, service=service)

        # ── 1. Fix & validate Start Date ─────────────────────────────────
        raw_date = g("date")
        parsed_dt = None
        if raw_date:
            parsed_dt = parse_date(raw_date)
            if parsed_dt:
                fixed = parsed_dt.strftime("%d/%m/%Y")
                if fixed != raw_date:
                    stats["date_fixed"] += 1
                col_name = col.get("date", "")
                if col_name:
                    row[col_name] = fixed
            else:
                errors["unparseable_date"].append({**ctx, "value": raw_date})
        else:
            errors["missing_date"].append({**ctx, "weekday": g("weekday") or "(empty)"})

        # ── 2. Fix & validate Weekday ─────────────────────────────────────
        raw_wd = g("weekday")
        wd_fixed = WEEKDAY_MAP.get(raw_wd.lower(), "")
        col_name_wd = col.get("weekday", "")
        if wd_fixed:
            if wd_fixed != raw_wd:
                stats["weekday_fixed"] += 1
            if col_name_wd:
                row[col_name_wd] = wd_fixed
        elif raw_wd:
            errors["unknown_weekday"].append({**ctx, "value": raw_wd})
        else:
            errors["missing_weekday"].append({**ctx, "date": raw_date or "(empty)"})

        # ── 3. Weekend check ──────────────────────────────────────────────
        if wd_fixed in WEEKEND_DAYS:
            date_str = row.get(col.get("date", ""), "")
            errors["weekend"].append({**ctx, "weekday": wd_fixed, "date": date_str})

        # ── 4. Parent name ────────────────────────────────────────────────
        if not parent_fn or not parent_ln:
            errors["missing_parent"].append({
                **ctx,
                "first_name": parent_fn or "(empty)",
                "last_name": parent_ln or "(empty)",
            })

        # ── 5. Service ID required ────────────────────────────────────────
        if not service_id:
            errors["missing_service_id"].append(ctx)

        # ── 6. Billing Cycle validation ───────────────────────────────────
        cycle = g("cycle")
        if cycle and cycle.lower() not in VALID_CYCLES:
            errors["invalid_cycle"].append({**ctx, "value": cycle})

        # ── 7. Manual plan → start date must be Monday ───────────────────
        manual_val = g("manual").lower()
        is_manual = manual_val in ("yes", "1", "true")
        if is_manual and parsed_dt and parsed_dt.weekday() != 0:   # 0 = Monday
            day_name = parsed_dt.strftime("%A")
            errors["manual_not_monday"].append({
                **ctx,
                "date": parsed_dt.strftime("%d/%m/%Y"),
                "day": day_name,
            })

        # ── 8. Limit / Fixed Amount checks ───────────────────────────────
        limit_raw = g("limit")
        fixed_raw = g("fixed_amount")

        if limit_raw:
            if is_numeric(limit_raw):
                if float(limit_raw) < 0:
                    errors["negative_limit"].append({**ctx, "value": limit_raw})
            # (non-numeric → the tool imports as 0, so we don't flag it here)

        if fixed_raw:
            if is_numeric(fixed_raw):
                if float(fixed_raw) < 0:
                    errors["negative_fixed"].append({**ctx, "value": fixed_raw})

        # Both amounts > 0 (ERROR_ONLY_ONE_AMOUNT_ALLOWED)
        try:
            limit_num = float(limit_raw) if limit_raw else 0
            fixed_num = float(fixed_raw) if fixed_raw else 0
            if limit_num > 0 and fixed_num > 0:
                errors["both_amounts"].append({
                    **ctx,
                    "limit": limit_raw,
                    "fixed": fixed_raw,
                })
        except (ValueError, TypeError):
            pass

        row["_row_num"] = i   # preserve original file row number for later mapping
        processed.append(row)

    return {
        "processed_rows": processed,
        "fieldnames": fieldnames,
        "errors": errors,
        "stats": stats,
    }


# ─────────────────────────── File writers ───────────────────────────────────

def _output_dir(filepath: str) -> Path:
    """Return (and create) the Output folder.

    If the source file lives inside a folder named 'Input', the Output folder
    is placed alongside that Input folder (i.e. at the same level).
    Otherwise it is placed next to the source file as before.
    """
    src = Path(filepath).parent
    base = src.parent if src.name.lower() == "input" else src
    out = base / "Output"
    out.mkdir(exist_ok=True)
    return out


def write_cleaned_csv(filepath: str, fieldnames: list, rows: list) -> str:
    out = _output_dir(filepath) / f"{Path(filepath).stem}_cleaned.csv"
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return str(out)


def load_service_mapping(csv_path: str) -> dict:
    """Load serviceIDs.csv → {qk_service_id_str: {"xplor_id": str, "name": str}}."""
    mapping = {}
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            qk_id = row.get("QKServiceID", "").strip()
            xplor_id = row.get("Xplor Service ID", "").strip()
            name = row.get("Service Name", "").strip()
            if qk_id:
                mapping[qk_id] = {"xplor_id": xplor_id, "name": name}
    return mapping


def _safe_filename(name: str) -> str:
    """Remove characters that are invalid in Windows filenames."""
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip()


def write_split_csvs(filepath: str, col: dict, rows: list, service_map: dict) -> dict:
    """
    Split processed rows by Service ID and write one CSV per service.

    - Rows whose Service ID is found in service_map  → Output/<ServiceName>_payment_plan_import.csv
    - Rows whose Service ID is NOT found             → Output/unknown/<id>_payment_plan_import.csv
    - Service Name column is replaced with the correct name from service_map.
    - All output files follow TEMPLATE_COLUMNS column order.

    Returns {"known": {service_name: path}, "unknown": {service_id: path}}
    """
    out_dir = _output_dir(filepath)
    unknown_dir = out_dir / "unknown"

    # Group rows by service_id value
    groups: dict[str, list] = {}
    for row in rows:
        sid = _entry(row, col, "service_id")
        groups.setdefault(sid, []).append(row)

    known_files: dict[str, str] = {}
    unknown_files: dict[str, str] = {}
    row_map: dict[int, int] = {}   # {original_row_num: service_file_row_num}

    for sid, group_rows in groups.items():
        svc = service_map.get(sid)
        is_known = svc is not None
        if is_known:
            service_name = svc["name"]
            xplor_id = svc["xplor_id"]
            dest = out_dir / (_safe_filename(service_name) + "_payment_plan_import.csv")
        else:
            unknown_dir.mkdir(exist_ok=True)
            label = sid if sid else "no_service_id"
            dest = unknown_dir / (_safe_filename(label) + "_payment_plan_import.csv")
            service_name = _entry(group_rows[0], col, "service")   # keep original
            xplor_id = sid

        template_rows = []
        for file_row_num, row in enumerate(group_rows, start=2):  # row 1 = header
            orig = row.get("_row_num")
            if orig is not None:
                row_map[orig] = file_row_num

            def g(key, _row=row): return _entry(_row, col, key)
            template_rows.append({
                "Service_ID": xplor_id,
                "Service_Name": service_name,
                "Parent_Legacy_Id": g("parent_id"),
                "Parent_First_Name": g("parent_fn"),
                "Parent_Last_Name": g("parent_ln"),
                "Child_Legacy_Id": f"{xplor_id}_{g('child_id')}" if g("child_id") else "",
                "Child_First_Name": g("child_fn"),
                "Child_Last_Name": g("child_ln"),
                "Direct_Debit_Start_Date": g("date"),
                "Direct_Debit_Day": g("weekday"),
                "Manual": g("manual"),
                "Billing_Cycle": g("cycle"),
                "Direct_Debit_Limit": g("limit"),
                "Fixed_Amount": g("fixed_amount"),
                "Gateway_Reference": g("gateway"),
            })

        with open(dest, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=TEMPLATE_COLUMNS)
            writer.writeheader()
            writer.writerows(template_rows)

        if is_known:
            known_files[service_name] = str(dest)
        else:
            unknown_files[sid or "(empty)"] = str(dest)

    return {"known": known_files, "unknown": unknown_files, "row_map": row_map}


def _translate_error_rows(errors: dict, row_map: dict) -> None:
    """Replace original-file row numbers with service-file row numbers in-place."""
    for items in errors.values():
        for e in items:
            orig = e.get("row")
            if orig in row_map:
                e["row"] = row_map[orig]


def write_error_report(filepath: str, errors: dict) -> str:
    """Write a colour-coded Excel error report (.xlsx)."""
    out = _output_dir(filepath) / f"{Path(filepath).stem}_error_report.xlsx"

    # ── Colour palette ───────────────────────────────────────────────────
    C_WEEKEND_HEADER = "C0392B"   # deep red
    C_WEEKEND_ROW = "FADBD8"   # light pink
    C_ERROR_HEADER = "E67E22"   # orange
    C_ERROR_ROW = "FDEBD0"   # light orange
    C_WARN_HEADER = "F1C40F"   # yellow
    C_WARN_ROW = "FEF9E7"   # light yellow
    C_COL_HEADER = "1A5276"   # navy (column header row)
    C_WHITE = "FFFFFF"
    C_LIGHT_GRAY = "F2F2F2"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def bold_font(color=None, size=10):
        return Font(bold=True, color=color or "000000", size=size)

    def std_font(color=None):
        return Font(color=color or "000000", size=10)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Error definitions: (key, tool_error_key, display_label, note_fn, severity)
    # severity: "weekend" | "error" | "warn"
    PRIORITY = [
        ("weekend",
         "ERROR_INVALID_PAYMENT_DAY",
         "WEEKEND — Sat / Sun",
         lambda e: f"Falls on a {
             'Sunday' if e.get('weekday') == 'Sun' else 'Saturday'} ({
             e.get(
                 'date',
                 '')}) — please verify before importing",
            "weekend"),
        ("missing_date",
         "ERROR_MISSING_BOOKING_START_DATE",
         "Missing Start Date",
         lambda e: "Start Date is empty — cannot import",
         "error"),
        ("missing_weekday",
         "ERROR_MISSING_PAYMENT_DAY",
         "Missing Weekday",
         lambda e: "Weekday is empty — cannot import",
         "error"),
        ("missing_parent",
         "ERROR_MISSING_GUARDIAN",
         "Missing Parent Name",
         lambda e: f"First: \"{
             e.get(
                 'first_name',
                 '')}\"  Last: \"{
             e.get(
                 'last_name',
                 '')}\"",
         "error"),
        ("missing_service_id",
         "ERROR_MISSING_SERVICE_ID",
         "Missing Service ID",
         lambda e: "Service ID is empty (required field)",
         "error"),
        ("invalid_cycle",
         "ERROR_INVALID_FREQUENCY",
         "Invalid Billing Cycle",
         lambda e: f"Value: \"{
             e.get(
                 'value',
                 '')}\" — must be Weekly / Fortnightly / Monthly",
         "error"),
        ("manual_not_monday",
         "MANUAL_PLAN_NOT_MONDAY",
         "Manual Plan — Not Monday",
         lambda e: f"Starts on {
             e.get(
                 'day',
                 '')} ({
             e.get(
                 'date',
                 '')}) — Manual plans must start on Monday",
         "error"),
        ("negative_limit",
         "ERROR_NEGATIVE_DIRECT_DEBIT_LIMIT",
         "Negative Direct Debit Limit",
         lambda e: f"Value = {
             e.get(
                 'value',
                 '')}  (must be >= 0)",
         "error"),
        ("negative_fixed",
         "ERROR_NEGATIVE_FIXED_LIMIT",
         "Negative Fixed Amount",
         lambda e: f"Value = {
             e.get(
                 'value',
                 '')}  (must be >= 0)",
         "error"),
        ("both_amounts",
         "ERROR_ONLY_ONE_AMOUNT_ALLOWED",
         "Both Limit + Fixed Amount Set",
         lambda e: f"Limit={
             e.get(
                 'limit',
                 '')}  AND  Fixed={
             e.get(
                 'fixed',
                 '')}  (only one allowed)",
         "error"),
        ("unparseable_date",
         "UNPARSEABLE_DATE",
         "Unparseable Date Value",
         lambda e: f"Cannot parse: \"{
             e.get(
                 'value',
                 '')}\"",
         "warn"),
        ("unknown_weekday",
         "UNKNOWN_WEEKDAY",
         "Unknown Weekday Value",
         lambda e: f"Unrecognised value: \"{
             e.get(
                 'value',
                 '')}\"",
         "warn"),
    ]

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════
    # Sheet 1: Summary
    # ════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    # Title
    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = "Payment Plan Import — Error Report"
    ws_sum["A1"].font = Font(bold=True, size=16, color=C_WHITE)
    ws_sum["A1"].fill = fill(C_COL_HEADER)
    ws_sum["A1"].alignment = center
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells("A2:D2")
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  Source: {Path(filepath).name}"
    ws_sum["A2"].font = Font(italic=True, size=9, color="666666")
    ws_sum["A2"].alignment = left
    ws_sum.row_dimensions[2].height = 16

    ws_sum.append([])

    # Column headers
    ws_sum.append(["Error Category", "Onboarding Tool Error Key", "Count", "Severity"])
    for col_idx, cell in enumerate(ws_sum[ws_sum.max_row], 1):
        cell.font = bold_font(C_WHITE)
        cell.fill = fill(C_COL_HEADER)
        cell.alignment = center
        cell.border = border
    ws_sum.row_dimensions[ws_sum.max_row].height = 20

    total = 0
    for key, tool_key, label, _, sev in PRIORITY:
        n = len(errors.get(key, []))
        total += n
        row_data = [label, tool_key, n, sev.upper()]
        ws_sum.append(row_data)
        r = ws_sum.max_row
        c_fill = fill(C_WEEKEND_ROW if sev == "weekend"
                      else C_ERROR_ROW if sev == "error"
                      else C_WARN_ROW)
        c_font_count = bold_font("C0392B") if n > 0 else std_font("2ECC71")
        for col_idx, cell in enumerate(ws_sum[r], 1):
            cell.fill = c_fill
            cell.alignment = left if col_idx <= 2 else center
            cell.border = border
            cell.font = std_font()
        ws_sum[r][2].font = c_font_count  # count cell
        ws_sum.row_dimensions[r].height = 18

    # Total row
    ws_sum.append(["TOTAL ERRORS", "", total, ""])
    r = ws_sum.max_row
    t_fill = fill("C0392B") if total > 0 else fill("1E8449")
    for cell in ws_sum[r]:
        cell.font = bold_font(C_WHITE, size=11)
        cell.fill = t_fill
        cell.alignment = center
        cell.border = border
    ws_sum.row_dimensions[r].height = 22

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 38
    ws_sum.column_dimensions["C"].width = 10
    ws_sum.column_dimensions["D"].width = 14

    # ════════════════════════════════════════════════════════════════════
    # Sheet 2: Detail (all errors)
    # ════════════════════════════════════════════════════════════════════
    ws_det = wb.create_sheet("Error Detail")
    ws_det.sheet_view.showGridLines = False
    ws_det.freeze_panes = "A3"

    # Title row
    ws_det.merge_cells("A1:I1")
    ws_det["A1"] = "Error Detail — All Issues Found"
    ws_det["A1"].font = Font(bold=True, size=13, color=C_WHITE)
    ws_det["A1"].fill = fill(C_COL_HEADER)
    ws_det["A1"].alignment = center
    ws_det.row_dimensions[1].height = 24

    # Column headers
    COLS = ["#", "Error Category", "Tool Error Key",
            "Row #", "Parent Legacy ID", "Child Legacy ID",
            "Parent Name", "Service Name", "Note / Detail"]
    ws_det.append(COLS)
    for cell in ws_det[2]:
        cell.font = bold_font(C_WHITE)
        cell.fill = fill(C_COL_HEADER)
        cell.alignment = center
        cell.border = border
    ws_det.row_dimensions[2].height = 20

    seq = 0
    for key, tool_key, label, note_fn, sev in PRIORITY:
        items = errors.get(key, [])
        if not items:
            continue

        # Section divider
        ws_det.append(["", f"── {label.upper()}  ({len(items)} rows) ──"] + [""] * 7)
        r = ws_det.max_row
        ws_det.merge_cells(f"B{r}:I{r}")
        c_hdr = (C_WEEKEND_HEADER if sev == "weekend"
                 else C_ERROR_HEADER if sev == "error"
                 else C_WARN_HEADER)
        for cell in ws_det[r]:
            cell.fill = fill(c_hdr)
            cell.font = bold_font(C_WHITE)
            cell.alignment = left
        ws_det.row_dimensions[r].height = 18

        c_row_fill = fill(C_WEEKEND_ROW if sev == "weekend"
                          else C_ERROR_ROW if sev == "error"
                          else C_WARN_ROW)
        alt_fill = fill(C_LIGHT_GRAY)

        for i, e in enumerate(items):
            seq += 1
            note = note_fn(e)
            ws_det.append([
                seq, label, tool_key,
                e.get("row", ""), e.get("parent_id", ""), e.get("child_id", ""),
                e.get("parent_name", ""), e.get("service", ""), note,
            ])
            r = ws_det.max_row
            row_fill = c_row_fill if i % 2 == 0 else alt_fill
            for col_idx, cell in enumerate(ws_det[r], 1):
                cell.fill = row_fill
                cell.border = border
                cell.alignment = center if col_idx in (1, 4) else left
                cell.font = std_font()
            ws_det.row_dimensions[r].height = 16

    ws_det.column_dimensions["A"].width = 5
    ws_det.column_dimensions["B"].width = 30
    ws_det.column_dimensions["C"].width = 34
    ws_det.column_dimensions["D"].width = 7
    ws_det.column_dimensions["E"].width = 16
    ws_det.column_dimensions["F"].width = 16
    ws_det.column_dimensions["G"].width = 28
    ws_det.column_dimensions["H"].width = 40
    ws_det.column_dimensions["I"].width = 55

    # ════════════════════════════════════════════════════════════════════
    # Sheet 3: Weekends only (quick-check sheet)
    # ════════════════════════════════════════════════════════════════════
    ws_wk = wb.create_sheet("Weekend Errors")
    ws_wk.sheet_view.showGridLines = False
    ws_wk.freeze_panes = "A3"

    ws_wk.merge_cells("A1:G1")
    ws_wk["A1"] = "*** WEEKEND ERRORS — Please verify before importing ***"
    ws_wk["A1"].font = Font(bold=True, size=13, color=C_WHITE)
    ws_wk["A1"].fill = fill(C_WEEKEND_HEADER)
    ws_wk["A1"].alignment = center
    ws_wk.row_dimensions[1].height = 24

    WK_COLS = ["Row #", "Weekday", "Start Date", "Parent Legacy ID",
               "Child Legacy ID", "Parent Name", "Service Name"]
    ws_wk.append(WK_COLS)
    for cell in ws_wk[2]:
        cell.font = bold_font(C_WHITE)
        cell.fill = fill(C_WEEKEND_HEADER)
        cell.alignment = center
        cell.border = border
    ws_wk.row_dimensions[2].height = 20

    weekend_items = errors.get("weekend", [])
    if weekend_items:
        for i, e in enumerate(weekend_items):
            ws_wk.append([
                e.get("row", ""), e.get("weekday", ""), e.get("date", ""),
                e.get("parent_id", ""), e.get("child_id", ""),
                e.get("parent_name", ""), e.get("service", ""),
            ])
            r = ws_wk.max_row
            row_fill = fill(C_WEEKEND_ROW) if i % 2 == 0 else fill("F5B7B1")
            for col_idx, cell in enumerate(ws_wk[r], 1):
                cell.fill = row_fill
                cell.border = border
                cell.alignment = center if col_idx <= 3 else left
                cell.font = std_font()
                if col_idx == 2:
                    cell.font = bold_font("C0392B")
            ws_wk.row_dimensions[r].height = 16
    else:
        ws_wk.merge_cells("A3:G3")
        ws_wk["A3"] = "✓ No Weekend Errors found"
        ws_wk["A3"].font = Font(bold=True, size=11, color="1E8449")
        ws_wk["A3"].alignment = center
        ws_wk["A3"].fill = fill("D5F5E3")

    ws_wk.column_dimensions["A"].width = 7
    ws_wk.column_dimensions["B"].width = 10
    ws_wk.column_dimensions["C"].width = 14
    ws_wk.column_dimensions["D"].width = 16
    ws_wk.column_dimensions["E"].width = 16
    ws_wk.column_dimensions["F"].width = 30
    ws_wk.column_dimensions["G"].width = 45

    wb.save(out)
    return str(out)


# ─────────────────────────── GUI ────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Payment Plan Import Checker  v2")
        self.resizable(True, True)
        self.minsize(820, 640)
        self.configure(bg="#F5F5F5")

        self.col_vars = {k: tk.StringVar(value=v) for k, v in DEFAULT_COLUMNS.items()}
        self.filepath = tk.StringVar()

        self._build_ui()
        self._center_window(900, 750)

    # ── Layout ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        PAD = {"padx": 12, "pady": 5}

        # Title
        tf = tk.Frame(self, bg="#1565C0")
        tf.pack(fill="x")
        tk.Label(tf, text="  Payment Plan Import Checker",
                 font=("Segoe UI", 14, "bold"), fg="white", bg="#1565C0",
                 anchor="w", pady=10).pack(fill="x", padx=12)
        tk.Label(tf, text="  Validates against Onboarding Tool error rules from technical docs",
                 font=("Segoe UI", 9), fg="#90CAF9", bg="#1565C0",
                 anchor="w", pady=2).pack(fill="x", padx=12)

        # File selector
        ff = tk.LabelFrame(self, text=" 1. Select CSV File ", font=("Segoe UI", 10, "bold"),
                           bg="#F5F5F5", fg="#333")
        ff.pack(fill="x", **PAD)
        inner = tk.Frame(ff, bg="#F5F5F5")
        inner.pack(fill="x", padx=8, pady=6)
        tk.Entry(inner, textvariable=self.filepath, font=("Segoe UI", 10),
                 width=60, state="readonly").pack(side="left", fill="x", expand=True)
        tk.Button(inner, text="Browse…", command=self._browse,
                  bg="#1976D2", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=10, cursor="hand2").pack(side="left", padx=(8, 0))

        # Column mapping
        cf = tk.LabelFrame(self, text=" 2. Column Mapping  (edit if your CSV uses different column names) ",
                           font=("Segoe UI", 10, "bold"), bg="#F5F5F5", fg="#333")
        cf.pack(fill="x", **PAD)

        grid = tk.Frame(cf, bg="#F5F5F5")
        grid.pack(fill="x", padx=8, pady=4)

        keys = list(COLUMN_LABELS.keys())
        for idx, key in enumerate(keys):
            r, c = divmod(idx, 3)
            label = COLUMN_LABELS[key]
            fg = "#C62828" if "*required" in label else "#555"
            lbl_text = label.replace("  *required", " *")
            tk.Label(grid, text=lbl_text + ":", font=("Segoe UI", 8),
                     bg="#F5F5F5", fg=fg, anchor="e").grid(
                row=r, column=c * 2, sticky="e", padx=(8, 3), pady=2)
            tk.Entry(grid, textvariable=self.col_vars[key],
                     font=("Segoe UI", 8), width=24).grid(
                row=r, column=c * 2 + 1, sticky="ew", padx=(0, 10), pady=2)
        for c in range(3):
            grid.columnconfigure(c * 2 + 1, weight=1)

        # Run button
        bf = tk.Frame(self, bg="#F5F5F5")
        bf.pack(fill="x", **PAD)
        self.run_btn = tk.Button(
            bf, text="▶  Validate and Generate Files",
            command=self._run, bg="#2E7D32", fg="white",
            font=("Segoe UI", 12, "bold"), relief="flat",
            padx=20, pady=8, cursor="hand2", state="disabled",
        )
        self.run_btn.pack(side="left")
        self.status_lbl = tk.Label(bf, text="", font=("Segoe UI", 10),
                                   bg="#F5F5F5", fg="#666")
        self.status_lbl.pack(side="left", padx=16)

        # Results
        rf = tk.LabelFrame(self, text=" 3. Results ",
                           font=("Segoe UI", 10, "bold"), bg="#F5F5F5", fg="#333")
        rf.pack(fill="both", expand=True, **PAD)

        self.results = tk.Text(rf, font=("Consolas", 9), bg="#1E1E1E", fg="#D4D4D4",
                               insertbackground="white", wrap="none", relief="flat",
                               state="disabled")
        sb_y = tk.Scrollbar(rf, command=self.results.yview)
        sb_x = tk.Scrollbar(rf, orient="horizontal", command=self.results.xview)
        self.results.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        self.results.pack(fill="both", expand=True, padx=4, pady=4)

        self.results.tag_config("header", foreground="#569CD6", font=("Consolas", 9, "bold"))
        self.results.tag_config("ok", foreground="#4EC9B0")
        self.results.tag_config("warn", foreground="#DCDCAA")
        self.results.tag_config("error", foreground="#F44747", font=("Consolas", 9, "bold"))
        self.results.tag_config("weekend", foreground="#FF4500", font=("Consolas", 9, "bold"))
        self.results.tag_config("orange", foreground="#FF8C00")
        self.results.tag_config("dim", foreground="#858585")
        self.results.tag_config("path", foreground="#9CDCFE")

        tk.Label(self, text="Payment Plan Import Checker  •  Xplor Technologies",
                 font=("Segoe UI", 8), bg="#F5F5F5", fg="#AAAAAA").pack(pady=(0, 4))

    # ── Helpers ─────────────────────────────────────────────────────────────

    def _center_window(self, w, h):
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _browse(self):
        input_dir = Path(__file__).parent / "Input"
        initial_dir = str(input_dir) if input_dir.is_dir() else str(Path(__file__).parent)
        path = filedialog.askopenfilename(
            title="Select CSV File",
            initialdir=initial_dir,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if path:
            self.filepath.set(path)
            self.run_btn.config(state="normal")
            self._clear_results()
            self.status_lbl.config(text="")

    def _clear_results(self):
        self.results.config(state="normal")
        self.results.delete("1.0", "end")
        self.results.config(state="disabled")

    def _append(self, text: str, tag: str = ""):
        self.results.config(state="normal")
        self.results.insert("end", text, tag)
        self.results.config(state="disabled")

    # ── Main runner ──────────────────────────────────────────────────────────

    def _run(self):
        fp = self.filepath.get()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror("Error", "Please select a valid CSV file.")
            return

        col = {k: v.get().strip() for k, v in self.col_vars.items()}
        self.run_btn.config(state="disabled")
        self.status_lbl.config(text="Processing...")
        self.update()

        try:
            result = process_csv(fp, col)
        except Exception as exc:
            messagebox.showerror("Processing Error", str(exc))
            self.run_btn.config(state="normal")
            self.status_lbl.config(text="")
            return

        cleaned_path = write_cleaned_csv(fp, result["fieldnames"], result["processed_rows"])

        # Load service mapping — search order:
        #   1. Input/ folder next to script
        #   2. Same folder as the selected file
        #   3. Script folder itself
        service_map = {}
        for candidate in [Path(__file__).parent / "Input" / "serviceIDs.csv",
                          Path(fp).parent / "serviceIDs.csv",
                          Path(__file__).parent / "serviceIDs.csv"]:
            if candidate.exists():
                service_map = load_service_mapping(str(candidate))
                break
        if not service_map:
            messagebox.showwarning(
                "Service Mapping Not Found",
                "serviceIDs.csv was not found.\n"
                "Output will not be split by service — all rows go to 'unknown'.",
            )

        # Split first so we get row_map, then translate row numbers before writing report
        split_result = write_split_csvs(fp, col, result["processed_rows"], service_map)
        _translate_error_rows(result["errors"], split_result["row_map"])

        error_path = write_error_report(fp, result["errors"])

        self.run_btn.config(state="normal")
        self.status_lbl.config(text="Done ✓")
        self._display_results(result, cleaned_path, error_path, split_result)

    # ── Display ──────────────────────────────────────────────────────────────

    def _display_results(self, result: dict, cleaned_path: str, error_path: str,
                         split_result: dict = None):
        self._clear_results()
        stats = result["stats"]
        errors = result["errors"]

        def ln(text="", tag=""):
            self._append(text + "\n", tag)

        # ── Header ────────────────────────────────────────────────────────
        ln("═" * 76, "header")
        ln("  PAYMENT PLAN IMPORT CHECKER  —  Results", "header")
        ln(f"  {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}  |  {Path(self.filepath.get()).name}", "dim")
        ln("═" * 76, "header")
        ln()

        # ── Fix stats ─────────────────────────────────────────────────────
        ln("  FIXES APPLIED", "header")
        ln(f"  Total rows processed    : {stats['total']:,}", "ok")
        ln(f"  Date format fixed       : {stats['date_fixed']:,}", "ok")
        ln(f"  Weekday abbreviated     : {stats['weekday_fixed']:,}", "ok")
        ln(f"  Trailing spaces removed : {stats['spaces_fixed']:,}", "ok")
        ln()

        # ── Error summary (maps to Onboarding Tool error keys) ────────────
        DISPLAY = [
            ("weekend", "*** WEEKEND (Sat/Sun)            ", "weekend"),
            ("missing_date", "ERROR_MISSING_BOOKING_START_DATE ", "error"),
            ("missing_weekday", "ERROR_MISSING_PAYMENT_DAY        ", "error"),
            ("missing_parent", "ERROR_MISSING_GUARDIAN           ", "error"),
            ("missing_service_id", "ERROR_MISSING_SERVICE_ID         ", "error"),
            ("invalid_cycle", "ERROR_INVALID_FREQUENCY          ", "error"),
            ("manual_not_monday", "Manual Plan — Not Monday         ", "orange"),
            ("negative_limit", "ERROR_NEGATIVE_DIRECT_DEBIT_LIMIT", "error"),
            ("negative_fixed", "ERROR_NEGATIVE_FIXED_LIMIT       ", "error"),
            ("both_amounts", "ERROR_ONLY_ONE_AMOUNT_ALLOWED    ", "error"),
            ("unparseable_date", "Unparseable Date                 ", "warn"),
            ("unknown_weekday", "Unknown Weekday Value            ", "warn"),
        ]

        total_err = sum(len(errors.get(k, [])) for k, *_ in DISPLAY)

        ln("  ERROR SUMMARY  (matches Onboarding Tool error keys)", "header")
        for key, label, tag in DISPLAY:
            n = len(errors.get(key, []))
            t = tag if n else "ok"
            ln(f"  {label}: {n}", t)
        ln(f"  {'─' * 42}", "dim")
        ln(f"  {'Total errors':<43}: {total_err}", "error" if total_err else "ok")
        ln()

        # ── Detail sections ───────────────────────────────────────────────

        def section(key, title, tag, body_fn):
            items = errors.get(key, [])
            if not items:
                return
            ln("─" * 76, tag)
            ln(f"  {title}  ({len(items)} rows)", tag)
            ln("─" * 76, tag)
            for e in items:
                body_fn(e)
            ln()

        def row_line(e, extra=""):
            ln(f"  Row {e['row']:>5}  ParentID={e['parent_id']}  ChildID={e['child_id']}", "warn")
            ln(f"          {e['parent_name']}", "warn")
            ln(f"          Service : {e['service']}", "dim")
            if extra:
                ln(f"          {extra}", "error")

        # Weekend
        def _show_weekend(e):
            day_name = "Sunday" if e["weekday"] == "Sun" else "Saturday"
            ln(f"  Row {e['row']:>5}  [{e['weekday']}]  ParentID={e['parent_id']}  ChildID={e['child_id']}", "weekend")
            ln(f"          {e['parent_name']}", "weekend")
            ln(f"          Service : {e['service']}", "dim")
            ln(f"          Date    : {e['date']}  ← Falls on a {day_name}", "weekend")
            ln()

        section("weekend", "*** WEEKEND ERRORS — Must verify before importing! ***", "weekend", _show_weekend)

        # Missing Start Date
        section("missing_date", "ERROR_MISSING_BOOKING_START_DATE", "error",
                lambda e: (row_line(e, f"Weekday={e.get('weekday', '')}  Date=(empty)"), ln()))

        # Missing Weekday
        section("missing_weekday", "ERROR_MISSING_PAYMENT_DAY", "error",
                lambda e: (row_line(e, f"Date={e.get('date', '')}  Weekday=(empty)"), ln()))

        # Missing Parent
        section(
            "missing_parent",
            "ERROR_MISSING_GUARDIAN",
            "error",
            lambda e: (
                row_line(
                    e,
                    f"First:\"{
                        e.get(
                            'first_name',
                            '')}\"  Last:\"{
                        e.get(
                            'last_name',
                            '')}\""),
                ln()))

        # Missing Service ID
        section("missing_service_id", "ERROR_MISSING_SERVICE_ID", "error",
                lambda e: (row_line(e, "Service ID is empty"), ln()))

        # Invalid Cycle
        section("invalid_cycle", "ERROR_INVALID_FREQUENCY  (must be Weekly / Fortnightly / Monthly)", "error",
                lambda e: (row_line(e, f"Billing Cycle value: \"{e.get('value', '')}\""), ln()))

        # Manual not Monday
        section(
            "manual_not_monday",
            "Manual (Paused) Plan — Start Date Must Be Monday",
            "orange",
            lambda e: (
                row_line(
                    e,
                    f"Start Date {
                        e.get(
                            'date',
                            '')} is a {
                        e.get(
                            'day',
                            '')} — must be Monday"),
                ln()))

        # Negative Limit
        section("negative_limit", "ERROR_NEGATIVE_DIRECT_DEBIT_LIMIT", "error",
                lambda e: (row_line(e, f"Direct Debit Limit = {e.get('value', '')}"), ln()))

        # Negative Fixed
        section("negative_fixed", "ERROR_NEGATIVE_FIXED_LIMIT", "error",
                lambda e: (row_line(e, f"Fixed Amount = {e.get('value', '')}"), ln()))

        # Both amounts
        section(
            "both_amounts",
            "ERROR_ONLY_ONE_AMOUNT_ALLOWED",
            "error",
            lambda e: (
                row_line(
                    e,
                    f"Limit={
                        e.get(
                            'limit',
                            '')}  AND  Fixed={
                        e.get(
                            'fixed',
                            '')}  (only one allowed)"),
                ln()))

        # Unparseable date
        section("unparseable_date", "Unparseable Date Values", "warn",
                lambda e: (ln(f"  Row {e['row']:>5}  \"{e['value']}\"  — {e['parent_name']}", "warn"), ln()))

        # Unknown weekday
        section("unknown_weekday", "Unknown Weekday Values", "warn",
                lambda e: (ln(f"  Row {e['row']:>5}  \"{e['value']}\"  — {e['parent_name']}", "warn"), ln()))

        # All clear
        if total_err == 0:
            ln("  ✓ No errors found — ready to import!", "ok")
            ln()

        # Output files
        ln("─" * 76, "dim")
        ln("  OUTPUT FILES", "header")
        ln(f"  Cleaned CSV   : {cleaned_path}", "path")
        ln(f"  Error Report  : {error_path}", "path")

        if split_result:
            known = split_result.get("known", {})
            unknown = split_result.get("unknown", {})

            if known:
                ln()
                ln(f"  SPLIT FILES — by Service  ({len(known)} service(s))", "header")
                for svc_name, path in sorted(known.items()):
                    ln(f"  ✓  {svc_name}", "ok")
                    ln(f"       {path}", "path")

            if unknown:
                ln()
                ln(f"  UNKNOWN / UNMAPPED Service IDs  ({len(unknown)} group(s))", "warn")
                for sid, path in sorted(unknown.items()):
                    ln(f"  ⚠  Service ID: {sid}", "warn")
                    ln(f"       {path}", "path")

        ln("─" * 76, "dim")

        self.results.see("1.0")


# ─────────────────────────── Entry point ────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
