"""
process_balance_adjustments.py

Reads input data files (.xlsx, .xls, or .csv) from the 'input' folder,
maps columns to the Balance Adjustments Details Upload Template,
and writes one output .xlsx file per centre to the 'output' folder.

Column mapping:
  Centre_Name                <- Center Name  (must match Service Name in serviceIDs.csv)
  Account Holder First Name  <- Account Name, parsed from "Last Name, First Name"
  Account Holder Last Name   <- Account Name, parsed from "Last Name, First Name"
  Credit                     <- Credit column  OR  negative Amount Due (as absolute value)
  Owing                      <- Owing column   OR  positive Amount Due

Output file name: [Center Name]_Balance_Import.xlsx

Notes:
  - .xls files exported as HTML (e.g. from Xplor/QK) are handled automatically.
  - Only centres whose name exactly matches a Service Name in serviceIDs.csv get output files.
"""

import os
import sys
import copy
import csv
import re

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
TEMPLATE_PATH = os.path.join(BASE_DIR, "Balance Adjustments Details Upload Template.xlsx")
SERVICE_IDS_PATH = os.path.join(INPUT_DIR, "serviceIDs.csv")


def load_service_names(service_ids_path):
    """Return a set of valid Service Names from serviceIDs.csv."""
    service_names = set()
    with open(service_ids_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("Service Name", "").strip()
            if name:
                service_names.add(name)
    return service_names


def find_input_files(input_dir):
    """Return list of data files in input_dir, excluding serviceIDs.csv."""
    files = []
    for fname in sorted(os.listdir(input_dir)):
        if fname.lower() == "serviceids.csv":
            continue
        if fname.lower().endswith((".xlsx", ".xls", ".csv")):
            files.append(os.path.join(input_dir, fname))
    return files


def _is_html_file(filepath):
    """Peek at the first 512 bytes to detect HTML disguised as .xls."""
    with open(filepath, "rb") as f:
        header = f.read(512)
    return b"<" in header and b">" in header


def read_input_file(filepath):
    """
    Read an input file into a DataFrame with normalised column names.
    Handles true .xlsx/.xls, HTML-exported .xls, and .csv.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        df = pd.read_csv(filepath, dtype=str)
    elif ext in (".xlsx", ".xls") and _is_html_file(filepath):
        # HTML table exported with an Excel extension
        tables = pd.read_html(filepath)
        df = tables[0].astype(str)
    elif ext == ".xlsx":
        df = pd.read_excel(filepath, engine="openpyxl", dtype=str)
    else:
        # True binary .xls — requires xlrd
        df = pd.read_excel(filepath, engine="xlrd", dtype=str)

    df.columns = [c.strip() for c in df.columns]
    return df


def parse_amount(val):
    """
    Convert a currency string like '$1,234.56' or '-$51.67' to float.
    Returns None if the value is empty/NaN/zero.
    """
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "nan", "None"):
        return None
    # Remove currency symbol and thousands separator
    s = re.sub(r"[,$]", "", s)
    try:
        num = float(s)
    except ValueError:
        return None
    return num if num != 0.0 else None


def parse_name(account_name):
    """
    Split 'Last Name, First Name' into (first_name, last_name).
    Falls back gracefully when no comma is present.
    """
    if pd.isna(account_name) or str(account_name).strip() == "":
        return "", ""
    account_name = str(account_name).strip()
    if "," in account_name:
        last, _, first = account_name.partition(",")
        return first.strip(), last.strip()
    return account_name, ""


def copy_template_header(template_ws, target_ws):
    """Copy rows 1–4 (title + blank rows + header row) with styles to target_ws."""
    # Column widths
    for col_idx in range(1, template_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in template_ws.column_dimensions:
            target_ws.column_dimensions[col_letter].width = (
                template_ws.column_dimensions[col_letter].width
            )

    for row_num in range(1, 5):
        if row_num in template_ws.row_dimensions:
            target_ws.row_dimensions[row_num].height = (
                template_ws.row_dimensions[row_num].height
            )
        for col_idx in range(1, template_ws.max_column + 1):
            src = template_ws.cell(row=row_num, column=col_idx)
            dst = target_ws.cell(row=row_num, column=col_idx)
            dst.value = src.value
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.fill = copy.copy(src.fill)
                dst.border = copy.copy(src.border)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format

    # Replicate merged cells that fall within the header area
    for merge in template_ws.merged_cells.ranges:
        if merge.min_row <= 4:
            target_ws.merge_cells(str(merge))


def write_output(centre_name, rows_df, template_wb, output_dir):
    """Create one output .xlsx file for a single centre."""
    template_ws = template_wb.active

    # Capture data-row style from template row 5
    style_ref = {
        col_idx: template_ws.cell(row=5, column=col_idx)
        for col_idx in range(1, template_ws.max_column + 1)
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    copy_template_header(template_ws, ws)

    has_credit_col = "Credit" in rows_df.columns
    has_owing_col = "Owing" in rows_df.columns
    has_amount_due = "Amount Due" in rows_df.columns

    for data_idx, (_, row) in enumerate(rows_df.iterrows()):
        excel_row = 5 + data_idx
        first_name, last_name = parse_name(row.get("Account Name", ""))

        # Resolve Credit / Owing
        if has_credit_col or has_owing_col:
            credit = parse_amount(row.get("Credit"))
            owing = parse_amount(row.get("Owing"))
        elif has_amount_due:
            amount = parse_amount(row.get("Amount Due"))
            if amount is None:
                credit, owing = None, None
            elif amount < 0:
                credit, owing = abs(amount), None
            else:
                credit, owing = None, amount
        else:
            credit, owing = None, None

        data = [centre_name, first_name, last_name, credit, owing]

        for col_idx, value in enumerate(data, start=1):
            dst = ws.cell(row=excel_row, column=col_idx)
            dst.value = value
            src = style_ref[col_idx]
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.fill = copy.copy(src.fill)
                dst.border = copy.copy(src.border)
                dst.alignment = copy.copy(src.alignment)

    safe_name = "".join(c if c not in r'\/:*?"<>|' else "_" for c in centre_name)
    out_path = os.path.join(output_dir, f"{safe_name}_Balance_Import.xlsx")
    wb.save(out_path)
    return out_path


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    service_names = load_service_names(SERVICE_IDS_PATH)
    print(f"Loaded {len(service_names)} service names from serviceIDs.csv")

    input_files = find_input_files(INPUT_DIR)
    if not input_files:
        print("No input data files found in the input folder.")
        sys.exit(0)

    template_wb = load_workbook(TEMPLATE_PATH)

    total_outputs = 0
    total_rows = 0
    skipped_centres = set()

    for filepath in input_files:
        print(f"\nProcessing: {os.path.basename(filepath)}")
        try:
            df = read_input_file(filepath)
        except Exception as exc:
            print(f"  ERROR reading file: {exc}")
            continue

        # Find center name column (flexible naming)
        centre_col = next(
            (c for c in df.columns if c.lower().replace(" ", "") in ("centername", "centrename")),
            None,
        )
        if centre_col is None:
            print(f"  ERROR: No 'Center Name' column found. Columns: {list(df.columns)}")
            continue
        if "Account Name" not in df.columns:
            print(f"  ERROR: No 'Account Name' column found. Columns: {list(df.columns)}")
            continue

        df[centre_col] = df[centre_col].fillna("").str.strip()

        # Remove demo/test accounts
        before = len(df)
        df = df[~df["Account Name"].str.contains("demo parent", case=False, na=False)]
        removed = before - len(df)
        if removed:
            print(f"  Removed {removed} 'demo parent' row(s)")

        # Preserve original row order within each group
        seen_centres = []
        for centre_name in df[centre_col]:
            if centre_name and centre_name not in seen_centres:
                seen_centres.append(centre_name)

        for centre_name in seen_centres:
            if centre_name not in service_names:
                skipped_centres.add(centre_name)
                continue

            group = df[df[centre_col] == centre_name]
            out_path = write_output(centre_name, group, template_wb, OUTPUT_DIR)
            row_count = len(group)
            total_rows += row_count
            total_outputs += 1
            print(f"  Created: {os.path.basename(out_path)}  ({row_count} row{'s' if row_count != 1 else ''})")

    print(f"\nDone. {total_outputs} output file(s) created, {total_rows} data row(s) written.")
    if skipped_centres:
        print(f"Skipped {len(skipped_centres)} centre(s) not in serviceIDs.csv.")


if __name__ == "__main__":
    main()
