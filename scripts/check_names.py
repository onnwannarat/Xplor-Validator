import csv
import re
from collections import defaultdict
from pathlib import Path

BASE = Path("c:/Users/AmyBoonyaratanakornk/OneDrive - Xplor Technologies/Room and Fee Names Checking")

def read_csv(filename, encoding='utf-8-sig'):
    path = BASE / filename
    with open(path, encoding=encoding, errors='replace') as f:
        return list(csv.DictReader(f))

# Load service ID mapping: QKServiceID -> {xplor_service_id, service_name}
service_map = {}  # qk_service_id (str) -> {xplor_id, name}
for row in read_csv('serviceIDs.csv'):
    qk_id = row['QKServiceID'].strip()
    service_map[qk_id] = {
        'xplor_id': row['Xplor Service ID'].strip(),
        'name': row['Service Name'].strip()
    }

# Load Xplor fees: xplor_service_id -> set of fee names
xplor_fees = defaultdict(set)
for row in read_csv('9682_Active Fees_16.04.2026.csv'):
    sid = row['Service ID'].strip()
    fee = row['Fee Name'].strip()
    xplor_fees[sid].add(fee)

# Load Xplor rooms: service_name -> set of room names
xplor_rooms = defaultdict(set)
for row in read_csv('9682_Active Rooms_16.04.2026.csv'):
    centre = row['Centre_Name'].strip()
    room = row['Room_Name'].strip()
    xplor_rooms[centre].add(room)

# Load QK bookings (recurring + casual)
qk_rows = read_csv('4258_QikKidsOnboardingBookings_20260416.csv')
qk_rows += read_csv('Casual_QikKidsOnboardingBookings_fixed.csv')

# Collect unique (qk_service_id, fee_name) and (qk_service_id, room_name) pairs
qk_fees = defaultdict(set)   # qk_service_id -> set of fee names used
qk_rooms = defaultdict(set)  # qk_service_id -> set of room names used

for row in qk_rows:
    qk_id = row.get('Service Legacy ID', '').strip()
    fee = row.get('Fee Name', '').strip()
    room = row.get('Room Name', '').strip()
    if qk_id:
        if fee:
            qk_fees[qk_id].add(fee)
        if room:
            qk_rooms[qk_id].add(room)

# Build fee mismatch report
fee_mismatches = []
for qk_id, fees_in_qk in sorted(qk_fees.items()):
    if qk_id not in service_map:
        continue
    svc = service_map[qk_id]
    xplor_id = svc['xplor_id']
    svc_name = svc['name']
    xplor_fee_set = xplor_fees.get(xplor_id, set())
    for f in sorted(fees_in_qk):
        if f not in xplor_fee_set:
            # Find closest matches (fees containing similar words)
            words = set(re.split(r'\W+', f.lower())) - {'', 'the', 'a', 'an'}
            suggestions = [xf for xf in sorted(xplor_fee_set)
                           if any(w in xf.lower() for w in words if len(w) > 2)]
            fee_mismatches.append({
                'QK Service ID': qk_id,
                'Xplor Service ID': xplor_id,
                'Service Name': svc_name,
                'Fee Name in QK': f,
                'Possible Match in Xplor': '; '.join(suggestions) if suggestions else '(no match found)'
            })

# Build room mismatch report
room_mismatches = []
for qk_id, rooms_in_qk in sorted(qk_rooms.items()):
    if qk_id not in service_map:
        continue
    svc = service_map[qk_id]
    xplor_id = svc['xplor_id']
    svc_name = svc['name']
    xplor_room_set = xplor_rooms.get(svc_name, set())
    for r in sorted(rooms_in_qk):
        if r not in xplor_room_set:
            words = set(re.split(r'\W+', r.lower())) - {'', 'the', 'a', 'an'}
            suggestions = [xr for xr in sorted(xplor_room_set)
                           if any(w in xr.lower() for w in words if len(w) > 2)]
            room_mismatches.append({
                'QK Service ID': qk_id,
                'Xplor Service ID': xplor_id,
                'Service Name': svc_name,
                'Room Name in QK': r,
                'Possible Match in Xplor': '; '.join(suggestions) if suggestions else '(no match found)'
            })

# Write Excel report
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook()

    def write_sheet(wb, title, data, headers):
        ws = wb.active if title == 'Fee Mismatches' else wb.create_sheet(title)
        ws.title = title
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(bold=True, color='FFFFFF')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row_idx, row in enumerate(data, 2):
            for col, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=row.get(h, ''))
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        return ws

    fee_headers = ['QK Service ID', 'Xplor Service ID', 'Service Name', 'Fee Name in QK', 'Possible Match in Xplor']
    room_headers = ['QK Service ID', 'Xplor Service ID', 'Service Name', 'Room Name in QK', 'Possible Match in Xplor']
    write_sheet(wb, 'Fee Mismatches', fee_mismatches, fee_headers)
    write_sheet(wb, 'Room Mismatches', room_mismatches, room_headers)

    out = BASE / 'Fee_Room_Name_Mismatch_Report.xlsx'
    wb.save(out)
    print(f"Report saved: {out}")
    print(f"Fee mismatches: {len(fee_mismatches)}")
    print(f"Room mismatches: {len(room_mismatches)}")
except ImportError:
    print("openpyxl not found, writing CSV instead")
    import csv as csv_mod
    for name, data, hdrs in [
        ('Fee_Mismatches.csv', fee_mismatches, ['QK Service ID','Xplor Service ID','Service Name','Fee Name in QK','Possible Match in Xplor']),
        ('Room_Mismatches.csv', room_mismatches, ['QK Service ID','Xplor Service ID','Service Name','Room Name in QK','Possible Match in Xplor']),
    ]:
        with open(BASE / name, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv_mod.DictWriter(f, fieldnames=hdrs)
            w.writeheader()
            w.writerows(data)
        print(f"Saved {name}: {len(data)} rows")
