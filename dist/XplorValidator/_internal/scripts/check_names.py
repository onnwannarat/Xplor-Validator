"""
check_names.py
==============
Compares fee names and room names used in QikKids booking exports against the
names that exist in Xplor, and produces an Excel report with two sheets:
  • Fee Mismatches
  • Room Mismatches

Each mismatch row includes word-overlap-based suggestions for likely matches.

Public API
----------
run_check_names(service_ids_bytes, fees_bytes, rooms_bytes, bookings_bytes_list)
    -> bytes  (the Excel report as in-memory bytes)
"""

import csv
import io
import re
from collections import defaultdict


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _read_csv_bytes(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    return list(csv.DictReader(io.StringIO(text)))


def _build_service_map(service_ids_bytes: bytes) -> dict:
    """Return {qk_service_id: {xplor_id, name}}."""
    service_map = {}
    for row in _read_csv_bytes(service_ids_bytes):
        qk_id = row.get("QKServiceID", "").strip()
        if qk_id:
            service_map[qk_id] = {
                "xplor_id": row.get("Xplor Service ID", "").strip(),
                "name": row.get("Service Name", "").strip(),
            }
    return service_map


def _build_xplor_fees(fees_bytes: bytes) -> defaultdict:
    """Return {xplor_service_id: set of fee names}."""
    xplor_fees: defaultdict = defaultdict(set)
    for row in _read_csv_bytes(fees_bytes):
        sid = row.get("Service ID", "").strip()
        fee = row.get("Fee Name", "").strip()
        if sid and fee:
            xplor_fees[sid].add(fee)
    return xplor_fees


def _build_xplor_rooms(rooms_bytes: bytes) -> defaultdict:
    """Return {centre_name: set of room names}."""
    xplor_rooms: defaultdict = defaultdict(set)
    for row in _read_csv_bytes(rooms_bytes):
        centre = row.get("Centre_Name", "").strip()
        room = row.get("Room_Name", "").strip()
        if centre and room:
            xplor_rooms[centre].add(room)
    return xplor_rooms


def _load_qk_bookings(bookings_bytes_list: list[bytes]) -> list[dict]:
    rows: list[dict] = []
    for data in bookings_bytes_list:
        rows.extend(_read_csv_bytes(data))
    return rows


def _build_qk_usage(qk_rows: list[dict]) -> tuple[defaultdict, defaultdict]:
    """Return (qk_fees, qk_rooms) keyed by qk_service_id."""
    qk_fees: defaultdict = defaultdict(set)
    qk_rooms: defaultdict = defaultdict(set)
    for row in qk_rows:
        qk_id = row.get("Service Legacy ID", "").strip()
        if not qk_id:
            continue
        fee = row.get("Fee Name", "").strip()
        room = row.get("Room Name", "").strip()
        if fee:
            qk_fees[qk_id].add(fee)
        if room:
            qk_rooms[qk_id].add(room)
    return qk_fees, qk_rooms


def _word_overlap_suggestions(name: str, candidates: set[str]) -> list[str]:
    words = set(re.split(r"\W+", name.lower())) - {"", "the", "a", "an"}
    return [c for c in sorted(candidates) if any(w in c.lower() for w in words if len(w) > 2)]


def _build_fee_mismatches(
    service_map: dict,
    xplor_fees: defaultdict,
    qk_fees: defaultdict,
) -> list[dict]:
    mismatches = []
    for qk_id, fees_in_qk in sorted(qk_fees.items()):
        if qk_id not in service_map:
            continue
        svc = service_map[qk_id]
        xplor_id = svc["xplor_id"]
        svc_name = svc["name"]
        xplor_fee_set = xplor_fees.get(xplor_id, set())
        for f in sorted(fees_in_qk):
            if f not in xplor_fee_set:
                suggestions = _word_overlap_suggestions(f, xplor_fee_set)
                mismatches.append({
                    "QK Service ID": qk_id,
                    "Xplor Service ID": xplor_id,
                    "Service Name": svc_name,
                    "Fee Name in QK": f,
                    "Possible Match in Xplor": "; ".join(suggestions) if suggestions else "(no match found)",
                })
    return mismatches


def _build_room_mismatches(
    service_map: dict,
    xplor_rooms: defaultdict,
    qk_rooms: defaultdict,
) -> list[dict]:
    mismatches = []
    for qk_id, rooms_in_qk in sorted(qk_rooms.items()):
        if qk_id not in service_map:
            continue
        svc = service_map[qk_id]
        xplor_id = svc["xplor_id"]
        svc_name = svc["name"]
        xplor_room_set = xplor_rooms.get(svc_name, set())
        for r in sorted(rooms_in_qk):
            if r not in xplor_room_set:
                suggestions = _word_overlap_suggestions(r, xplor_room_set)
                mismatches.append({
                    "QK Service ID": qk_id,
                    "Xplor Service ID": xplor_id,
                    "Service Name": svc_name,
                    "Room Name in QK": r,
                    "Possible Match in Xplor": "; ".join(suggestions) if suggestions else "(no match found)",
                })
    return mismatches


def _write_excel(fee_mismatches: list[dict], room_mismatches: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()

    def write_sheet(wb, title, data, headers):
        ws = wb.active if title == "Fee Mismatches" else wb.create_sheet(title)
        ws.title = title
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(data, 2):
            for col, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=row.get(h, ""))
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    fee_headers = ["QK Service ID", "Xplor Service ID", "Service Name", "Fee Name in QK", "Possible Match in Xplor"]
    room_headers = ["QK Service ID", "Xplor Service ID", "Service Name", "Room Name in QK", "Possible Match in Xplor"]
    write_sheet(wb, "Fee Mismatches", fee_mismatches, fee_headers)
    write_sheet(wb, "Room Mismatches", room_mismatches, room_headers)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def run_check_names(
    service_ids_bytes: bytes,
    fees_bytes: bytes,
    rooms_bytes: bytes,
    bookings_bytes_list: list[bytes],
) -> tuple[bytes, int, int]:
    """Run the fee/room name check and return (excel_bytes, fee_count, room_count)."""
    service_map = _build_service_map(service_ids_bytes)
    xplor_fees = _build_xplor_fees(fees_bytes)
    xplor_rooms = _build_xplor_rooms(rooms_bytes)
    qk_rows = _load_qk_bookings(bookings_bytes_list)
    qk_fees, qk_rooms = _build_qk_usage(qk_rows)

    fee_mismatches = _build_fee_mismatches(service_map, xplor_fees, qk_fees)
    room_mismatches = _build_room_mismatches(service_map, xplor_rooms, qk_rooms)

    excel_bytes = _write_excel(fee_mismatches, room_mismatches)
    return excel_bytes, len(fee_mismatches), len(room_mismatches)


# ─────────────────────────────────────────────────────────────────────────────
# Standalone entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    from pathlib import Path

    if len(sys.argv) < 5:
        print("Usage: python check_names.py <serviceIDs.csv> <fees.csv> <rooms.csv> <bookings1.csv> [bookings2.csv ...]")
        sys.exit(1)

    svc_bytes = Path(sys.argv[1]).read_bytes()
    fees_bytes_ = Path(sys.argv[2]).read_bytes()
    rooms_bytes_ = Path(sys.argv[3]).read_bytes()
    bookings_list = [Path(p).read_bytes() for p in sys.argv[4:]]

    excel, n_fees, n_rooms = run_check_names(svc_bytes, fees_bytes_, rooms_bytes_, bookings_list)
    out = Path("Fee_Room_Name_Mismatch_Report.xlsx")
    out.write_bytes(excel)
    print(f"Report saved: {out}")
    print(f"Fee mismatches : {n_fees}")
    print(f"Room mismatches: {n_rooms}")
