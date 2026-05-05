# prepare_bookings_import.py
# Created by Amy B.
# Processes QikKids booking exports and prepares them for Xplor import.

import datetime
import glob
import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR        = os.path.join(SCRIPT_DIR, "Input")
OUTPUT_DIR       = os.path.join(SCRIPT_DIR, "Output")
SERVICE_IDS_FILE = os.path.join(INPUT_DIR, "serviceIDs.csv")
DEFAULT_END_DATE = "31/12/2026"
TODAY            = datetime.date.today().strftime("%Y%m%d")

# Final column order must match booking_onboarding_tools.csv exactly
TEMPLATE_COLUMNS = [
    "ServiceID", "Service_Name", "Child_Legacy_Id",
    "Child_First_Name", "Child_Last_Name",
    "StartDate", "EndDate",
    "ImportedFee", "ImportedRoom", "WeekType",
    "MON1", "TUE1", "WED1", "THU1", "FRI1", "SAT1", "SUN1",
    "MON2", "TUE2", "WED2", "THU2", "FRI2", "SAT2", "SUN2",
    "QKCreatedDate", "QKCreatedVia",
]

# Source column → Template column mapping
COLUMN_MAP = {
    "Service Legacy ID": "ServiceID",       # will be replaced with Xplor Service ID
    "Service Name":      "Service_Name",    # will be replaced with serviceIDs.csv name
    "Child Legacy ID":   "Child_Legacy_Id", # will be reformatted
    "Child First Name":  "Child_First_Name",
    "Child Last Name":   "Child_Last_Name",
    "Start Date":        "StartDate",
    "End Date":          "EndDate",
    "Fee Name":          "ImportedFee",
    "Room Name":         "ImportedRoom",
    "Frequency":         "WeekType",
    "Monday1":    "MON1",  "Tuesday1":   "TUE1",  "Wednesday1": "WED1",
    "Thursday1":  "THU1",  "Friday1":    "FRI1",  "Saturday1":  "SAT1",
    "Sunday1":    "SUN1",
    "Monday2":    "MON2",  "Tuesday2":   "TUE2",  "Wednesday2": "WED2",
    "Thursday2":  "THU2",  "Friday2":    "FRI2",  "Saturday2":  "SAT2",
    "Sunday2":    "SUN2",
    "Created Date": "QKCreatedDate",
    "Created via":  "QKCreatedVia",
}

# Weekday number (Monday=0) → day column prefix
WEEKDAY_TO_COL = {0: "MON", 1: "TUE", 2: "WED", 3: "THU", 4: "FRI", 5: "SAT", 6: "SUN"}


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling & duplicate-detection constants  ←── Raj's contribution
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="2F75B6")
HEADER_FONT   = Font(color="FFFFFF", bold=True)
WARN_FILL     = PatternFill("solid", fgColor="FFEB9C")   # yellow – used for removed/flagged rows
REMOVED_FILL  = PatternFill("solid", fgColor="F4CCCC")   # red-tint – overlap-removed casuals
GROUP_COLOURS = [
    "FFF2CC", "FCE4D6", "DDEBF7", "E2EFDA", "F4CCCC",
    "D9D2E9", "D0E0E3", "FFE599", "CFE2F3", "EAD1DC",
]
_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Key that identifies an exact-duplicate booking row  ←── Raj's contribution
DUPE_KEY = [
    "Service Legacy ID",
    "Child Legacy ID", "Child First Name", "Child Last Name",
    "Start Date", "End Date",
    "Fee Name", "Room Name", "Frequency",
    "Monday1", "Tuesday1", "Wednesday1", "Thursday1", "Friday1", "Saturday1", "Sunday1",
    "Monday2", "Tuesday2", "Wednesday2", "Thursday2", "Friday2", "Saturday2", "Sunday2",
]

# Source day-column → short label (used in duplicate report)  ←── Raj's contribution
_DAY_LABELS = {
    "Monday1": "Mon1", "Tuesday1": "Tue1", "Wednesday1": "Wed1", "Thursday1": "Thu1",
    "Friday1": "Fri1", "Saturday1": "Sat1", "Sunday1":   "Sun1",
    "Monday2": "Mon2", "Tuesday2": "Tue2", "Wednesday2": "Wed2", "Thursday2": "Thu2",
    "Friday2": "Fri2", "Saturday2": "Sat2", "Sunday2":   "Sun2",
}


# ─────────────────────────────────────────────────────────────────────────────
# Excel helper functions  ←── Raj's contribution
# ─────────────────────────────────────────────────────────────────────────────

def _style_header(ws) -> None:
    """Apply blue header styling to the first row of a worksheet."""
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER
    ws.row_dimensions[1].height = 30


def _add_notes_sheet(wb: Workbook, rows: list[tuple[str, str]]) -> None:
    """Add an 'About This Report' worksheet with label–description pairs.

    rows: list of (label, description) — pass ("", "") for a blank spacer row.
    """
    ws = wb.create_sheet("About This Report")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    LABEL_FONT = Font(bold=True, color="1F497D")
    NOTE_FILL  = PatternFill("solid", fgColor="F2F2F2")

    for label, text in rows:
        ws.append([label, text])
        r = ws.max_row
        ws.cell(r, 1).font      = LABEL_FONT
        ws.cell(r, 1).fill      = NOTE_FILL
        ws.cell(r, 2).fill      = NOTE_FILL
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, min(15 * ((len(text) // 90) + 1), 60))


def _auto_width(ws, max_width: int = 40) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=8,
        )
        ws.column_dimensions[letter].width = min(best + 3, max_width)


def _write_row(ws, values: list, fill=None, bold: bool = False) -> None:
    """Append a styled row to a worksheet."""
    ws.append(values)
    for cell in ws[ws.max_row]:
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True)
        cell.border = _BORDER


def _add_schedule_overlap_sheet(wb: Workbook, df_conflicts: pd.DataFrame) -> None:
    """Add a 'Recurring Schedule Overlaps' sheet to an existing workbook.

    Also appends a summary section to the existing 'About This Report' sheet.
    Both conflicting rows in each pair are marked REMOVE — neither is kept.
    """
    ws = wb.create_sheet("Recurring Schedule Overlaps")
    report_cols = TEMPLATE_COLUMNS + ["ConflictReason"]
    ws.append(report_cols)
    _style_header(ws)

    if df_conflicts.empty:
        ws.append(["No recurring schedule overlaps found."] + [""] * (len(report_cols) - 1))
    else:
        df_sorted  = df_conflicts.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)
        group_num  = 0
        prev_key: tuple | None = None
        for _, row in df_sorted.iterrows():
            key = (str(row.get("ServiceID", "")), str(row.get("Child_Legacy_Id", "")))
            if key != prev_key:
                if prev_key is not None:
                    ws.append([""] * len(report_cols))   # blank separator between groups
                group_num += 1
                prev_key = key
            fill   = PatternFill("solid", fgColor=GROUP_COLOURS[(group_num - 1) % len(GROUP_COLOURS)])
            values = [row.get(c, "") for c in report_cols]
            _write_row(ws, values, fill=fill)

    _auto_width(ws)
    ws.freeze_panes = "A2"

    # Append notes to the existing 'About This Report' sheet
    if "About This Report" in wb.sheetnames:
        notes_ws   = wb["About This Report"]
        LABEL_FONT = Font(bold=True, color="1F497D")
        NOTE_FILL  = PatternFill("solid", fgColor="F2F2F2")
        n_groups   = (
            df_conflicts.groupby(["ServiceID", "Child_Legacy_Id"]).ngroups
            if not df_conflicts.empty else 0
        )
        extra_rows = [
            ("", ""),
            ("RECURRING SCHEDULE OVERLAPS", ""),
            ("", ""),
            ("What is a recurring schedule overlap?",
             "Two (or more) recurring booking rows for the same child and service "
             "whose date ranges overlap AND which share at least one common booked weekday. "
             "This means the child would be double-booked in the same slot."),
            ("Action taken",
             "ALL conflicting rows have been removed from the Recurring import files. "
             "They appear in the 'Recurring Schedule Overlaps' sheet of this workbook."),
            ("If removal looks incorrect",
             "Verify the date ranges and day patterns in the source data. "
             "Correct the source and re-run the script."),
            ("", ""),
            ("Total conflict groups found", str(n_groups)),
            ("Total rows removed",          str(len(df_conflicts))),
        ]
        for label, text in extra_rows:
            notes_ws.append([label, text])
            r = notes_ws.max_row
            notes_ws.cell(r, 1).font      = LABEL_FONT
            notes_ws.cell(r, 1).fill      = NOTE_FILL
            notes_ws.cell(r, 2).fill      = NOTE_FILL
            notes_ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
            notes_ws.row_dimensions[r].height = max(15, min(15 * ((len(text) // 90) + 1), 60))


def _day_val(val) -> int | str:
    """Return 1 if the day cell is booked, blank string otherwise – no zeros."""
    try:
        return 1 if int(float(val)) == 1 else ""
    except (ValueError, TypeError):
        return ""


def _normalize_room(s: str) -> str:
    """Normalize room name for comparison, handling encoding variations.
    Different source files may represent the same non-ASCII characters
    differently (e.g. apostrophes as '?' vs 'â' encoding artifacts).
    Strips all non-alphanumeric characters and lowercases before comparing.
    """
    s = re.sub(r'[^\x00-\x7F]', ' ', s)   # remove non-ASCII (â artifacts)
    s = re.sub(r'[^a-zA-Z0-9 ]', ' ', s)  # remove ? - ( ) etc.
    return re.sub(r'\s+', ' ', s).strip().lower()


def _day_summary(row) -> str:
    """Human-readable booked-days string, e.g. 'Mon1, Wed1, Fri2'."""
    return ", ".join(
        lbl for col, lbl in _DAY_LABELS.items() if row.get(col, 0) == 1
    ) or "None"


# ─────────────────────────────────────────────
# Pipeline helpers
# ─────────────────────────────────────────────

def parse_date(series: pd.Series) -> pd.Series:
    """Parse a date series (with or without time) and return DD/MM/YYYY strings.
    Tries common formats used in QK exports before falling back to inference.
    """
    formats_to_try = [
        "%d/%m/%Y %I:%M:%S %p",  # 30/03/2026 12:00:00 AM
        "%d/%m/%Y %H:%M:%S",      # 30/03/2026 00:00:00
        "%d/%m/%Y",               # 30/03/2026
        "%Y-%m-%d %H:%M:%S",      # 2026-03-30 00:00:00
        "%Y-%m-%d",               # 2026-03-30
    ]
    result = pd.Series([pd.NaT] * len(series), index=series.index)
    remaining = series.copy()

    for fmt in formats_to_try:
        mask = result.isna() & remaining.str.strip().ne("")
        if not mask.any():
            break
        attempt = pd.to_datetime(remaining[mask], format=fmt, errors="coerce")
        result[mask] = attempt

    still_na = result.isna() & remaining.str.strip().ne("")
    if still_na.any():
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            fallback = pd.to_datetime(remaining[still_na], dayfirst=True, errors="coerce")
        result[still_na] = fallback

    return result.dt.strftime("%d/%m/%Y")


def load_service_mapping(path) -> dict:
    """Return {QKServiceID (str): (xplor_id (str), service_name (str))}.

    path may be a file path string or a file-like object (e.g. io.StringIO).
    """
    df = pd.read_csv(path, dtype=str).dropna(subset=["QKServiceID", "Xplor Service ID"])
    df["QKServiceID"]       = df["QKServiceID"].str.strip()
    df["Xplor Service ID"]  = df["Xplor Service ID"].str.strip()
    df["Service Name"]      = df["Service Name"].str.strip()
    return {
        row["QKServiceID"]: (row["Xplor Service ID"], row["Service Name"])
        for _, row in df.iterrows()
    }


def sanitize_filename(name: str) -> str:
    """Convert a service name to a safe filename component."""
    safe = re.sub(r"[^\w\s'-]", "", name)
    safe = re.sub(r"[\s']+", "_", safe.strip())
    safe = re.sub(r"_+", "_", safe)
    return safe.strip("_")


# ─────────────────────────────────────────────────────────────────────────────
# Duplicate detection & Excel report  ←── Raj's logic
# ─────────────────────────────────────────────────────────────────────────────

def detect_duplicates_and_report(
    df_raw: pd.DataFrame,
    out_dir: str,
) -> tuple[pd.DataFrame, int, int, Workbook, str]:
    """Detect exact-duplicate booking rows and build a styled Excel report.

    The workbook is returned (not saved) so that additional sheets can be
    appended before the final wb.save() call in main().

    Uses Raj's DUPE_KEY and colour-coded grouping logic.
    Duplicates are identified on the raw source columns (before transformation).

    Returns:
        df_clean      – de-duplicated raw DataFrame (first occurrence kept)
        n_dupe_rows   – total number of duplicate rows found
        n_dupe_groups – number of distinct duplicate groups
    """
    # Work on a copy; normalise day columns to int for reliable comparison
    df_cmp = df_raw.copy()
    df_cmp = df_cmp.map(lambda x: x.strip() if isinstance(x, str) else x)
    for c in _DAY_LABELS:
        if c in df_cmp.columns:
            df_cmp[c] = pd.to_numeric(df_cmp[c], errors="coerce").fillna(0).astype(int)
        else:
            df_cmp[c] = 0

    key = [c for c in DUPE_KEY if c in df_cmp.columns]

    # Fill NaN in key columns so duplicated() and groupby() treat them consistently
    df_cmp[key] = df_cmp[key].fillna("")

    is_dupe   = df_cmp.duplicated(subset=key, keep=False)
    dupe_rank = df_cmp.groupby(key).cumcount()

    # Rows to keep: non-dupes + first occurrence of each dupe group
    keep_mask = ~is_dupe | (dupe_rank == 0)
    df_clean  = df_raw[keep_mask].reset_index(drop=True)

    dupe_df       = df_cmp[is_dupe].copy()
    dupe_df["_dupe_rank"] = dupe_rank[is_dupe].values
    n_dupe_rows   = len(dupe_df)
    n_dupe_groups = dupe_df.groupby(key).ngroups if n_dupe_rows else 0

    # ── Write Excel report ───────────────────────────────────────────────────
    orig_cols   = [c for c in df_cmp.columns if not c.startswith("_")]
    report_path = os.path.join(out_dir, f"duplicate_bookings_report_{TODAY}.xlsx")

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Duplicate Booking Patterns"

    header = orig_cols + ["Booked Days Summary", "Duplicate Group", "Status"]
    ws.append(header)
    _style_header(ws)

    if n_dupe_rows == 0:
        ws.append(["No duplicate booking patterns found."] + [""] * (len(header) - 1))
    else:
        # Assign sequential group numbers
        group_map: dict = {}
        g = 1
        for _, grp in dupe_df.groupby(key, sort=False):
            for idx in grp.index:
                group_map[idx] = g
            g += 1

        dupe_df["_group_num"] = dupe_df.index.map(group_map)
        dupe_df_sorted = dupe_df.sort_values("_group_num")

        prev_group = None
        for _, row in dupe_df_sorted.iterrows():
            grp_num = int(row["_group_num"])
            fill    = PatternFill("solid", fgColor=GROUP_COLOURS[(grp_num - 1) % len(GROUP_COLOURS)])
            is_first = row["_dupe_rank"] == 0
            status   = "KEEP (first)" if is_first else "REMOVE (duplicate)"

            # Blank separator row between groups
            if prev_group is not None and grp_num != prev_group:
                ws.append([""] * len(header))

            values = [_day_val(row[c]) if c in _DAY_LABELS else (row.get(c, "") or "")
                      for c in orig_cols]
            values += [_day_summary(row), f"Group {grp_num}", status]
            _write_row(ws, values, fill=fill, bold=is_first)
            prev_group = grp_num

    _auto_width(ws)
    ws.freeze_panes = "A2"

    _add_notes_sheet(wb, [
        ("DUPLICATE BOOKING PATTERNS REPORT", ""),
        ("", ""),
        ("What is a duplicate?",
         "A booking row where ALL of the following fields are identical to another row: "
         "Service Legacy ID, Child Legacy ID, Child Name, Start Date, End Date, Fee Name, "
         "Room Name, Frequency, and Day Pattern (Monday–Sunday, Weeks 1 & 2)."),
        ("", ""),
        ("HOW TO READ THIS REPORT", ""),
        ("Colour coding",
         "Each colour group represents one set of duplicate rows. "
         "All rows sharing the same colour are exact duplicates of each other."),
        ("Bold row → KEEP (first)",
         "The first occurrence of the duplicate. This row IS included in the import file."),
        ("Normal rows → REMOVE (duplicate)",
         "All subsequent occurrences. These rows are NOT included in the import file."),
        ("Booked Days Summary",
         "Human-readable list of booked days, e.g. 'Mon1, Wed1, Fri2'. "
         "Week 1 = standard week; Week 2 = second week of a fortnightly cycle."),
        ("Duplicate Group",
         "Sequential group number. All rows with the same number are duplicates of each other."),
        ("Status",
         "KEEP (first) = included in import file.  |  REMOVE (duplicate) = excluded from import file."),
        ("", ""),
        ("ACTION REQUIRED", ""),
        ("If duplicates look correct",
         "No action needed — the script has already excluded the extra rows from all import files."),
        ("If a row should NOT be a duplicate",
         "The two rows may have genuinely different intended bookings (e.g. different dates). "
         "Correct the source data and re-run the script."),
        ("", ""),
        (f"Total duplicate groups found", str(n_dupe_groups)),
        (f"Total duplicate rows removed", str(max(0, n_dupe_rows - n_dupe_groups))),
    ])

    return df_clean, n_dupe_rows, n_dupe_groups, wb, report_path


# ─────────────────────────────────────────────
# Pipeline: transform raw → template format
# ─────────────────────────────────────────────

def process_df(df_raw: pd.DataFrame, service_map: dict) -> tuple[pd.DataFrame, set]:
    """Transform a raw (source-column) DataFrame into the template format.

    Returns (transformed_df, set_of_unmapped_qk_ids).
    """
    df = df_raw.copy().fillna("")
    df.rename(columns=COLUMN_MAP, inplace=True)
    # Drop duplicate column names that can arise when the source file already
    # contains a column matching a template name alongside the original name.
    df = df.loc[:, ~df.columns.duplicated(keep="first")]

    unmapped_ids: set = set()
    xplor_ids: list   = []
    service_names: list = []
    original_child_ids  = df["Child_Legacy_Id"].tolist()

    for _, row in df.iterrows():
        qk_id = str(row["ServiceID"]).strip()
        if qk_id in service_map:
            xplor_id, svc_name = service_map[qk_id]
            xplor_ids.append(xplor_id)
            service_names.append(svc_name)
        else:
            unmapped_ids.add(qk_id)
            xplor_ids.append("")
            service_names.append(row.get("Service_Name", ""))

    df["ServiceID"]     = xplor_ids
    df["Service_Name"]  = service_names

    # Format Child_Legacy_Id as {XplorServiceID}_{OriginalChildID}
    df["Child_Legacy_Id"] = [
        f"{xid}_{cid}" if xid else cid
        for xid, cid in zip(xplor_ids, original_child_ids)
    ]

    df["StartDate"]     = parse_date(df["StartDate"])
    df["EndDate"]       = parse_date(df["EndDate"])
    df["QKCreatedDate"] = parse_date(df["QKCreatedDate"])

    # Default empty EndDate → 31/12/2026
    df["EndDate"] = df["EndDate"].replace("NaT", "").apply(
        lambda v: DEFAULT_END_DATE if (pd.isna(v) or str(v).strip() in ("", "NaT", "nan")) else v
    )

    # Frequency: 'single' → 'CASUAL'
    df["WeekType"] = df["WeekType"].str.strip().apply(
        lambda v: "CASUAL" if v.lower() == "single" else v
    )

    for col in TEMPLATE_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    return df[TEMPLATE_COLUMNS].copy(), unmapped_ids


# ─────────────────────────────────────────────
# Casual / recurring overlap detection
# ─────────────────────────────────────────────

def detect_casual_overlaps(
    df_casual: pd.DataFrame,
    df_recurring: pd.DataFrame,
) -> list[tuple[bool, str]]:
    """Check every casual booking against all recurring bookings for the same child.

    A casual booking is flagged as an overlap when ALL of the following are true:
      1. Same Child_Legacy_Id
      2. Same ServiceID
      3. Same ImportedRoom
      4. The casual date falls within the recurring booking's StartDate – EndDate
      5. The casual date's weekday is a booked day in the recurring pattern
         (for Fortnightly: the correct Week 1 or Week 2 column; for Weekly and
         all other frequencies: the Week 1 columns only)

    Two overlap sub-types are distinguished in the reason string:
      • Same fee  – the casual and recurring fees match (direct double-charge risk)
      • Diff fee  – same room/day but a different fee (still occupying the same
                    slot, charged at a different rate)

    Returns a list of (is_overlap: bool, reason: str) — one entry per casual row.
    """
    results: list[tuple[bool, str]] = []

    # Pre-index recurring rows by child for speed.
    # Strip the key so that leading/trailing whitespace in the source data
    # does not prevent a match (Bug 3 fix).
    rec_by_child: dict[str, pd.DataFrame] = {}
    for child_id, grp in df_recurring.groupby("Child_Legacy_Id", sort=False):
        key = str(child_id).strip()
        if key in rec_by_child:
            rec_by_child[key] = pd.concat([rec_by_child[key], grp])
        else:
            rec_by_child[key] = grp

    for _, cas in df_casual.iterrows():
        child_id = str(cas["Child_Legacy_Id"]).strip()   # strip for Bug 3
        relevant = rec_by_child.get(child_id)
        if relevant is None:
            results.append((False, ""))
            continue

        # Bug 1 fix: use errors="coerce" so that None/blank StartDate values
        # (which parse_date can produce for unparseable dates) become NaT
        # instead of raising an exception.  In newer pandas, pd.to_datetime(None)
        # already returns NaT without raising, so the old try/except was silently
        # passing NaT through — NaT comparisons evaluate to False, causing the
        # entire casual row to be ignored.  The explicit pd.isna() guard makes
        # this failure visible and deterministic.
        cas_date = pd.to_datetime(cas["StartDate"], format="%d/%m/%Y", errors="coerce")
        if pd.isna(cas_date):
            results.append((False, ""))
            continue

        cas_service = str(cas.get("ServiceID",    "")).strip()
        cas_fee     = str(cas.get("ImportedFee",  "")).strip()
        cas_room    = str(cas.get("ImportedRoom", "")).strip()

        matched_reason = ""
        for _, rec in relevant.iterrows():
            # Bug 1 fix (cont.): same pattern for recurring dates.
            rec_start = pd.to_datetime(rec["StartDate"], format="%d/%m/%Y", errors="coerce")
            rec_end   = pd.to_datetime(rec["EndDate"],   format="%d/%m/%Y", errors="coerce")
            if pd.isna(rec_start) or pd.isna(rec_end):
                # Unparseable recurring date — skip this row but don't silently
                # fail the whole casual booking.
                continue

            # Condition 1–3: same service, same room, date within range
            if not (rec_start <= cas_date <= rec_end):
                continue
            if str(rec.get("ServiceID",    "")).strip() != cas_service:
                continue
            if _normalize_room(str(rec.get("ImportedRoom", ""))) != _normalize_room(cas_room):
                continue

            # Condition 4–5: the casual date's weekday must be a booked day in
            # the recurring pattern.
            #
            # FORTNIGHTLY: rec_start is Monday of Week 1.
            #   Week 1 = days 0–6 from rec_start, Week 2 = days 7–13, repeating
            #   every 14 days.  Check the column for the appropriate week suffix.
            # WEEKLY (and all other frequencies): full pattern is in Week 1
            #   columns (MON1–SUN1) only.
            cas_weekday  = cas_date.weekday()           # 0=Mon … 6=Sun
            day_prefix   = WEEKDAY_TO_COL[cas_weekday]  # e.g. "MON"
            week_type    = str(rec.get("WeekType", "")).strip().lower()

            if week_type == "fortnightly":
                delta_days    = (cas_date - rec_start).days
                week_in_cycle = (delta_days // 7) % 2   # 0=Week1, 1=Week2
                suffix        = "1" if week_in_cycle == 0 else "2"
            else:
                suffix = "1"

            # Bug 2 fix: use _day_val() instead of a raw string-equality check.
            # The source CSV may store booked days as "1.0" (float string) rather
            # than "1", causing str(...) == "1" to return False even when the day
            # IS booked.  _day_val() converts via float() first, so "1", "1.0",
            # and 1 all compare equal.
            col_to_check = day_prefix + suffix
            if _day_val(rec.get(col_to_check, "")) != 1:
                continue  # casual date's weekday not booked in this pattern

            # Build the shared part of the reason string (use _day_val here too)
            day_cols = [c for c in ["MON1","TUE1","WED1","THU1","FRI1","SAT1","SUN1",
                                    "MON2","TUE2","WED2","THU2","FRI2","SAT2","SUN2"]
                        if _day_val(rec.get(c, "")) == 1]
            days_str  = ", ".join(day_cols) if day_cols else "—"
            rec_fee   = str(rec.get("ImportedFee", "")).strip()
            period    = (
                f"Casual date {cas['StartDate']} ({day_prefix}, "
                f"Week {suffix} of {rec.get('WeekType','')}) falls within "
                f"recurring period {rec['StartDate']} – {rec['EndDate']} "
                f"(Days: {days_str}, Room: {cas_room}"
            )

            if rec_fee == cas_fee:
                matched_reason = period + f", Fee: {cas_fee})"
            else:
                matched_reason = (
                    period
                    + f") — same room, different fee: "
                    f"casual fee '{cas_fee}' vs recurring fee '{rec_fee}'"
                )
            break

        results.append((bool(matched_reason), matched_reason))

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Recurring schedule overlap detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_recurring_schedule_overlaps(
    df_recurring: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Find recurring bookings for the same child+service that overlap in both
    date range AND booked weekdays.

    Two rows conflict when ALL of the following are true:
      1. Same Child_Legacy_Id
      2. Same ServiceID
      3. Date ranges overlap  (start_A <= end_B  AND  start_B <= end_A)
      4. At least one shared booked day (MON1–SUN2 columns both = 1)

    BOTH rows in each conflicting pair are removed — neither is kept, because
    it is impossible to tell automatically which date range or day pattern is
    correct.

    Returns:
        df_clean     – df_recurring with all conflicting rows removed
        df_conflicts – the conflicting rows with an extra 'ConflictReason' column
    """
    DAY_COLS = [
        "MON1", "TUE1", "WED1", "THU1", "FRI1", "SAT1", "SUN1",
        "MON2", "TUE2", "WED2", "THU2", "FRI2", "SAT2", "SUN2",
    ]

    conflict_indices: set  = set()
    index_to_reason:  dict = {}

    for _, grp in df_recurring.groupby(
        ["Child_Legacy_Id", "ServiceID"], sort=False
    ):
        if len(grp) < 2:
            continue

        idx_list = list(grp.index)
        for i in range(len(idx_list)):
            idx_a = idx_list[i]
            row_a = grp.loc[idx_a]
            start_a = pd.to_datetime(row_a["StartDate"], format="%d/%m/%Y", errors="coerce")
            end_a   = pd.to_datetime(row_a["EndDate"],   format="%d/%m/%Y", errors="coerce")
            if pd.isna(start_a) or pd.isna(end_a):
                continue

            for j in range(i + 1, len(idx_list)):
                idx_b = idx_list[j]
                row_b = grp.loc[idx_b]
                start_b = pd.to_datetime(row_b["StartDate"], format="%d/%m/%Y", errors="coerce")
                end_b   = pd.to_datetime(row_b["EndDate"],   format="%d/%m/%Y", errors="coerce")
                if pd.isna(start_b) or pd.isna(end_b):
                    continue

                # Condition 3: date ranges must overlap
                if not (start_a <= end_b and start_b <= end_a):
                    continue

                # Condition 4: same room AND same fee
                if _normalize_room(str(row_a.get("ImportedRoom", ""))) != \
                   _normalize_room(str(row_b.get("ImportedRoom", ""))):
                    continue
                if str(row_a.get("ImportedFee", "")).strip().lower() != \
                   str(row_b.get("ImportedFee", "")).strip().lower():
                    continue

                # Condition 5: at least one shared booked weekday
                shared_days = [
                    c for c in DAY_COLS
                    if _day_val(row_a.get(c, "")) == 1 and _day_val(row_b.get(c, "")) == 1
                ]
                if not shared_days:
                    continue

                ov_start = max(start_a, start_b).strftime("%d/%m/%Y")
                ov_end   = min(end_a,   end_b  ).strftime("%d/%m/%Y")
                reason = (
                    f"Overlapping recurring schedules: "
                    f"date overlap {ov_start}–{ov_end}, "
                    f"shared booked days: {', '.join(shared_days)}"
                )
                conflict_indices.add(idx_a)
                conflict_indices.add(idx_b)
                index_to_reason.setdefault(idx_a, reason)
                index_to_reason.setdefault(idx_b, reason)

    df_conflicts = df_recurring.loc[sorted(conflict_indices)].copy()
    df_conflicts["ConflictReason"] = df_conflicts.index.map(index_to_reason)

    df_clean = df_recurring.drop(index=list(conflict_indices)).reset_index(drop=True)
    return df_clean, df_conflicts.reset_index(drop=True)


# ─────────────────────────────────────────────
# Service mapping — bytes variant
# ─────────────────────────────────────────────

def load_service_mapping_bytes(data: bytes) -> dict:
    """Like load_service_mapping() but reads from in-memory bytes."""
    import io as _io
    return load_service_mapping(_io.StringIO(data.decode("utf-8-sig", errors="replace")))


# ─────────────────────────────────────────────
# Save split CSVs by service (shared by both main variants)
# ─────────────────────────────────────────────

def save_by_service(
    df: pd.DataFrame,
    xplor_id_to_name: dict,
    target_dir: str,
    file_suffix: str,
) -> list[tuple[str, int]]:
    """Write per-service CSVs to target_dir and return [(filename, row_count)]."""
    files: list[tuple[str, int]] = []
    if df.empty:
        return files
    for xplor_id, group in df.groupby("ServiceID", sort=True):
        xplor_id_str = str(xplor_id).strip()
        if not xplor_id_str:
            filename = f"UNMAPPED_{file_suffix}.csv"
        else:
            svc_name  = xplor_id_to_name.get(xplor_id_str, xplor_id_str)
            safe_name = sanitize_filename(svc_name)
            filename  = f"{safe_name}_{file_suffix}.csv"
        group.to_csv(os.path.join(target_dir, filename), index=False, encoding="utf-8-sig")
        files.append((filename, len(group)))
    return files


# ─────────────────────────────────────────────
# MAIN (Streamlit-friendly: accepts bytes)
# ─────────────────────────────────────────────

def main(
    input_files: list[tuple[str, bytes]],
    service_ids_bytes: bytes,
    output_dir: str,
) -> dict:
    """Process booking files and write output CSVs + reports to output_dir.

    Parameters
    ----------
    input_files:
        List of (filename, file_bytes) for each booking CSV/XLSX.
    service_ids_bytes:
        Raw bytes of serviceIDs.csv.
    output_dir:
        Absolute path to the destination folder.  Sub-folders Recurring/ and
        Casual/ will be created inside it.

    Returns
    -------
    dict with keys:
        n_input_files, n_raw_rows, n_dupe_rows, n_dupe_groups,
        n_sched_conflict_rows, n_sched_conflict_groups,
        n_recurring, n_casual, n_casual_removed,
        unmapped_ids (set), recurring_files, casual_files,
        output_files (list of all saved paths)
    """
    import io as _io

    os.makedirs(output_dir, exist_ok=True)

    # 2. Load service mapping
    service_map      = load_service_mapping_bytes(service_ids_bytes)
    xplor_id_to_name = {v[0]: v[1] for v in service_map.values()}

    # 3. Load ALL files into one combined DataFrame (pre-transformation)
    RAW_COLS = list(COLUMN_MAP.keys())
    raw_frames: list[pd.DataFrame] = []
    for filename, file_bytes in input_files:
        ext = os.path.splitext(filename)[1].lower()
        if ext in (".xlsx", ".xls"):
            raw = pd.read_excel(_io.BytesIO(file_bytes), dtype=str)
        else:
            raw = pd.read_csv(_io.BytesIO(file_bytes), dtype=str)
        for col in RAW_COLS:
            if col not in raw.columns:
                raw[col] = ""
        raw_frames.append(raw[RAW_COLS])

    if not raw_frames:
        return {"error": "No readable rows found in the uploaded files."}

    df_raw_all = pd.concat(raw_frames, ignore_index=True).fillna("")

    # 4. Duplicate detection → duplicate_bookings_report.xlsx
    df_raw_clean, n_dupe_rows, n_dupe_groups, dup_wb, dup_report_path = \
        detect_duplicates_and_report(df_raw_all, output_dir)

    # 5. Transform de-duplicated raw data into template format
    df_all, all_unmapped = process_df(df_raw_clean, service_map)

    # 6. Separate recurring and casual bookings
    is_casual_mask = df_all["WeekType"].str.upper() == "CASUAL"
    df_recurring   = df_all[~is_casual_mask].reset_index(drop=True)
    df_casual      = df_all[is_casual_mask].reset_index(drop=True)

    # 6b. Detect recurring schedule overlaps
    df_recurring, df_recurring_conflicts = detect_recurring_schedule_overlaps(df_recurring)
    n_sched_conflict_rows   = len(df_recurring_conflicts)
    n_sched_conflict_groups = (
        df_recurring_conflicts.groupby(["ServiceID", "Child_Legacy_Id"]).ngroups
        if n_sched_conflict_rows else 0
    )

    _add_schedule_overlap_sheet(dup_wb, df_recurring_conflicts)
    dup_wb.save(dup_report_path)

    # 7. Remove casual bookings that overlap with recurring bookings
    removed_count = 0
    df_removed    = pd.DataFrame(columns=TEMPLATE_COLUMNS)
    if not df_recurring.empty and not df_casual.empty:
        overlap_results = detect_casual_overlaps(df_casual, df_recurring)
        overlap_mask    = pd.Series([r[0] for r in overlap_results], index=df_casual.index)
        reasons         = [r[1] for r in overlap_results]

        removed_count  = int(overlap_mask.sum())
        df_removed     = df_casual[overlap_mask].copy()
        df_removed["OverlapReason"] = [
            reason for reason, flag in zip(reasons, overlap_mask) if flag
        ]
        df_casual = df_casual[~overlap_mask].reset_index(drop=True)

    # 8. Sort
    df_recurring = df_recurring.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)
    df_casual    = df_casual.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)

    # 11. Create Output sub-folders
    recurring_dir = os.path.join(output_dir, "Recurring")
    casual_dir    = os.path.join(output_dir, "Casual")
    os.makedirs(recurring_dir, exist_ok=True)
    os.makedirs(casual_dir,    exist_ok=True)

    output_files: list[str] = [dup_report_path]

    # 12. Save removed-overlap report
    if not df_removed.empty:
        df_removed_sorted = df_removed.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)
        report_cols = TEMPLATE_COLUMNS + ["OverlapReason"]

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Removed Casual Bookings"
        ws.append(report_cols)
        _style_header(ws)

        for _, row in df_removed_sorted.iterrows():
            values = [row.get(c, "") for c in report_cols]
            _write_row(ws, values, fill=REMOVED_FILL)

        _auto_width(ws)
        ws.freeze_panes = "A2"

        _add_notes_sheet(wb, [
            ("REMOVED CASUAL BOOKINGS REPORT", ""),
            ("", ""),
            ("Why were these rows removed?",
             "Each casual booking in this report conflicts with an existing recurring booking "
             "for the same child. The casual booking falls on a day that is already covered "
             "by the recurring pattern (same room, same day-of-week, within the recurring "
             "date range). Uploading it would place two bookings in the same room slot."),
            ("", ""),
            ("OVERLAP CONDITIONS", ""),
            ("Conditions 1–5 must all be true",
             "1. Same child (Child_Legacy_Id)  "
             "2. Same service (ServiceID)  "
             "3. Same room (ImportedRoom)  "
             "4. Casual date falls within the recurring booking's Start–End Date range  "
             "5. Casual date's weekday is a booked day in the recurring pattern "
             "(for Fortnightly: checked against the correct Week 1 or Week 2 column; "
             "for Weekly and other frequencies: checked against Week 1 columns only)"),
            ("", ""),
            ("ACTION REQUIRED", ""),
            ("If removal looks correct",
             "No action needed. The row is already excluded from the Casual import files."),
            ("If removal looks incorrect",
             "The casual booking may be a genuinely different session. "
             "Verify with the service and, if needed, re-add it manually after the "
             "recurring import is complete."),
            ("", ""),
            ("Total casual bookings removed", str(len(df_removed_sorted))),
        ])

        overlap_report_path = os.path.join(output_dir, f"removed_overlap_report_{TODAY}.xlsx")
        wb.save(overlap_report_path)
        output_files.append(overlap_report_path)

    # 13. Save import CSVs split by service
    recurring_files = save_by_service(df_recurring, xplor_id_to_name, recurring_dir, f"bookings_import_{TODAY}")
    casual_files    = save_by_service(df_casual,    xplor_id_to_name, casual_dir,    f"casualbookings_import_{TODAY}")

    output_files += [os.path.join(recurring_dir, f) for f, _ in recurring_files]
    output_files += [os.path.join(casual_dir,    f) for f, _ in casual_files]

    return {
        "n_input_files":           len(input_files),
        "n_raw_rows":              len(df_raw_all),
        "n_dupe_rows":             n_dupe_rows,
        "n_dupe_groups":           n_dupe_groups,
        "n_sched_conflict_rows":   n_sched_conflict_rows,
        "n_sched_conflict_groups": n_sched_conflict_groups,
        "n_recurring":             len(df_recurring),
        "n_casual":                len(df_casual),
        "n_casual_removed":        removed_count,
        "unmapped_ids":            all_unmapped,
        "recurring_files":         recurring_files,
        "casual_files":            casual_files,
        "output_files":            output_files,
    }


# ─────────────────────────────────────────────
# CLI fallback (reads from the original Input/ folder structure)
# ─────────────────────────────────────────────

def _main_cli():
    # 1. Find all source files in the Input folder
    os.makedirs(INPUT_DIR, exist_ok=True)
    pattern = os.path.join(INPUT_DIR, "*.csv")
    matches = sorted(glob.glob(pattern))
    if not matches:
        print(f"ERROR: No CSV files found in:")
        print(f"       {INPUT_DIR}")
        return

    print(f"Input folder          : {INPUT_DIR}")
    print(f"Source files detected : {len(matches)}")
    for m in matches:
        print(f"  - {os.path.basename(m)}")

    # Build input_files list
    input_files = [(os.path.basename(m), open(m, "rb").read()) for m in matches]
    service_ids_bytes = open(SERVICE_IDS_FILE, "rb").read()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    result = main(input_files, service_ids_bytes, OUTPUT_DIR)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        return

    print(f"\nService mapping loaded : {len(result['unmapped_ids'])} unmapped IDs" if result["unmapped_ids"] else "")
    print()
    print("=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    print(f"  Total input files processed : {result['n_input_files']}")
    print(f"  Raw rows loaded             : {result['n_raw_rows']}")
    if result["n_dupe_rows"]:
        print(f"  Exact duplicates removed    : {result['n_dupe_rows']} ({result['n_dupe_groups']} groups)")
    if result["n_sched_conflict_rows"]:
        print(f"  Recurring sched. conflicts  : {result['n_sched_conflict_rows']} rows removed")
    print(f"  Recurring bookings          : {result['n_recurring']}")
    print(f"  Casual bookings kept        : {result['n_casual']}")
    if result["n_casual_removed"]:
        print(f"  Casual bookings removed     : {result['n_casual_removed']} (overlap with recurring)")
    if result["unmapped_ids"]:
        print()
        print("-" * 55)
        print("  WARNING: unmapped QK Service Legacy IDs:")
        for uid in sorted(result["unmapped_ids"]):
            print(f"    >> {uid}")
        print("-" * 55)
    print()
    print(f"  Output/Recurring/  ({len(result['recurring_files'])} files)")
    for fname, row_count in sorted(result["recurring_files"]):
        print(f"    [OK]  {fname}  ({row_count} rows)")
    print()
    print(f"  Output/Casual/  ({len(result['casual_files'])} files)")
    for fname, row_count in sorted(result["casual_files"]):
        print(f"    [OK]  {fname}  ({row_count} rows)")
    print("=" * 60)

    # 4. Duplicate detection (Raj's logic) → duplicate_bookings_report.xlsx
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("\nChecking for exact duplicate bookings …")
    df_raw_clean, n_dupe_rows, n_dupe_groups, dup_wb, dup_report_path = detect_duplicates_and_report(df_raw_all, OUTPUT_DIR)
    if n_dupe_rows:
        print(f"  Found {n_dupe_rows} duplicate rows across {n_dupe_groups} groups "
              f"→ duplicate_bookings_report_{TODAY}.xlsx")
    else:
        print("  No duplicates found.")

    # 5. Transform de-duplicated raw data into template format
    print("\nTransforming data …")
    df_all, all_unmapped = process_df(df_raw_clean, service_map)
    print(f"  {len(df_all)} rows after transformation")

    # 6. Separate recurring and casual bookings
    is_casual_mask = df_all["WeekType"].str.upper() == "CASUAL"
    df_recurring   = df_all[~is_casual_mask].reset_index(drop=True)
    df_casual      = df_all[is_casual_mask].reset_index(drop=True)

    print()
    print(f"Recurring bookings : {len(df_recurring)}")
    print(f"Casual bookings    : {len(df_casual)}")

    # 6b. Detect recurring schedule overlaps → remove BOTH conflicting rows
    n_sched_conflict_rows   = 0
    n_sched_conflict_groups = 0
    print("\nChecking for recurring schedule overlaps …")
    df_recurring, df_recurring_conflicts = detect_recurring_schedule_overlaps(df_recurring)
    n_sched_conflict_rows = len(df_recurring_conflicts)
    if n_sched_conflict_rows:
        n_sched_conflict_groups = df_recurring_conflicts.groupby(
            ["ServiceID", "Child_Legacy_Id"]
        ).ngroups
        print(f"  Removed {n_sched_conflict_rows} recurring booking(s) across "
              f"{n_sched_conflict_groups} conflict group(s)"
              f"  → duplicate_bookings_report_{TODAY}.xlsx  (sheet: Recurring Schedule Overlaps)")
    else:
        print("  No recurring schedule overlaps found.")

    _add_schedule_overlap_sheet(dup_wb, df_recurring_conflicts)
    dup_wb.save(dup_report_path)

    # 7. Remove casual bookings that overlap with recurring bookings
    #    Condition: same child + casual date within recurring period + same Fee + same Room
    removed_count = 0
    df_removed    = pd.DataFrame(columns=TEMPLATE_COLUMNS)
    if not df_recurring.empty and not df_casual.empty:
        print("\nChecking for casual/recurring overlaps …")
        overlap_results = detect_casual_overlaps(df_casual, df_recurring)
        overlap_mask    = pd.Series([r[0] for r in overlap_results], index=df_casual.index)
        reasons         = [r[1] for r in overlap_results]

        removed_count  = int(overlap_mask.sum())
        df_removed     = df_casual[overlap_mask].copy()
        df_removed["OverlapReason"] = [
            reason for reason, flag in zip(reasons, overlap_mask) if flag
        ]
        df_casual = df_casual[~overlap_mask].reset_index(drop=True)
        if removed_count:
            print(f"  Removed {removed_count} casual booking(s) that overlap with recurring bookings")
        else:
            print("  No overlaps found.")

    # 8. Sort
    df_recurring = df_recurring.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)
    df_casual    = df_casual.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)

    # 10. Warn about unmapped service IDs
    if all_unmapped:
        print()
        print("-" * 55)
        print("  WARNING: The following QK Service Legacy IDs could NOT")
        print("  be matched in serviceIDs.csv:")
        for uid in sorted(all_unmapped):
            print(f"    >> {uid}")
        print("  These rows have been included with a blank ServiceID.")
        print("-" * 55)

    # 11. Create Output sub-folders
    RECURRING_DIR = os.path.join(OUTPUT_DIR, "Recurring")
    CASUAL_DIR    = os.path.join(OUTPUT_DIR, "Casual")
    os.makedirs(RECURRING_DIR, exist_ok=True)
    os.makedirs(CASUAL_DIR,    exist_ok=True)

    # 12. Save removed-overlap report as styled Excel
    if not df_removed.empty:
        df_removed_sorted = df_removed.sort_values(["ServiceID", "Child_Legacy_Id"]).reset_index(drop=True)
        report_cols = TEMPLATE_COLUMNS + ["OverlapReason"]

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Removed Casual Bookings"
        ws.append(report_cols)
        _style_header(ws)

        for _, row in df_removed_sorted.iterrows():
            values = [row.get(c, "") for c in report_cols]
            _write_row(ws, values, fill=REMOVED_FILL)

        _auto_width(ws)
        ws.freeze_panes = "A2"

        _add_notes_sheet(wb, [
            ("REMOVED CASUAL BOOKINGS REPORT", ""),
            ("", ""),
            ("Why were these rows removed?",
             "Each casual booking in this report conflicts with an existing recurring booking "
             "for the same child. The casual booking falls on a day that is already covered "
             "by the recurring pattern (same room, same day-of-week, within the recurring "
             "date range). Uploading it would place two bookings in the same room slot."),
            ("", ""),
            ("OVERLAP CONDITIONS", ""),
            ("Conditions 1–5 must all be true",
             "1. Same child (Child_Legacy_Id)  "
             "2. Same service (ServiceID)  "
             "3. Same room (ImportedRoom)  "
             "4. Casual date falls within the recurring booking's Start–End Date range  "
             "5. Casual date's weekday is a booked day in the recurring pattern "
             "(for Fortnightly: checked against the correct Week 1 or Week 2 column; "
             "for Weekly and other frequencies: checked against Week 1 columns only)"),
            ("", ""),
            ("OVERLAP SUB-TYPES", ""),
            ("Same fee (direct double-charge)",
             "The casual and recurring bookings have the same fee. Uploading both would "
             "charge the parent twice at the same rate."),
            ("Same room, different fee",
             "The casual and recurring bookings share the same room and day but use "
             "different fee names. The slot is already occupied by the recurring booking; "
             "the casual fee would charge the parent an additional (usually higher) amount. "
             "The OverlapReason column identifies which sub-type applies to each row."),
            ("", ""),
            ("UPLOAD ORDER", ""),
            ("Step 1", "Upload ALL files in Output/Recurring/ first."),
            ("Step 2", "Upload ALL files in Output/Casual/ after Step 1 is complete."),
            ("Note",
             "The casual bookings in this report have already been excluded from the "
             "Output/Casual/ files — no further action is needed before uploading."),
            ("", ""),
            ("COLUMN GUIDE", ""),
            ("OverlapReason",
             "Describes which recurring booking caused the conflict: the casual date, "
             "day-of-week, week number (for Fortnightly), the recurring period "
             "(Start – End Date), booked days, Room, and Fee. For different-fee overlaps, "
             "both the casual and recurring fee names are shown."),
            ("", ""),
            ("ACTION REQUIRED", ""),
            ("If removal looks correct",
             "No action needed. The row is already excluded from the Casual import files."),
            ("If removal looks incorrect",
             "The casual booking may be a genuinely different session. "
             "Verify with the service and, if needed, re-add it manually after the "
             "recurring import is complete."),
            ("", ""),
            ("Total casual bookings removed", str(len(df_removed_sorted))),
        ])

        report_path = os.path.join(OUTPUT_DIR, f"removed_overlap_report_{TODAY}.xlsx")
        wb.save(report_path)
        print(f"\n  Overlap report saved : removed_overlap_report_{TODAY}.xlsx  ({len(df_removed_sorted)} rows)")

    # 13. Save import CSVs split by service
    def save_by_service(df: pd.DataFrame, target_dir: str, file_suffix: str) -> list[tuple[str, int]]:
        files: list[tuple[str, int]] = []
        if df.empty:
            return files
        for xplor_id, group in df.groupby("ServiceID", sort=True):
            xplor_id_str = str(xplor_id).strip()
            if not xplor_id_str:
                filename = f"UNMAPPED_{file_suffix}.csv"
            else:
                svc_name  = xplor_id_to_name.get(xplor_id_str, xplor_id_str)
                safe_name = sanitize_filename(svc_name)
                filename  = f"{safe_name}_{file_suffix}.csv"
            group.to_csv(os.path.join(target_dir, filename), index=False, encoding="utf-8-sig")
            files.append((filename, len(group)))
        return files

    recurring_files = save_by_service(df_recurring, RECURRING_DIR, f"bookings_import_{TODAY}")
    casual_files    = save_by_service(df_casual,    CASUAL_DIR,    f"casualbookings_import_{TODAY}")

    # 14. Summary
    print()
    print()
    print("=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    print(f"  Total input files processed : {len(matches)}")
    print(f"  Raw rows loaded             : {len(df_raw_all)}")
    if n_dupe_rows:
        print(f"  Exact duplicates removed    : {n_dupe_rows} ({n_dupe_groups} groups)"
              f"  → duplicate_bookings_report_{TODAY}.xlsx")
    if n_sched_conflict_rows:
        print(f"  Recurring sched. conflicts  : {n_sched_conflict_rows} rows removed "
              f"({n_sched_conflict_groups} groups)"
              f"  → duplicate_bookings_report_{TODAY}.xlsx")
    print(f"  Recurring bookings          : {len(df_recurring)}")
    print(f"  Casual bookings kept        : {len(df_casual)}")
    if removed_count:
        print(f"  Casual bookings removed     : {removed_count} (overlap with recurring)"
              f"  → removed_overlap_report_{TODAY}.xlsx")
    print()
    print(f"  Output/Recurring/  ({len(recurring_files)} files)")
    for fname, row_count in sorted(recurring_files):
        print(f"    [OK]  {fname}  ({row_count} rows)")
    print()
    print(f"  Output/Casual/  ({len(casual_files)} files)")
    for fname, row_count in sorted(casual_files):
        print(f"    [OK]  {fname}  ({row_count} rows)")
    print("=" * 60)


if __name__ == "__main__":
    _main_cli()
