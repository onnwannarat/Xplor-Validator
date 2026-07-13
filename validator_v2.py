#!/usr/bin/env python3
"""
Xplor Data Migration — CSV Validation & Transformation Tool  v2.0
==================================================================
Extends v1 (audit-only) with automatic transformation capabilities:

  • Service ID mapping        — maps QK legacy IDs to Xplor IDs via serviceIDs.csv
  • State normalisation       — converts free-text state names to standard abbreviations
  • In-row email deduplication — removes duplicate emails within Parent / Emergency Contact fields,
                                 including clearing EC emails that duplicate a parent email
  • Multi-service file split  — writes one ready_to_import_{Service_Name}.csv per service
  • Excel audit report (.xlsx) — one tab per service, includes an 'Action_Taken' column

Strict audit rules from v1 (mandatory fields, CRN format, duplicate detection,
waitlist logic, parent CRN/email registry, etc.) are all preserved unchanged.

Usage:
    py validator_v2.py <input_csv> [service_map_csv]

    input_csv        — the migration CSV prepared for Xplor import
    service_map_csv  — optional path to serviceIDs.csv
                       (defaults to serviceIDs.csv in the same folder as this script)

Output (written to the same folder as the input CSV):
    ready_to_import_{Service_Name}.csv  — one file per service (transformed data)
    validation_audit_report_v2.xlsx     — Excel report with one tab per service

Dependencies:
    pip install pandas openpyxl

Author: Amy Boonyaratanakornkit (Onboarding team)
"""

import csv
import io
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — Mandatory Fields
# Source: OFF-Families_ Data mapping & ON-Onboarding Tool - Parent Child Import
# ─────────────────────────────────────────────────────────────────────────────

MANDATORY_CHILD_FIELDS = [
    "ServiceID",
    "Child_Legacy_Id",
    "Child_First_Name",
    "Child_Last_Name",
    "DOB",
    "Status",
    "Child_CRN",
    "Room_Name",
    "Enrolment_Start_Date",
]

# Waitlist records are exempt from these fields — enrolment date is not yet confirmed.
MANDATORY_ACTIVE_ONLY_FIELDS = {"Enrolment_Start_Date"}

MANDATORY_PARENT1_FIELDS = [
    "Parent1_CRN",
    "Parent1_Legacy_Account_ID",
    "Parent1_DOB",
    "Parent1_Email",
]

MANDATORY_PARENT2_FIELDS = [
    "Parent2_CRN",
    "Parent2_Legacy_Account_ID",
    "Parent2_DOB",
    "Parent2_Email",
]

EMERGENCY_CONTACT_LEGACY_ID_FIELDS = [
    ("EmergencyContact1_LegacyID", "EmergencyContact1_First_Name"),
    ("EmergencyContact2_LegacyID", "EmergencyContact2_First_Name"),
    ("EmergencyContact3_LegacyID", "EmergencyContact3_First_Name"),
    ("EmergencyContact4_LegacyID", "EmergencyContact4_First_Name"),
    ("EmergencyContact5_LegacyID", "EmergencyContact5_First_Name"),
]

PAIRED_NAME_FIELDS = [
    ("Parent1_First_Name",           "Parent1_Last_Name"),
    ("Parent2_First_Name",           "Parent2_Last_Name"),
    ("EmergencyContact1_First_Name", "EmergencyContact1_Last_Name"),
    ("EmergencyContact2_First_Name", "EmergencyContact2_Last_Name"),
    ("EmergencyContact3_First_Name", "EmergencyContact3_Last_Name"),
    ("EmergencyContact4_First_Name", "EmergencyContact4_Last_Name"),
    ("EmergencyContact5_First_Name", "EmergencyContact5_Last_Name"),
]

# Accepted Status values — compared case-insensitively
VALID_STATUSES = {"active", "inactive", "waitlist"}

VALID_GENDERS = {"Male", "Female"}

STANDARD_GENDER_IDENTITIES = {"Male", "Female", "Non-Binary", "Trans Female", "Trans Male"}

# The Migration Tool auto-converts Yes/No and True/False to 1/0
VALID_BOOLEAN_VALUES = {"0", "1", "yes", "no", "true", "false"}

VALID_AU_STATES = {"NSW", "VIC", "QLD", "SA", "WA", "TAS", "ACT", "NT"}

DATE_FIELDS = ["DOB", "Enrolment_Start_Date", "Medicare_Expiry_Date", "Parent1_DOB", "Parent2_DOB"]

EMAIL_FIELDS = [
    "Parent1_Email", "Parent1_Work_Email",
    "Parent2_Email", "Parent2_Work_Email",
    "EmergencyContact1_Email", "EmergencyContact2_Email",
    "EmergencyContact3_Email", "EmergencyContact4_Email", "EmergencyContact5_Email",
]

PHONE_FIELDS = [
    "Parent1_Contact_Mobile", "Parent1_Contact_Home", "Parent1_Work_Phone",
    "Parent2_Contact_Mobile", "Parent2_Contact_Home", "Parent2_Work_Phone",
    "Medical_Practitioner_Phone",
    "EmergencyContact1_Contact_Number", "EmergencyContact2_Contact_Number",
    "EmergencyContact3_Contact_Number", "EmergencyContact4_Contact_Number",
    "EmergencyContact5_Contact_Number",
]

CRN_FIELDS = ["Child_CRN", "Parent1_CRN", "Parent2_CRN", "Enrolment_Parent_CRN"]

BOOLEAN_FIELDS = [
    "Consents_Photos_Videos", "Epipen/Anipen",
    "EmergencyContact1_Emergency_Contact", "EmergencyContact1_Medical_Nominee",
    "EmergencyContact1_Collection_Nominee", "EmergencyContact1_Excursion_Nominee",
    "EmergencyContact2_Emergency_Contact", "EmergencyContact2_Medical_Nominee",
    "EmergencyContact2_Collection_Nominee", "EmergencyContact2_Excursion_Nominee",
    "EmergencyContact3_Emergency_Contact", "EmergencyContact3_Medical_Nominee",
    "EmergencyContact3_Collection_Nominee", "EmergencyContact3_Excursion_Nominee",
    "EmergencyContact4_Emergency_Contact", "EmergencyContact4_Medical_Nominee",
    "EmergencyContact4_Collection_Nominee", "EmergencyContact4_Excursion_Nominee",
    "EmergencyContact5_Emergency_Contact", "EmergencyContact5_Medical_Nominee",
    "EmergencyContact5_Collection_Nominee", "EmergencyContact5_Excursion_Nominee",
]

STATE_FIELDS = [
    "State", "Parent1_State", "Parent1_Work_State",
    "Parent2_State", "Parent2_Work_State",
    "EmergencyContact1_State", "EmergencyContact2_State",
    "EmergencyContact3_State", "EmergencyContact4_State", "EmergencyContact5_State",
]

POSTCODE_FIELDS = [
    "PostCode",
    "Parent1_Post_Code", "Parent1_Work_Postcode",
    "Parent2_Post_Code", "Parent2_Work_Postcode",
    "EmergencyContact1_Postcode", "EmergencyContact2_Postcode",
    "EmergencyContact3_Postcode", "EmergencyContact4_Postcode", "EmergencyContact5_Postcode",
]

FIELD_LENGTH_LIMITS = {
    "Child_First_Name": 100, "Child_Middle_Name": 100, "Child_Last_Name": 100,
    "Gender": 100, "Gender_Identity": 100, "School": 200, "Class": 255,
    "Address": 255, "Suburb": 255, "Country": 255, "State": 255, "PostCode": 30,
    "Religion": 100, "Language": 255, "Cultural_Background": 255,
    "Cultural_Requirements": 255, "Indigenous_Status": 255,
    "Medicare_Number": 255, "Medicare_Expiry_Date": 255, "Ambulance_Cover_Number": 255,
    "Health_Care_Centre": 255, "Medical_Practitioner_Name": 255,
    "Medical_Practitioner_Phone": 255, "Medical_Practitioner_Address": 255,
    "Child_CRN": 255, "Child_Legacy_Id": 255,
    "Parent1_CRN": 255, "Parent1_Legacy_Account_ID": 100, "Parent1_Title": 100,
    "Parent1_First_Name": 100, "Parent1_Middle_Name": 100, "Parent1_Last_Name": 100,
    "Parent1_Email": 200, "Parent1_Contact_Home": 20, "Parent1_Gender": 50,
    "Parent1_Address_2": 100, "Parent1_Suburb": 255, "Parent1_State": 255,
    "Parent1_Post_Code": 255, "Parent1_Indigenous_Status": 255, "Parent1_Language": 255,
    "Parent1_Cultural_Background": 255, "Parent1_Work_Email": 255, "Parent1_Work_Phone": 255,
    "Parent1_Work_Address": 255, "Parent1_Work_Suburb": 255, "Parent1_Work_Postcode": 255,
    "Parent1_Work_Country": 255, "Parent1_Work_State": 255, "Parent1_Country": 255,
    "Parent2_CRN": 255, "Parent2_Legacy_Account_ID": 100,
    "Parent2_Email": 200, "Parent2_Work_Email": 255,
    "EmergencyContact1_Email": 60, "EmergencyContact2_Email": 60,
    "EmergencyContact3_Email": 60, "EmergencyContact4_Email": 60, "EmergencyContact5_Email": 60,
    "EmergencyContact1_Contact_Number": 30, "EmergencyContact2_Contact_Number": 30,
    "EmergencyContact3_Contact_Number": 30, "EmergencyContact4_Contact_Number": 30,
    "EmergencyContact5_Contact_Number": 30,
    "EmergencyContact1_First_Name": 100, "EmergencyContact1_Last_Name": 100,
    "EmergencyContact1_Postcode": 20,
    "EmergencyContact2_First_Name": 100, "EmergencyContact2_Last_Name": 100,
    "EmergencyContact2_Postcode": 20,
    "EmergencyContact3_First_Name": 100, "EmergencyContact3_Last_Name": 100,
    "EmergencyContact3_Postcode": 20,
    "EmergencyContact4_First_Name": 100, "EmergencyContact4_Last_Name": 100,
    "EmergencyContact4_Postcode": 20,
    "EmergencyContact5_First_Name": 100, "EmergencyContact5_Last_Name": 100,
    "EmergencyContact5_Postcode": 20,
}


# ─────────────────────────────────────────────────────────────────────────────
# STATE NORMALISATION MAP
# Maps common full names, abbreviations, and typo variants to standard codes.
# ─────────────────────────────────────────────────────────────────────────────

STATE_NORMALISATION_MAP: dict[str, str] = {
    # New South Wales
    "new south wales": "NSW", "new south wale": "NSW", "nsw": "NSW",
    "n.s.w": "NSW", "n.s.w.": "NSW",
    # Victoria
    "victoria": "VIC", "vic": "VIC", "v": "VIC",
    # Queensland
    "queensland": "QLD", "qld": "QLD", "queesland": "QLD", "queenslnd": "QLD",
    # South Australia
    "south australia": "SA", "sa": "SA", "sth australia": "SA",
    "south aust": "SA", "s.a": "SA", "s.a.": "SA",
    # Western Australia
    "western australia": "WA", "wa": "WA", "west australia": "WA",
    "w.a": "WA", "w.a.": "WA",
    # Tasmania
    "tasmania": "TAS", "tas": "TAS", "tassie": "TAS",
    # Australian Capital Territory
    "australian capital territory": "ACT", "act": "ACT",
    "a.c.t": "ACT", "a.c.t.": "ACT", "canberra": "ACT",
    # Northern Territory
    "northern territory": "NT", "nt": "NT",
    "n.t": "NT", "n.t.": "NT",
}

# Emergency contact email fields in priority order (EC1 takes precedence)
EC_EMAIL_FIELDS = [
    "EmergencyContact1_Email",
    "EmergencyContact2_Email",
    "EmergencyContact3_Email",
    "EmergencyContact4_Email",
    "EmergencyContact5_Email",
]

# Parent phone fields used for cross-service duplicate matching
PARENT_PHONE_FIELDS_BY_PREFIX = {
    "Parent1": ["Parent1_Contact_Mobile", "Parent1_Contact_Home", "Parent1_Work_Phone"],
    "Parent2": ["Parent2_Contact_Mobile", "Parent2_Contact_Home", "Parent2_Work_Phone"],
}

# Parent email fields used for cross-service duplicate matching
PARENT_EMAIL_FIELDS_BY_PREFIX = {
    "Parent1": ["Parent1_Email", "Parent1_Work_Email"],
    "Parent2": ["Parent2_Email", "Parent2_Work_Email"],
}


# ─────────────────────────────────────────────────────────────────────────────
# REGEX PATTERNS
# ─────────────────────────────────────────────────────────────────────────────

CRN_PATTERN    = re.compile(r"^\d{9}[A-Za-z]$")
PHONE_PATTERN  = re.compile(
    r"^(\+?61[\s\-]?|0)"
    r"([2378]\d[\s\-]?\d{4}[\s\-]?\d{4}|4\d{2}[\s\-]?\d{3}[\s\-]?\d{3})$"
)
EMAIL_PATTERN  = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
DATE_PATTERN_ISO = re.compile(r"^\d{4}-\d{2}-\d{2}$")
DATE_PATTERN_DMY = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
POSTCODE_PATTERN = re.compile(r"^\d{4}$")


# ─────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def is_blank(value) -> bool:
    """Returns True if the value is None, empty, or contains only whitespace."""
    return value is None or str(value).strip() == ""


def get_child_name(row: dict) -> str:
    """Constructs a display name for the child from the row data."""
    first = row.get("Child_First_Name", "").strip()
    last  = row.get("Child_Last_Name",  "").strip()
    return f"{first} {last}".strip() if (first or last) else "Unknown"


def is_valid_date(value: str) -> bool:
    """
    Returns True if the value is a valid date in YYYY-MM-DD or D/MM/YYYY format.
    Rejects impossible dates such as 29/02/2023.
    """
    v = value.strip()
    if DATE_PATTERN_ISO.match(v):
        fmt = "%Y-%m-%d"
    elif DATE_PATTERN_DMY.match(v):
        fmt = "%d/%m/%Y"
    else:
        return False
    try:
        datetime.strptime(v, fmt)
        return True
    except ValueError:
        return False


def parse_date(value: str) -> date:
    """Parses a date string in YYYY-MM-DD or D/MM/YYYY and returns a date object."""
    v = value.strip()
    if DATE_PATTERN_ISO.match(v):
        return datetime.strptime(v, "%Y-%m-%d").date()
    return datetime.strptime(v, "%d/%m/%Y").date()


def is_valid_crn(value: str) -> bool:
    return bool(CRN_PATTERN.match(value.strip()))


def is_valid_phone(value: str) -> bool:
    """
    Accepts standard Australian mobile/landline formats and 9-digit numbers
    beginning with '4' (the Migration Tool zero-pads these automatically).
    """
    cleaned = re.sub(r"[\s\-\(\)]", "", value)
    if re.match(r"^4\d{8}$", cleaned):
        return True
    return bool(PHONE_PATTERN.match(cleaned))


def is_valid_email(value: str) -> bool:
    return bool(EMAIL_PATTERN.match(value.strip()))


def normalise_key(col: str) -> str:
    return col.strip() if col else col


def safe_str(value) -> str:
    """Coerces a value to string, treating None as empty string.

    openpyxl reads integer-valued Excel cells as Python floats, so pandas
    produces "248992.0" instead of "248992".  Strip the spurious ".0" so IDs
    compare correctly when the same data arrives from both CSV and XLSX sources.
    """
    if value is None:
        return ""
    s = str(value)
    if s.endswith(".0") and len(s) > 2:
        core = s[:-2]
        if core.lstrip("-").isdigit():
            return core
    return s


def load_input_file(input_path: str) -> tuple[list[dict], list[str]]:
    """
    Reads a migration file in either CSV (.csv) or Excel (.xlsx / .xls) format
    and returns (rows, fieldnames).

    For XLSX files the first sheet is used, mirroring how the Bulk PC_Import
    template is structured.  All cell values are coerced to strings and None
    values are replaced with empty strings so that downstream logic is
    format-agnostic.

    Returns:
        rows       — list of dicts keyed by normalised column name
        fieldnames — ordered list of column names (preserving original order)
    """
    suffix = Path(input_path).suffix.lower()

    if suffix in (".xlsx", ".xls"):
        # Read the first sheet; keep_default_na=False prevents pandas from
        # silently converting empty cells to NaN strings.
        df = pd.read_excel(
            input_path,
            sheet_name=0,
            dtype=str,
            keep_default_na=False,
            engine="openpyxl" if suffix == ".xlsx" else None,
        )
        # Normalise column names
        df.columns = [normalise_key(str(c)) for c in df.columns]
        fieldnames = list(df.columns)
        rows: list[dict] = []
        for raw_row in df.to_dict(orient="records"):
            rows.append({k: safe_str(v) for k, v in raw_row.items()})
        return rows, fieldnames

    else:
        # CSV path — keep existing utf-8-sig handling
        rows = []
        fieldnames = []
        with open(input_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("The CSV file appears to be empty or has no header row.")
            reader.fieldnames = [normalise_key(h) for h in reader.fieldnames]
            fieldnames = reader.fieldnames[:]
            for raw_row in reader:
                rows.append({normalise_key(k): safe_str(v) for k, v in raw_row.items()})
        return rows, fieldnames


def load_input_bytes(raw: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """
    In-memory variant of load_input_file for Streamlit / bytes-based callers.
    Accepts the raw file bytes and the original filename (used to detect format).
    """
    suffix = Path(filename).suffix.lower()

    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(
            io.BytesIO(raw),
            sheet_name=0,
            dtype=str,
            keep_default_na=False,
            engine="openpyxl" if suffix == ".xlsx" else None,
        )
        df.columns = [normalise_key(str(c)) for c in df.columns]
        fieldnames = list(df.columns)
        rows = [{k: safe_str(v) for k, v in r.items()} for r in df.to_dict(orient="records")]
        return rows, fieldnames

    else:
        stream = io.StringIO(raw.decode("utf-8-sig"))
        reader = csv.DictReader(stream)
        if not reader.fieldnames:
            raise ValueError("The file appears to be empty or has no header row.")
        reader.fieldnames = [normalise_key(h) for h in reader.fieldnames]
        fieldnames = reader.fieldnames[:]
        rows = [{normalise_key(k): safe_str(v) for k, v in r.items()} for r in reader]
        return rows, fieldnames


# ─────────────────────────────────────────────────────────────────────────────
# EXISTING SERVICE — PARENT PROFILE LOADER
# Supports cross-service duplicate parent detection.
# ─────────────────────────────────────────────────────────────────────────────

def _extract_parent_profiles_from_rows(rows: list[dict], source_file: str) -> list[dict]:
    """
    Extracts a flat list of parent profile dicts from a list of row dicts.

    Each profile contains normalised (lowercase, stripped) identity fields so
    that comparisons against the input file are case-insensitive.
    """
    profiles: list[dict] = []
    for row in rows:
        for prefix in ("Parent1", "Parent2"):
            first_name = row.get(f"{prefix}_First_Name", "").strip().lower()
            last_name  = row.get(f"{prefix}_Last_Name",  "").strip().lower()
            if not first_name and not last_name:
                continue  # Skip empty parent slots

            dob = row.get(f"{prefix}_DOB", "").strip().lower()

            contacts: set[str] = set()
            for cf in PARENT_PHONE_FIELDS_BY_PREFIX.get(prefix, []):
                v = row.get(cf, "").strip()
                if v:
                    contacts.add(re.sub(r"[\s\-\(\)]", "", v).lower())

            emails: set[str] = set()
            for ef in PARENT_EMAIL_FIELDS_BY_PREFIX.get(prefix, []):
                v = row.get(ef, "").strip().lower()
                if v:
                    emails.add(v)

            profiles.append({
                "first_name":  first_name,
                "last_name":   last_name,
                "dob":         dob,
                "contacts":    contacts,
                "emails":      emails,
                "source_file": source_file,
                "service_id":  row.get("ServiceID", "").strip(),
                "legacy_id":   row.get(f"{prefix}_Legacy_Account_ID", "").strip().lower(),
                "parent_crn":  row.get(f"{prefix}_CRN", "").strip().lower(),
            })
    return profiles


def load_existing_parent_profiles_from_paths(file_paths: list[str]) -> list[dict]:
    """
    Loads parent profiles from one or more existing-service CSV/XLSX files on disk.
    Returns a flat list of profile dicts for use in cross-service duplicate checking.
    Logs a warning and continues if any individual file cannot be read.
    """
    all_profiles: list[dict] = []
    for path in file_paths:
        try:
            rows, _ = load_input_file(path)
            all_profiles.extend(_extract_parent_profiles_from_rows(rows, Path(path).name))
        except Exception as exc:
            print(f"  [WARNING] Could not load existing-service file '{path}': {exc}")
    return all_profiles


def load_existing_parent_profiles_from_bytes(
    files: list[tuple[bytes, str]],
) -> list[dict]:
    """
    In-memory variant of load_existing_parent_profiles_from_paths.
    Accepts a list of (file_bytes, filename) tuples — for Streamlit / bytes-based callers.
    """
    all_profiles: list[dict] = []
    for raw, filename in files:
        try:
            rows, _ = load_input_bytes(raw, filename)
            all_profiles.extend(_extract_parent_profiles_from_rows(rows, filename))
        except Exception as exc:
            print(f"  [WARNING] Could not parse existing-service file '{filename}': {exc}")
    return all_profiles


# ─────────────────────────────────────────────────────────────────────────────
# SERVICE MAPPING
# ─────────────────────────────────────────────────────────────────────────────

class ServiceMapping:
    """
    Loads serviceIDs.csv and provides lookups between QK legacy IDs,
    Xplor Service IDs, and Service Names.

    Supports the actual serviceIDs.csv column names:
      Service Name, Service Type, QKDBID, QKServiceID, Xplor Service ID
    """

    def __init__(self, csv_path: str):
        self._qk_to_xplor:   dict[str, str] = {}
        self._qk_to_name:    dict[str, str] = {}
        self._xplor_to_name: dict[str, str] = {}
        self._name_to_xplor: dict[str, str] = {}  # lowercase name -> xplor id
        self._loaded = False

        if not os.path.exists(csv_path):
            print(f"  [WARNING] serviceIDs.csv not found at: {csv_path}")
            print(f"            Service ID mapping will be skipped.")
            return

        try:
            with open(csv_path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Support both "Service Name" (actual) and "Service_Name" (legacy)
                    name  = safe_str(row.get("Service Name",    row.get("Service_Name",    ""))).strip()
                    xplor = safe_str(row.get("Xplor Service ID",row.get("Xplor_Service_ID",""))).strip()
                    # QK IDs: numeric short form and full QKDBID form
                    qk    = safe_str(row.get("QKServiceID",     row.get("QK_Service_ID",   ""))).strip()
                    qkdb  = safe_str(row.get("QKDBID",          "")).strip()

                    if xplor:
                        self._xplor_to_name[xplor] = name
                    if name and xplor:
                        self._name_to_xplor[name.lower()] = xplor
                    if qk and xplor:
                        self._qk_to_xplor[qk]   = xplor
                        self._qk_to_name[qk]     = name
                    if qkdb and xplor:
                        self._qk_to_xplor[qkdb] = xplor
                        self._qk_to_name[qkdb]   = name

            self._loaded = True
            print(f"  [OK] Service mapping loaded: {len(self._xplor_to_name)} service(s).")
        except Exception as exc:
            print(f"  [WARNING] Could not load serviceIDs.csv: {exc}")

    @property
    def is_loaded(self) -> bool:
        return self._loaded

    def lookup_by_qk(self, qk_id: str) -> tuple[str | None, str | None]:
        """
        Returns (xplor_service_id, service_name) for a given QK_Service_ID,
        or (None, None) if not found.
        """
        key = qk_id.strip()
        return self._qk_to_xplor.get(key), self._qk_to_name.get(key)

    def lookup_by_name(self, name: str) -> tuple[str, str] | tuple[None, None]:
        """
        Returns (xplor_service_id, canonical_service_name) for a given service name
        (case-insensitive), or (None, None) if not found.
        """
        key = name.strip().lower()
        xplor = self._name_to_xplor.get(key)
        if xplor:
            return xplor, self._xplor_to_name.get(xplor, "")
        return None, None

    def is_valid_xplor_id(self, xplor_id: str) -> bool:
        """Returns True if the given ID is a known Xplor Service ID."""
        return xplor_id.strip() in self._xplor_to_name

    def get_name_by_xplor(self, xplor_id: str) -> str:
        """Returns the Service Name for a given Xplor_Service_ID, or the ID itself."""
        return self._xplor_to_name.get(xplor_id.strip(), xplor_id.strip())


# ─────────────────────────────────────────────────────────────────────────────
# ISSUE / ACTION RECORDER
# ─────────────────────────────────────────────────────────────────────────────

class IssueRecorder:
    """
    Collects both validation issues and transformation actions across all rows.

    Severity levels:
      ERROR   — must be resolved before importing
      WARNING — review recommended before importing
      FIXED   — issue was detected and automatically corrected by the tool
    """

    REPORT_FIELDNAMES = [
        "Row", "Child_Name", "Field",
        "Issue_Description", "Severity_Level", "Action_Taken",
    ]

    def __init__(self):
        self.issues: list[dict] = []

    def add(
        self,
        row_num:     int,
        child_name:  str,
        field:       str,
        description: str,
        severity:    str,
        action:      str = "",
        tag:         str = "",
        **meta,
    ) -> None:
        """Records a single validation issue or transformation action.

        tag  — optional internal category used to filter specialist reports
               (e.g. 'duplicate_parent_email', 'redundant_ec',
               'cross_service_duplicate_parent').  Not written to main report columns.
        meta — optional arbitrary keyword arguments stored as '_key: value' pairs
               alongside the issue.  Used by specialist report writers to access
               structured data without parsing the description string.
               Example: source_file='abc.csv', existing_service_id='12345'
        """
        entry = {
            "Row":               row_num,
            "Child_Name":        child_name,
            "Field":             field,
            "Issue_Description": description,
            "Severity_Level":    severity,
            "Action_Taken":      action,
            "_tag":              tag,
        }
        entry.update({f"_{k}": v for k, v in meta.items()})
        self.issues.append(entry)

    def error_count(self) -> int:
        return sum(1 for i in self.issues if i["Severity_Level"] == "ERROR")

    def warning_count(self) -> int:
        return sum(1 for i in self.issues if i["Severity_Level"] == "WARNING")

    def fixed_count(self) -> int:
        return sum(1 for i in self.issues if i["Severity_Level"] == "FIXED")

    def to_dataframe(self) -> pd.DataFrame:
        if not self.issues:
            return pd.DataFrame(columns=self.REPORT_FIELDNAMES)
        return pd.DataFrame(self.issues, columns=self.REPORT_FIELDNAMES)

    def to_csv_bytes(self) -> bytes:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=self.REPORT_FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(self.issues)
        return buf.getvalue().encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# TRANSFORMATION FUNCTIONS
# These mutate the row dict in-place and log every change to the recorder.
# ─────────────────────────────────────────────────────────────────────────────

def transform_service_id(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder, service_map: ServiceMapping,
) -> None:
    """
    Resolves Xplor_Service_ID in this priority order:
      1. If the value is already a known Xplor Service ID → validate the
         Service_Name column (if present) and correct it if it doesn't match.
      2. If the value is a QK legacy ID (numeric or QKDBID form) → map it to
         the correct Xplor Service ID and update the row.
      3. If the value looks like a service name → look it up by name and
         replace with the correct Xplor Service ID.
      4. If not found in any direction → log a WARNING.

    After resolving the Xplor_Service_ID, if the row contains a Service_Name
    column that does not match the canonical name in serviceIDs.csv, the name
    is corrected automatically (FIXED).
    """
    if not service_map.is_loaded:
        return

    field = "ServiceID"
    value = row.get(field, "").strip()
    if is_blank(value):
        return  # Mandatory field check will catch this

    resolved_xplor_id: str | None = None
    canonical_name:    str | None = None

    # ── Step 1: Already a known Xplor Service ID? ────────────────────────────
    if service_map.is_valid_xplor_id(value):
        resolved_xplor_id = value
        canonical_name    = service_map.get_name_by_xplor(value)
        # No ID change needed — but fall through to name correction below.

    # ── Step 2: QK legacy ID lookup (numeric "182" or full "4258-182") ───────
    if resolved_xplor_id is None:
        xplor_id, svc_name = service_map.lookup_by_qk(value)
        if xplor_id is not None:
            original = value
            row[field] = xplor_id
            resolved_xplor_id = xplor_id
            canonical_name    = svc_name
            recorder.add(
                row_num, child_name, field,
                f"QK Service ID '{original}' mapped to Xplor Service ID '{xplor_id}' "
                f"({svc_name}).",
                "FIXED",
                action=f"Service ID updated: {original} → {xplor_id} ({svc_name})",
            )

    # ── Step 3: Name-based lookup ─────────────────────────────────────────────
    if resolved_xplor_id is None:
        xplor_by_name, svc_name = service_map.lookup_by_name(value)
        if xplor_by_name is not None:
            original = value
            row[field] = xplor_by_name
            resolved_xplor_id = xplor_by_name
            canonical_name    = svc_name
            recorder.add(
                row_num, child_name, field,
                f"Service name '{original}' resolved to Xplor Service ID '{xplor_by_name}' "
                f"({svc_name}).",
                "FIXED",
                action=f"Service ID updated from name: '{original}' → '{xplor_by_name}' ({svc_name})",
            )

    # ── Step 4: Not found anywhere ────────────────────────────────────────────
    if resolved_xplor_id is None:
        recorder.add(
            row_num, child_name, field,
            f"Service ID '{value}' was not found in the service mapping file "
            f"(checked as Xplor Service ID, QK Service ID, and Service Name). "
            f"Please verify the correct Xplor Service ID.",
            "WARNING",
        )
        return

    # ── Service_Name column correction ────────────────────────────────────────
    # If the input row has a Service_Name column, ensure it matches the canonical
    # name from serviceIDs.csv.  Fix it automatically if it differs.
    name_field = "Service_Name"
    if name_field in row and canonical_name:
        current_name = row.get(name_field, "").strip()
        if not is_blank(current_name) and current_name != canonical_name:
            row[name_field] = canonical_name
            recorder.add(
                row_num, child_name, name_field,
                f"Service_Name '{current_name}' does not match the canonical name "
                f"'{canonical_name}' for Xplor Service ID '{resolved_xplor_id}'.",
                "FIXED",
                action=f"Service_Name corrected: '{current_name}' → '{canonical_name}'",
            )


def normalise_state_value(raw: str, fallback: str = "") -> tuple[str, bool]:
    """
    Attempts to normalise a raw state string to a standard 2–3 letter abbreviation.

    Returns:
        (normalised_value, was_changed)
        If the value is already a valid abbreviation, returns it unchanged.
        If it matches a known variant, returns the normalised form.
        If unrecognisable and a fallback is provided, returns the fallback.
        If unrecognisable with no fallback, returns the original value unchanged.
    """
    stripped = raw.strip()
    if is_blank(stripped):
        return stripped, False

    # Already a valid abbreviation — no change needed
    if stripped.upper() in VALID_AU_STATES:
        return stripped.upper(), stripped.upper() != stripped

    # Try the normalisation map (case-insensitive)
    mapped = STATE_NORMALISATION_MAP.get(stripped.lower())
    if mapped:
        return mapped, True

    # Unrecognisable — use the service fallback if available
    if fallback:
        return fallback, True

    return stripped, False  # Cannot normalise; return as-is


def build_service_state_fallbacks(all_rows: list) -> dict[str, str]:
    """
    Builds a per-service fallback state by finding the most frequently occurring
    valid (already normalised) state value across all rows for each service.

    This is used for rows where the state value is entirely unrecognisable — those
    rows receive the modal state for their service.
    """
    # service_id -> Counter of valid state codes
    service_state_counts: defaultdict[str, Counter[str]] = defaultdict(Counter)

    for entry in all_rows:
        row       = entry["row"]
        svc_id    = row.get("ServiceID", "").strip()
        child_state = row.get("State", "").strip()

        if not is_blank(svc_id) and not is_blank(child_state):
            # Only count already-valid abbreviations
            if child_state.upper() in VALID_AU_STATES:
                service_state_counts[svc_id][child_state.upper()] += 1

    # For each service, pick the most common valid state
    fallbacks: dict[str, str] = {}
    for svc_id, counter in service_state_counts.items():
        if counter:
            fallbacks[svc_id] = counter.most_common(1)[0][0]

    return fallbacks


def transform_states(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder,
    service_fallbacks: dict[str, str],
) -> None:
    """
    Normalises all state fields in the row to standard Australian abbreviations.
    Uses the service-level fallback state for values that cannot be mapped.
    Logs a FIXED entry for every state that is changed.
    """
    svc_id   = row.get("ServiceID", "").strip()
    fallback = service_fallbacks.get(svc_id, "")

    for field in STATE_FIELDS:
        original = row.get(field, "").strip()
        if is_blank(original):
            continue

        normalised, changed = normalise_state_value(original, fallback)
        if changed:
            row[field] = normalised
            action_detail = f"State normalised: '{original}' → '{normalised}'"
            if normalised == fallback and original.upper() not in VALID_AU_STATES:
                action_detail += f" (using service modal state as fallback)"
            recorder.add(
                row_num, child_name, field,
                f"State value '{original}' normalised to '{normalised}'.",
                "FIXED",
                action=action_detail,
            )


def transform_email_dedup(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder, headers: set,
) -> None:
    """
    Removes duplicate email addresses within a single row, applying a strict
    priority order:

    Priority:
      Parent1_Email > Parent2_Email > EC1 > EC2 > EC3 > EC4 > EC5

    Rules:
      1. If Parent2_Email is identical to Parent1_Email, Parent2_Email is cleared.
      2. Within the Emergency Contact block (EC1–EC5), if a later contact's email
         matches any earlier contact's email, the later one is cleared.

    Note: Cross-row email deduplication is handled separately by the cross-row
    validation checks (check_duplicate_parent_emails).
    """
    # ── Rule 1: Parent2 vs Parent1 ───────────────────────────────────────────
    p1_email = row.get("Parent1_Email", "").strip().lower()
    p2_field = "Parent2_Email"
    if p2_field in headers:
        p2_email = row.get(p2_field, "").strip().lower()
        if p2_email and p1_email and p2_email == p1_email:
            original = row[p2_field]
            row[p2_field] = ""
            recorder.add(
                row_num, child_name, p2_field,
                f"Parent2_Email was identical to Parent1_Email ('{original}'). "
                f"Parent2_Email has been cleared to avoid a duplicate account conflict.",
                "FIXED",
                action=f"Duplicate email removed from Parent2_Email (same as Parent1_Email: '{original}')",
                tag="duplicate_parent_email",
            )

    # ── Rule 2: Emergency Contact email deduplication (EC1 wins) ─────────────
    seen_ec_emails: set[str] = set()

    for field in EC_EMAIL_FIELDS:
        if field not in headers:
            continue
        ec_email = row.get(field, "").strip().lower()
        if is_blank(ec_email):
            continue

        if ec_email in seen_ec_emails:
            original = row[field]
            row[field] = ""
            # Identify which earlier contact holds this email for a helpful message
            recorder.add(
                row_num, child_name, field,
                f"'{field}' email ('{original}') is a duplicate of an earlier emergency contact's "
                f"email. The value has been cleared — only the first occurrence is retained.",
                "FIXED",
                action=f"Duplicate email removed from '{field}': '{original}'",
            )
        else:
            seen_ec_emails.add(ec_email)


def transform_blank_first_names(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder, headers: set,
) -> None:
    """
    If any person's First_Name field is blank but their Last_Name is not,
    copies the Last_Name value into First_Name.

    Applies to: Child, Parent1, Parent2, EmergencyContact 1-5.
    """
    name_pairs = [
        ("Child_First_Name",           "Child_Last_Name"),
        ("Parent1_First_Name",         "Parent1_Last_Name"),
        ("Parent2_First_Name",         "Parent2_Last_Name"),
        ("EmergencyContact1_First_Name", "EmergencyContact1_Last_Name"),
        ("EmergencyContact2_First_Name", "EmergencyContact2_Last_Name"),
        ("EmergencyContact3_First_Name", "EmergencyContact3_Last_Name"),
        ("EmergencyContact4_First_Name", "EmergencyContact4_Last_Name"),
        ("EmergencyContact5_First_Name", "EmergencyContact5_Last_Name"),
    ]

    for first_field, last_field in name_pairs:
        if first_field not in headers or last_field not in headers:
            continue
        first_val = row.get(first_field, "").strip()
        last_val  = row.get(last_field,  "").strip()
        if is_blank(first_val) and not is_blank(last_val):
            row[first_field] = last_val
            recorder.add(
                row_num, child_name, first_field,
                f"'{first_field}' was blank while '{last_field}' was '{last_val}'. "
                f"First name has been set to the last name value.",
                "FIXED",
                action=f"Copied '{last_val}' from {last_field} into {first_field}",
            )


def transform_legacy_ids(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder, headers: set,
) -> None:
    """
    1. Prefixes Child_Legacy_Id with the Xplor Service ID (e.g. "1234_29085").
    2. Detects when Parent1 and Parent2 are the same person (matching first name,
       last name, DOB, AND Legacy Account ID) and makes P2's Legacy ID unique by
       appending '_1' (e.g. "38889_1").
    Parent legacy IDs are left as-is (no service prefix).
    All changes are logged as FIXED entries.
    """
    service_id = row.get("ServiceID", "").strip()
    if not service_id:
        return

    # ── 1. Child Legacy ID ────────────────────────────────────────────────────
    child_legacy_field = "Child_Legacy_Id"
    if child_legacy_field in headers:
        original = row.get(child_legacy_field, "").strip()
        if original and original != "0":
            new_val = f"{service_id}_{original}"
            row[child_legacy_field] = new_val
            recorder.add(
                row_num, child_name, child_legacy_field,
                f"Child Legacy ID prefixed with Xplor Service ID: '{original}' → '{new_val}'.",
                "FIXED",
                action=f"Child Legacy ID prefixed: '{original}' → '{new_val}'",
            )

    # ── 2. Detect P1 == P2 (same physical person) ────────────────────────────
    p1_first  = row.get("Parent1_First_Name",        "").strip().lower()
    p1_last   = row.get("Parent1_Last_Name",         "").strip().lower()
    p1_dob    = row.get("Parent1_DOB",               "").strip()
    p1_legacy = row.get("Parent1_Legacy_Account_ID", "").strip()

    p2_first  = row.get("Parent2_First_Name",        "").strip().lower()
    p2_last   = row.get("Parent2_Last_Name",         "").strip().lower()
    p2_dob    = row.get("Parent2_DOB",               "").strip()
    p2_legacy = row.get("Parent2_Legacy_Account_ID", "").strip()

    if (
        p1_first and p2_first
        and p1_first  == p2_first
        and p1_last   == p2_last
        and p1_dob    and p2_dob  and p1_dob   == p2_dob
        and p1_legacy and p2_legacy and p1_legacy == p2_legacy
    ):
        unique_p2 = f"{p2_legacy}_1"
        row["Parent2_Legacy_Account_ID"] = unique_p2
        recorder.add(
            row_num, child_name, "Parent2_Legacy_Account_ID",
            f"Parent1 and Parent2 appear to be the same person "
            f"(name, DOB, and Legacy ID all match: '{p2_legacy}'). "
            f"Parent2 Legacy ID made unique: '{p2_legacy}' → '{unique_p2}'.",
            "FIXED",
            action=f"Parent2 Legacy ID uniquified: '{p2_legacy}' → '{unique_p2}'",
        )



def transform_phone_leading_zero(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder, headers: set,
) -> None:
    """
    Prepends a leading '0' to any phone/contact-number field that consists
    entirely of digits (with optional spaces or hyphens as separators) but does
    not already start with '0' or '+' (international prefix).
    Logs a FIXED entry for every number corrected.
    """
    for field in PHONE_FIELDS:
        if field not in headers:
            continue
        original = row.get(field, "").strip()
        if is_blank(original):
            continue
        if original.startswith("+"):
            continue  # International format — leave as-is
        stripped = re.sub(r"[\s\-\(\)]", "", original)
        if stripped.isdigit() and not stripped.startswith("0"):
            new_val = "0" + original
            row[field] = new_val
            recorder.add(
                row_num, child_name, field,
                f"Phone number '{original}' was missing a leading zero. "
                f"Prepended '0': '{new_val}'.",
                "FIXED",
                action=f"Leading zero added: '{original}' → '{new_val}'",
            )


def transform_consents_photos(
    row: dict, row_num: int, child_name: str,
    recorder: IssueRecorder, headers: set,
) -> None:
    """
    Sets Consents_Photos (and Consents_Photos_Videos if present) to 'N' when
    the field is blank, rather than leaving it empty and generating a warning.
    """
    for field in ("Consents_Photos", "Consents_Photos_Videos"):
        if field not in headers:
            continue
        if is_blank(row.get(field, "")):
            row[field] = "N"
            recorder.add(
                row_num, child_name, field,
                f"'{field}' was blank — defaulted to 'N' (no photo consent).",
                "FIXED",
                action=f"'{field}' set to 'N' (was blank)",
            )


# ─────────────────────────────────────────────────────────────────────────────
# PER-ROW VALIDATION FUNCTIONS  (unchanged from v1)
# ─────────────────────────────────────────────────────────────────────────────

def validate_mandatory_child_fields(row, row_num, child_name, recorder, headers):
    """Checks mandatory child fields; Waitlist records are exempt from Enrolment_Start_Date."""
    is_waitlist = row.get("Status", "").strip().lower() == "waitlist"
    for field in MANDATORY_CHILD_FIELDS:
        if field not in headers:
            continue
        if is_waitlist and field in MANDATORY_ACTIVE_ONLY_FIELDS:
            continue
        if is_blank(row.get(field, "")):
            recorder.add(row_num, child_name, field,
                         f"Mandatory field '{field}' is missing or empty.", "ERROR")


def validate_mandatory_parent_fields(row, row_num, child_name, recorder, headers):
    """Checks mandatory parent fields when the parent block is populated."""
    def _block_active(indicator_fields):
        return any(not is_blank(row.get(f, "")) for f in indicator_fields if f in headers)

    if _block_active(["Parent1_First_Name", "Parent1_Last_Name", "Parent1_Email", "Parent1_CRN"]):
        for field in MANDATORY_PARENT1_FIELDS:
            if field in headers and is_blank(row.get(field, "")):
                recorder.add(row_num, child_name, field,
                             f"Mandatory Parent 1 field '{field}' is missing or empty.", "ERROR")

    if _block_active(["Parent2_First_Name", "Parent2_Last_Name", "Parent2_Email", "Parent2_CRN"]):
        for field in MANDATORY_PARENT2_FIELDS:
            if field in headers and is_blank(row.get(field, "")):
                recorder.add(row_num, child_name, field,
                             f"Mandatory Parent 2 field '{field}' is missing or empty.", "ERROR")


def validate_emergency_contact_legacy_ids(row, row_num, child_name, recorder, headers):
    """Each populated emergency contact block must include a Legacy ID."""
    for legacy_id_field, indicator_field in EMERGENCY_CONTACT_LEGACY_ID_FIELDS:
        if indicator_field not in headers:
            continue
        if not is_blank(row.get(indicator_field, "")):
            if legacy_id_field not in headers or is_blank(row.get(legacy_id_field, "")):
                recorder.add(row_num, child_name, legacy_id_field,
                             f"Emergency contact '{indicator_field}' is populated but "
                             f"'{legacy_id_field}' is missing. A Legacy ID is required.",
                             "ERROR")


def validate_date_fields(row, row_num, child_name, recorder, headers):
    """Validates date fields against YYYY-MM-DD or DD/MM/YYYY formats."""
    for field in DATE_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if not is_valid_date(value):
            recorder.add(row_num, child_name, field,
                         f"Date field '{field}' contains an invalid value '{value}'. "
                         f"Expected format: YYYY-MM-DD or DD/MM/YYYY.", "ERROR")


def validate_crn_format(row, row_num, child_name, recorder, headers):
    """Validates CRN fields: must be exactly 9 digits followed by 1 letter."""
    for field in CRN_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if not is_valid_crn(value):
            recorder.add(row_num, child_name, field,
                         f"CRN field '{field}' contains an invalid value '{value}'. "
                         f"Expected format: 9 digits + 1 letter (e.g. 123456789A).", "ERROR")


def validate_phone_fields(row, row_num, child_name, recorder, headers):
    """Validates phone fields against recognised Australian formats."""
    for field in PHONE_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if not is_valid_phone(value):
            recorder.add(row_num, child_name, field,
                         f"Phone field '{field}' does not match a recognised Australian format: "
                         f"'{value}'. Expected: 04xx xxx xxx (mobile) or 0x xxxx xxxx (landline).",
                         "WARNING")


def validate_email_fields(row, row_num, child_name, recorder, headers):
    """Validates email address format for all email fields."""
    for field in EMAIL_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if not is_valid_email(value):
            recorder.add(row_num, child_name, field,
                         f"Email field '{field}' contains an invalid address: '{value}'.", "ERROR")


def validate_status_field(row, row_num, child_name, recorder, headers):
    """Validates Status — case-insensitive. Accepted: Active, Inactive, Waitlist."""
    field = "Status"
    if field not in headers:
        return
    value = row.get(field, "").strip()
    if is_blank(value):
        return
    if value.lower() not in VALID_STATUSES:
        recorder.add(row_num, child_name, field,
                     f"Status contains an invalid value: '{value}'. "
                     f"Accepted values: Active, Inactive, Waitlist (case-insensitive).", "ERROR")


def validate_gender_fields(row, row_num, child_name, recorder, headers):
    """Validates Gender (Male/Female) and Gender_Identity (standard set or free text)."""
    if "Gender" in headers:
        value = row.get("Gender", "").strip()
        if not is_blank(value) and value not in VALID_GENDERS:
            recorder.add(row_num, child_name, "Gender",
                         f"Gender contains an unexpected value: '{value}'. "
                         f"Accepted values: Male, Female.", "WARNING")
    if "Gender_Identity" in headers:
        value = row.get("Gender_Identity", "").strip()
        if not is_blank(value) and value not in STANDARD_GENDER_IDENTITIES:
            recorder.add(row_num, child_name, "Gender_Identity",
                         f"Gender Identity contains a non-standard value: '{value}'. "
                         f"Standard values: Male, Female, Non-Binary, Trans Female, Trans Male. "
                         f"Free-text entries are permitted but please verify this is intentional.",
                         "WARNING")


def validate_boolean_fields(row, row_num, child_name, recorder, headers):
    """Validates boolean fields; accepts 0/1, Yes/No, True/False (case-insensitive)."""
    for field in BOOLEAN_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if value.lower() not in VALID_BOOLEAN_VALUES:
            recorder.add(row_num, child_name, field,
                         f"Boolean field '{field}' contains an invalid value: '{value}'. "
                         f"Accepted values: 0, 1, Yes, No, True, False.", "ERROR")


def validate_state_fields(row, row_num, child_name, recorder, headers):
    """
    Validates state fields after transformation has run.
    Any remaining non-standard values (untransformable) are flagged as errors.
    """
    for field in STATE_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if value.upper() not in VALID_AU_STATES:
            recorder.add(row_num, child_name, field,
                         f"State field '{field}' still contains a non-standard value: '{value}' "
                         f"after attempted normalisation. "
                         f"Expected: NSW, VIC, QLD, SA, WA, TAS, ACT, NT.", "ERROR")


def validate_postcode_fields(row, row_num, child_name, recorder, headers):
    """Validates postcode fields: must be a 4-digit Australian postcode."""
    for field in POSTCODE_FIELDS:
        if field not in headers:
            continue
        value = row.get(field, "").strip()
        if is_blank(value):
            continue
        if not POSTCODE_PATTERN.match(value):
            recorder.add(row_num, child_name, field,
                         f"Postcode field '{field}' contains an invalid value: '{value}'. "
                         f"Expected a 4-digit Australian postcode (e.g. 2000).", "WARNING")


def validate_field_lengths(row, row_num, child_name, recorder, headers):
    """Validates that field values do not exceed their maximum permitted character length."""
    for field, limit in FIELD_LENGTH_LIMITS.items():
        if field not in headers:
            continue
        value = row.get(field, "")
        if value and len(value) > limit:
            recorder.add(row_num, child_name, field,
                         f"Field '{field}' exceeds the maximum of {limit} characters "
                         f"(current length: {len(value)}).", "ERROR")


def validate_service_id(row, row_num, child_name, recorder, headers):
    """Validates that ServiceID is a positive integer."""
    field = "ServiceID"
    if field not in headers:
        return
    value = row.get(field, "").strip()
    if is_blank(value):
        return
    try:
        if int(value) <= 0:
            raise ValueError
    except ValueError:
        recorder.add(row_num, child_name, field,
                     f"Service ID '{value}' is not a valid positive integer. "
                     f"Please verify the correct Xplor Service ID.", "ERROR")


def validate_paired_name_fields(row, row_num, child_name, recorder, headers):
    """First name requires last name, and vice versa, for all paired name fields."""
    for first_field, last_field in PAIRED_NAME_FIELDS:
        if first_field not in headers or last_field not in headers:
            continue
        first = row.get(first_field, "").strip()
        last  = row.get(last_field,  "").strip()
        if first and not last:
            recorder.add(row_num, child_name, last_field,
                         f"'{first_field}' is populated ('{first}') but '{last_field}' is missing. "
                         f"Both must be provided together.", "ERROR")
        elif last and not first:
            recorder.add(row_num, child_name, first_field,
                         f"'{last_field}' is populated ('{last}') but '{first_field}' is missing. "
                         f"Both must be provided together.", "ERROR")


def validate_waitlist_logic(row, row_num, child_name, recorder, headers):
    """
    Validates Waitlist-specific business rules:
      1. Incomplete Waitlist — Waitlist status but no Parent 1 guardian information.
      2. Stale Waitlist — Enrolment_Start_Date has already passed.
    """
    if row.get("Status", "").strip().lower() != "waitlist":
        return

    # Rule 1 — Incomplete Waitlist
    has_parent = any(
        not is_blank(row.get(f, ""))
        for f in ["Parent1_Email", "Parent1_CRN", "Parent1_Legacy_Account_ID"]
        if f in headers
    )
    if not has_parent:
        recorder.add(row_num, child_name, "Status / Parent1_Email",
                     "Child has 'Waitlist' status but no Parent 1 guardian information is provided. "
                     "Please supply guardian details or confirm whether this record should remain "
                     "on the waitlist.", "WARNING")

    # Rule 2 — Stale Waitlist
    enrolment_str = row.get("Enrolment_Start_Date", "").strip()
    if not is_blank(enrolment_str) and is_valid_date(enrolment_str):
        if parse_date(enrolment_str) < date.today():
            recorder.add(row_num, child_name, "Status / Enrolment_Start_Date",
                         f"Child has 'Waitlist' status but the Enrolment_Start_Date "
                         f"({enrolment_str}) has already passed. "
                         f"The status should likely be updated to 'Active' before import.",
                         "ERROR")


def transform_crn_child_parent_equality(row, row_num, child_name, recorder, headers):
    """
    If Child_CRN is identical to Parent1_CRN or Parent2_CRN, clear Child_CRN
    automatically and log a FIXED entry.  A child and parent cannot share a CRN.
    """
    child_crn = row.get("Child_CRN", "").strip()
    if is_blank(child_crn):
        return
    for parent_crn_field in ["Parent1_CRN", "Parent2_CRN"]:
        if parent_crn_field not in headers:
            continue
        parent_crn = row.get(parent_crn_field, "").strip()
        if parent_crn and child_crn == parent_crn:
            row["Child_CRN"] = ""
            recorder.add(
                row_num, child_name, "Child_CRN",
                f"Child_CRN '{child_crn}' was identical to {parent_crn_field}. "
                f"Child_CRN has been removed — a child and parent cannot share the same CRN.",
                "FIXED",
                action=f"Child_CRN cleared (was '{child_crn}', same as {parent_crn_field})",
            )
            return  # Only need to clear once; no point checking the second parent


# Keep a thin validation wrapper for any remaining cases (should not trigger
# after the transform above, but acts as a safety net).
def validate_crn_child_parent_equality(row, row_num, child_name, recorder, headers):
    """Child CRN must not be identical to any Parent CRN (post-transform safety check)."""
    child_crn = row.get("Child_CRN", "").strip()
    if is_blank(child_crn):
        return
    for parent_crn_field in ["Parent1_CRN", "Parent2_CRN"]:
        if parent_crn_field not in headers:
            continue
        parent_crn = row.get(parent_crn_field, "").strip()
        if parent_crn and child_crn == parent_crn:
            recorder.add(row_num, child_name, f"Child_CRN / {parent_crn_field}",
                         f"Child CRN '{child_crn}' is identical to {parent_crn_field}. "
                         f"A child's CRN and parent's CRN must differ.", "ERROR")


def validate_future_dob(row, row_num, child_name, recorder, headers):
    """Child DOB must not be a future date."""
    dob_str = row.get("DOB", "").strip()
    if is_blank(dob_str) or not is_valid_date(dob_str):
        return
    if parse_date(dob_str) > date.today():
        recorder.add(row_num, child_name, "DOB",
                     f"Child's date of birth ({dob_str}) is set to a future date. "
                     f"Please verify this is correct.", "ERROR")


def validate_enrolment_date_not_before_dob(row, row_num, child_name, recorder, headers):
    """Enrolment_Start_Date must not precede the child's DOB."""
    dob_str       = row.get("DOB",                "").strip()
    enrolment_str = row.get("Enrolment_Start_Date","").strip()
    if is_blank(dob_str) or is_blank(enrolment_str):
        return
    if not is_valid_date(dob_str) or not is_valid_date(enrolment_str):
        return
    if parse_date(enrolment_str) < parse_date(dob_str):
        recorder.add(row_num, child_name, "Enrolment_Start_Date",
                     f"Enrolment_Start_Date ({enrolment_str}) is earlier than the child's "
                     f"date of birth ({dob_str}).", "ERROR")


def validate_medicare_number(row, row_num, child_name, recorder, headers):
    """Medicare Number must be numeric if provided."""
    if "Medicare_Number" not in headers:
        return
    value = row.get("Medicare_Number", "").strip()
    if not is_blank(value) and not value.isdigit():
        recorder.add(row_num, child_name, "Medicare_Number",
                     f"Medicare Number '{value}' contains non-numeric characters.", "WARNING")


def validate_consents_photos_videos(row, row_num, child_name, recorder, headers):
    """Flags blank Consents_Photos_Videos — commonly overlooked before import."""
    field = "Consents_Photos_Videos"
    if field not in headers:
        return
    if is_blank(row.get(field, "")):
        recorder.add(row_num, child_name, field,
                     f"'{field}' is blank. Please confirm consent for photos/videos "
                     f"(1 = yes, 0 = no).", "WARNING")


def validate_ec_parent_email_redundancy(row, row_num, child_name, recorder, headers):
    """
    If any Emergency Contact email matches Parent 1 or Parent 2 email in the same
    row, clears the EC email and records a FIXED entry.  A parent listed as an
    emergency contact creates a duplicate/redundant contact record in Xplor.

    Note: Parent2_Email deduplication (against Parent1_Email) runs as a FIXED
    transformation before this check, so Parent2_Email will already be cleared
    when it duplicates Parent1_Email.
    """
    p1_email = row.get("Parent1_Email", "").strip().lower()
    p2_email = row.get("Parent2_Email", "").strip().lower()
    parent_emails: dict[str, str] = {}
    if p1_email:
        parent_emails[p1_email] = "Parent1_Email"
    if p2_email:
        parent_emails[p2_email] = "Parent2_Email"

    if not parent_emails:
        return

    for ec_field in EC_EMAIL_FIELDS:
        if ec_field not in headers:
            continue
        ec_email = row.get(ec_field, "").strip().lower()
        if is_blank(ec_email):
            continue
        if ec_email in parent_emails:
            matched_parent = parent_emails[ec_email]
            original = row[ec_field]
            row[ec_field] = ""
            recorder.add(
                row_num, child_name, ec_field,
                f"Redundant Contact: '{ec_field}' ('{original.strip()}') "
                f"is identical to {matched_parent}. A parent must not also be listed "
                f"as an Emergency Contact — the email has been cleared.",
                "FIXED",
                action=f"Duplicate email removed from '{ec_field}': '{original.strip()}' (same as {matched_parent})",
                tag="redundant_ec",
            )


# ─────────────────────────────────────────────────────────────────────────────
# CROSS-ROW VALIDATION FUNCTIONS  (unchanged from v1)
# ─────────────────────────────────────────────────────────────────────────────

def _get_parent_identity_key(row: dict, prefix: str) -> str:
    """Returns the best available unique identifier for a parent (CRN > Legacy ID > name)."""
    crn = row.get(f"{prefix}_CRN", "").strip()
    if not is_blank(crn):
        return crn.lower()
    legacy = row.get(f"{prefix}_Legacy_Account_ID", "").strip()
    if not is_blank(legacy):
        return legacy.lower()
    first = row.get(f"{prefix}_First_Name", "").strip().lower()
    last  = row.get(f"{prefix}_Last_Name",  "").strip().lower()
    if first or last:
        return f"{first}|{last}"
    return ""


def check_duplicates(all_rows, recorder, headers):
    """Checks Child_Legacy_Id and Child_CRN for uniqueness across all rows."""
    config = {"Child_Legacy_Id": "ERROR", "Child_CRN": "ERROR"}
    seen: defaultdict[str, defaultdict[str, list[tuple[int, str]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for entry in all_rows:
        for field in config:
            if field not in headers:
                continue
            value = entry["row"].get(field, "").strip()
            if not is_blank(value):
                seen[field][value.lower()].append((entry["row_num"], entry["child_name"]))
    for field, severity in config.items():
        if field not in headers:
            continue
        for value, occurrences in seen[field].items():
            if len(occurrences) > 1:
                row_nums_str = ", ".join(str(r) for r, _ in occurrences)
                for row_num, child_name in occurrences:
                    recorder.add(row_num, child_name, field,
                                 f"Duplicate value in '{field}': '{value}' appears in rows "
                                 f"{row_nums_str}. Each record must have a unique value.",
                                 severity)


def check_duplicate_parent_emails(all_rows, recorder, headers):
    """
    Parent 1 Email: same email + different CRN/identity = ERROR.
    Parent 2 Email: any duplicate = WARNING.
    Same email + same identity (one parent, many children) = no issue.
    """
    for prefix, error_on_diff_identity in [("Parent1", True), ("Parent2", False)]:
        email_field = f"{prefix}_Email"
        if email_field not in headers:
            continue
        email_to_identities: defaultdict[str, set[str]] = defaultdict(set)
        email_to_rows: defaultdict[str, list[tuple[int, str]]] = defaultdict(list)
        for entry in all_rows:
            email = entry["row"].get(email_field, "").strip().lower()
            if is_blank(email):
                continue
            identity = _get_parent_identity_key(entry["row"], prefix)
            email_to_identities[email].add(identity)
            email_to_rows[email].append((entry["row_num"], entry["child_name"]))
        for email, identities in email_to_identities.items():
            occurrences = email_to_rows[email]
            if len(occurrences) <= 1:
                continue
            if error_on_diff_identity:
                if len(identities) > 1:
                    row_nums_str = ", ".join(str(r) for r, _ in occurrences)
                    for row_num, child_name in occurrences:
                        recorder.add(row_num, child_name, email_field,
                                     f"Duplicate '{email_field}': '{email}' appears in rows "
                                     f"{row_nums_str} linked to different parent identities. "
                                     f"The same email cannot belong to two different parents.",
                                     "ERROR")
            else:
                row_nums_str = ", ".join(str(r) for r, _ in occurrences)
                for row_num, child_name in occurrences:
                    recorder.add(row_num, child_name, email_field,
                                 f"Duplicate '{email_field}': '{email}' appears in rows "
                                 f"{row_nums_str}. Please verify this is intentional.",
                                 "WARNING")


def check_parent_crn_email_registry(all_rows, recorder, headers):
    """
    Enforces '1 CRN = 1 Email'. A CRN linked to multiple distinct emails = ERROR.
    Complements check_duplicate_parent_emails (which checks the reverse direction).
    """
    for prefix in ("Parent1", "Parent2"):
        crn_field, email_field = f"{prefix}_CRN", f"{prefix}_Email"
        if crn_field not in headers or email_field not in headers:
            continue
        crn_to_emails: defaultdict[str, set[str]] = defaultdict(set)
        crn_to_rows:   defaultdict[str, list[tuple[int, str]]] = defaultdict(list)
        for entry in all_rows:
            row = entry["row"]
            crn   = row.get(crn_field,   "").strip().lower()
            email = row.get(email_field, "").strip().lower()
            if is_blank(crn):
                continue
            crn_to_rows[crn].append((entry["row_num"], entry["child_name"]))
            if not is_blank(email):
                crn_to_emails[crn].add(email)
        for crn, emails in crn_to_emails.items():
            if len(emails) <= 1:
                continue
            emails_str   = ", ".join(sorted(emails))
            row_nums_str = ", ".join(str(r) for r, _ in crn_to_rows[crn])
            for row_num, child_name in crn_to_rows[crn]:
                recorder.add(row_num, child_name, f"{crn_field} / {email_field}",
                             f"Parent CRN '{crn.upper()}' is linked to multiple email addresses "
                             f"across rows {row_nums_str}: {emails_str}. "
                             f"One CRN must map to exactly one email address.", "ERROR")


def check_enrolment_parent_crn_consistency(all_rows, recorder, headers):
    """Enrolment_Parent_CRN must match Parent1_CRN or Parent2_CRN in the same row."""
    field = "Enrolment_Parent_CRN"
    if field not in headers:
        return
    for entry in all_rows:
        row = entry["row"]
        enrolment_crn = row.get(field, "").strip()
        if is_blank(enrolment_crn):
            continue
        p1 = row.get("Parent1_CRN", "").strip() if "Parent1_CRN" in headers else ""
        p2 = row.get("Parent2_CRN", "").strip() if "Parent2_CRN" in headers else ""
        if enrolment_crn not in {crn for crn in [p1, p2] if crn}:
            recorder.add(entry["row_num"], entry["child_name"], field,
                         f"Enrolment_Parent_CRN '{enrolment_crn}' does not match any listed "
                         f"guardian's CRN (Parent1: '{p1}', Parent2: '{p2}'). "
                         f"The CCS parent must be one of the child's listed guardians.", "ERROR")


def check_cross_service_parent_duplicates(
    all_rows:          list,
    existing_profiles: list[dict],
    recorder:          "IssueRecorder",
) -> None:
    """
    Cross-checks every Parent 1 and Parent 2 in the input file against parent
    profiles extracted from existing-service data files.

    A match is flagged as an ERROR when:
      • First name AND last name both match (case-insensitive), AND
      • At least one of DOB, contact number, or email also matches.

    This combination minimises false positives from common names while reliably
    catching genuine duplicate accounts that would be created on import.

    Action required (written into the report):
      Link the affected children to the already-existing parent profile, then
      delete the newly created duplicate profile.
    """
    if not existing_profiles:
        return

    for entry in all_rows:
        row        = entry["row"]
        row_num    = entry["row_num"]
        child_name = entry["child_name"]

        for prefix in ("Parent1", "Parent2"):
            first_name = row.get(f"{prefix}_First_Name", "").strip().lower()
            last_name  = row.get(f"{prefix}_Last_Name",  "").strip().lower()
            if not first_name and not last_name:
                continue  # No parent data in this slot

            dob = row.get(f"{prefix}_DOB", "").strip().lower()

            contacts: set[str] = set()
            for cf in PARENT_PHONE_FIELDS_BY_PREFIX.get(prefix, []):
                v = row.get(cf, "").strip()
                if v:
                    contacts.add(re.sub(r"[\s\-\(\)]", "", v).lower())

            emails: set[str] = set()
            for ef in PARENT_EMAIL_FIELDS_BY_PREFIX.get(prefix, []):
                v = row.get(ef, "").strip().lower()
                if v:
                    emails.add(v)

            input_legacy_id = row.get(f"{prefix}_Legacy_Account_ID", "").strip().lower()

            # Compare against each existing profile
            for profile in existing_profiles:
                if profile["first_name"] != first_name or profile["last_name"] != last_name:
                    continue  # Names don't match — skip quickly

                # Matching Legacy IDs mean the system will link the accounts automatically
                # rather than creating a new profile — no duplicate will occur.
                if input_legacy_id and profile["legacy_id"] and input_legacy_id == profile["legacy_id"]:
                    continue

                # Names match: check at least one secondary field
                matched_fields: list[str] = []

                if dob and profile["dob"] and dob == profile["dob"]:
                    matched_fields.append(f"DOB: {row.get(f'{prefix}_DOB', '').strip()}")

                shared_contacts = contacts & profile["contacts"]
                if shared_contacts:
                    matched_fields.append(f"Contact: {next(iter(shared_contacts))}")

                shared_emails = emails & profile["emails"]
                if shared_emails:
                    matched_fields.append(f"Email: {next(iter(shared_emails))}")

                if not matched_fields:
                    continue  # Name match only — not enough confidence

                display_name = (
                    f"{row.get(f'{prefix}_First_Name', '').strip()} "
                    f"{row.get(f'{prefix}_Last_Name', '').strip()}".strip()
                )
                svc_label = (
                    f"Service ID {profile['service_id']}"
                    if profile["service_id"]
                    else "unknown service"
                )
                matched_str = ", ".join(matched_fields)
                recorder.add(
                    row_num,
                    child_name,
                    f"{prefix}_First_Name / {prefix}_Last_Name",
                    f"Potential duplicate parent: '{display_name}' matches an existing parent "
                    f"profile in '{profile['source_file']}' ({svc_label}). "
                    f"Matched on: {matched_str}. "
                    f"Action required: (1) Link this parent's children to the existing profile "
                    f"in the existing service. (2) Delete the newly created duplicate profile.",
                    "ERROR",
                    tag="cross_service_duplicate_parent",
                    parent_slot=prefix,
                    parent_name=display_name,
                    matched_on=matched_str,
                    duplicate_source=f"Existing file: {profile['source_file']}",
                    duplicate_service_id=profile["service_id"],
                    duplicate_parent_crn=profile["parent_crn"],
                    duplicate_legacy_id=profile["legacy_id"],
                )
                break  # Report once per parent slot — first match is sufficient


def check_intra_file_parent_duplicates(
    all_rows: list,
    recorder: "IssueRecorder",
) -> None:
    """
    Detects duplicate parent profiles within the upload file.

    A duplicate profile is detected when two parent slots (P1 or P2, in any
    combination, across any services) satisfy ALL of:

      1. First name + last name match (case-insensitive)
      2. Date of birth matches
      3. At least one contact number matches (normalised — spaces and hyphens stripped)
      4. Legacy Account IDs differ  ← confirms two distinct profiles for the same person

    Condition 4 is the key: if the same person correctly shares one Legacy ID
    across multiple child rows, the system will link them automatically and no
    action is needed.  Differing Legacy IDs mean two separate accounts would be
    created for the same physical person.

    This check covers two real-world scenarios:
      • Cross-service  — parent appears in different services with different IDs
      • Same-service   — parent appears as P1 for one child and P2 for another
                         with different IDs (e.g. data-entry error in source system)

    Auto-link action: once a duplicate pair is found, the later profile's
    Legacy Account ID and CRN are overwritten in-place with the first-created
    profile's (the one with the lower row number) so both children link to the
    same parent profile post-publishing, instead of creating two accounts. The
    row mutation happens before CSVs are written, so the fix carries through to
    the import file.  The action is logged as FIXED (not WARNING) for both rows.

    Each pair is reported once — from the perspective of the earlier row.
    """
    # Build a flat profile list with row context
    profiles: list[dict] = []
    for entry in all_rows:
        row        = entry["row"]
        row_num    = entry["row_num"]
        child_name = entry["child_name"]
        service_id = row.get("ServiceID", "").strip()

        for prefix in ("Parent1", "Parent2"):
            first_name = row.get(f"{prefix}_First_Name", "").strip().lower()
            last_name  = row.get(f"{prefix}_Last_Name",  "").strip().lower()
            if not first_name and not last_name:
                continue

            dob = row.get(f"{prefix}_DOB", "").strip()

            contacts: set[str] = set()
            for cf in PARENT_PHONE_FIELDS_BY_PREFIX.get(prefix, []):
                v = row.get(cf, "").strip()
                if v:
                    contacts.add(re.sub(r"[\s\-\(\)]", "", v).lower())

            crn       = row.get(f"{prefix}_CRN",                 "").strip().lower()
            legacy_id = row.get(f"{prefix}_Legacy_Account_ID",   "").strip().lower()
            display   = (
                f"{row.get(f'{prefix}_First_Name', '').strip()} "
                f"{row.get(f'{prefix}_Last_Name',  '').strip()}".strip()
            )

            profiles.append({
                "row":         row,
                "prefix":      prefix,
                "row_num":     row_num,
                "child_name":  child_name,
                "parent_slot": prefix,
                "display":     display,
                "first_name":  first_name,
                "last_name":   last_name,
                "dob":         dob,
                "contacts":    contacts,
                "crn":         crn,
                "legacy_id":   legacy_id,
                "service_id":  service_id,
            })

    reported_pairs: set[frozenset] = set()

    for i, pa in enumerate(profiles):
        for pb in profiles[i + 1:]:
            # ── Condition 1: names must match ─────────────────────────────────
            if pa["first_name"] != pb["first_name"] or pa["last_name"] != pb["last_name"]:
                continue

            # ── Condition 2: DOB must match ───────────────────────────────────
            if not pa["dob"] or not pb["dob"] or pa["dob"] != pb["dob"]:
                continue

            # ── Condition 3: at least one contact number must match ───────────
            shared_contacts = pa["contacts"] & pb["contacts"]
            if not shared_contacts:
                continue

            # ── Condition 4: legacy IDs must differ ───────────────────────────
            # Same base legacy ID = intentional shared profile; skip.
            # Strip any '_\d+' uniquifier suffix added by transform_legacy_ids
            # so that e.g. '248992' and '248992_1' are treated as the same person.
            def _base(lid: str) -> str:
                return re.sub(r'_\d+$', '', lid) if lid else lid

            if pa["legacy_id"] and pb["legacy_id"] and _base(pa["legacy_id"]) == _base(pb["legacy_id"]):
                continue

            # Avoid reporting the same pair twice (A→B and B→A)
            pair_key = frozenset([
                (pa["row_num"], pa["parent_slot"]),
                (pb["row_num"], pb["parent_slot"]),
            ])
            if pair_key in reported_pairs:
                continue
            reported_pairs.add(pair_key)

            same_service = pa["service_id"] == pb["service_id"]
            context = (
                f"Service {pa['service_id']}"
                if same_service
                else f"Service {pa['service_id']} and Service {pb['service_id']}"
            )
            matched_contact = next(iter(shared_contacts))
            matched_str = f"DOB: {pa['dob']}, Contact: {matched_contact}"

            # ── Auto-link: keep the first-created profile's identity, and
            # overwrite the later profile's Legacy Account ID / CRN with it so
            # both children resolve to the same parent profile post-publishing.
            first, later = (pa, pb) if pa["row_num"] <= pb["row_num"] else (pb, pa)

            first_legacy_raw = first["row"].get(f"{first['prefix']}_Legacy_Account_ID", "").strip()
            first_crn_raw    = first["row"].get(f"{first['prefix']}_CRN", "").strip()
            later_legacy_before = later["row"].get(f"{later['prefix']}_Legacy_Account_ID", "").strip()
            later_crn_before    = later["row"].get(f"{later['prefix']}_CRN", "").strip()

            if first_legacy_raw:
                later["row"][f"{later['prefix']}_Legacy_Account_ID"] = first_legacy_raw
            if first_crn_raw:
                later["row"][f"{later['prefix']}_CRN"] = first_crn_raw

            msg = (
                f"Duplicate parent profile: '{pa['display']}' (legacy '{pa['legacy_id']}') and "
                f"'{pb['display']}' (legacy '{pb['legacy_id']}') appear to be the same person "
                f"in {context}. Matched on: {matched_str}. "
                f"Auto-linked: {later['child_name']}'s {later['parent_slot']} profile "
                f"(legacy '{later_legacy_before}', CRN '{later_crn_before}') was reassigned to "
                f"{first['child_name']}'s {first['parent_slot']} profile "
                f"(legacy '{first_legacy_raw}', CRN '{first_crn_raw}', first created at row {first['row_num']}) "
                f"so both children link to the same parent profile post-publishing."
            )

            # Record once for each side so both legacy IDs are searchable in the report
            for src, dst in ((pa, pb), (pb, pa)):
                recorder.add(
                    src["row_num"],
                    src["child_name"],
                    f"{src['parent_slot']}_Legacy_Account_ID",
                    msg,
                    "FIXED",
                    action=(
                        f"Linked to first-created profile at row {first['row_num']} "
                        f"(legacy '{first_legacy_raw}', CRN '{first_crn_raw}')"
                        if src is later else
                        f"Kept as first-created profile; row {later['row_num']} linked to it"
                    ),
                    tag="intra_file_duplicate_parent",
                    parent_slot=src["parent_slot"],
                    parent_name=src["display"],
                    matched_on=matched_str,
                    duplicate_source="Input file",
                    duplicate_row_num=dst["row_num"],
                    duplicate_service_id=dst["service_id"],
                    duplicate_parent_crn=dst["crn"],
                    duplicate_legacy_id=dst["legacy_id"],
                )


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT — SPLIT BY SERVICE
# ─────────────────────────────────────────────────────────────────────────────

def sanitise_filename(name: str) -> str:
    """Removes characters not permitted in filenames across Windows and macOS."""
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def write_split_csvs(
    all_rows: list,
    fieldnames: list[str],
    output_dir: str,
    service_map: "ServiceMapping | None" = None,
) -> tuple[dict[str, str], dict[int, int]]:
    """
    Writes one CSV per distinct Xplor Service ID found in the transformed data.
    Files are named: {Service_Name}_families_import.csv
    Falls back to {Service_ID}_families_import.csv when the name is unavailable.

    Returns:
        output_paths — dict mapping service_id -> output file path
        row_num_map  — dict mapping original row_num -> row number within the
                       service's output file (row 2 = first data row after header)
    """
    service_entries: defaultdict[str, list[dict]] = defaultdict(list)
    for entry in all_rows:
        svc_id = entry["row"].get("ServiceID", "").strip()
        service_entries[svc_id or "Unknown"].append(entry)

    output_paths: dict[str, str] = {}
    row_num_map:  dict[int, int]  = {}

    for svc_id, entries in service_entries.items():
        # Prefer the human-readable service name for the filename.
        # Try (in order): Service_Name field in the row data, then service_map lookup.
        svc_name = ""
        if entries:
            svc_name = entries[0]["row"].get("Service_Name", "").strip()
        if not svc_name and service_map and service_map.is_loaded:
            candidate = service_map.get_name_by_xplor(svc_id)
            if candidate != svc_id:   # returns the ID itself when not found
                svc_name = candidate
        label    = sanitise_filename(svc_name or svc_id or "Unknown")
        filename = f"{label}_families_import.csv"
        out_path = os.path.join(output_dir, filename)

        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for out_row_idx, entry in enumerate(entries, start=2):
                writer.writerow(entry["row"])
                row_num_map[entry["row_num"]] = out_row_idx

        output_paths[svc_id] = out_path
        print(f"  [CSV] {len(entries):>4} rows -> {filename}")

    return output_paths, row_num_map


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT — EXCEL AUDIT REPORT (.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

# Colour palettes for the report — pastel tones, easy on the eyes
COLOUR_ERROR   = "FFE4E6"   # Pastel blush pink
COLOUR_WARNING = "FFF8DC"   # Pastel cornsilk yellow
COLOUR_FIXED   = "E6F4EA"   # Pastel mint green
COLOUR_HEADER  = "4472C4"   # Xplor cornflower blue
COLOUR_SUMMARY = "EBF3FB"   # Very light sky blue

# Tags that identify client-facing issues (used to populate client_audit_report.xlsx)
CLIENT_ISSUE_TAGS = {"duplicate_parent_email", "redundant_ec"}
CLIENT_TAG_LABELS = {
    "duplicate_parent_email": "Duplicate Parent Emails",
    "redundant_ec":           "Redundant Emergency Contacts",
}


def _svc_sort_key(svc_id: str, service_map: "ServiceMapping") -> tuple:
    """Sort key: alphabetical by service name, 'Unknown' always last."""
    if svc_id == "Unknown":
        return (1, "")
    name = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
    return (0, name.lower())


def _apply_header_style(ws, row_num: int, num_cols: int) -> None:
    """Applies a dark navy header style to a row in the given worksheet."""
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor=COLOUR_HEADER)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_size_columns(ws, min_width: int = 12, max_width: int = 60) -> None:
    """Sets column widths based on the maximum content length in each column."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def _remap_row_numbers_in_text(text: str, row_num_map: dict) -> str:
    """
    Replaces input row numbers with output row numbers inside description strings.
    Only replaces numbers that appear in 'rows X, Y, Z' / 'rows X' patterns so
    that row numbers embedded in other values (CRNs, postcodes, etc.) are untouched.
    """
    def _replace(match: re.Match) -> str:
        nums_part = match.group(1)
        remapped = []
        for token in re.split(r",\s*", nums_part):
            try:
                old = int(token.strip())
                remapped.append(str(row_num_map.get(old, old)))
            except ValueError:
                remapped.append(token.strip())
        return "rows " + ", ".join(remapped)

    return re.sub(r"\brows\s+([\d,\s]+)", _replace, text)


def write_excel_report(
    recorder: IssueRecorder,
    all_rows: list,
    output_path: str,
    service_map: ServiceMapping,
    row_num_map: dict | None = None,
) -> None:
    """
    Writes a colour-coded Excel audit report with one tab per service and a
    Summary tab showing overall counts.

    Strategy for performance on large files (thousands of rows):
      - Data is bulk-written via pandas ExcelWriter (fast).
      - openpyxl styling is applied in a second pass over cells only
        (header row + Severity_Level column), avoiding per-cell fill loops
        on every data row which would be prohibitively slow at scale.
      - Row background colours are applied only to the Severity_Level cell
        and a narrow 'indicator' band rather than every column.

    Severity colour coding:
      ERROR   — light red background
      WARNING — light amber background
      FIXED   — light green background
    """
    df = recorder.to_dataframe()

    # Build a map from row_num -> service_id for grouping report entries
    row_to_service: dict[int, str] = {}
    for entry in all_rows:
        row_to_service[entry["row_num"]] = entry["row"].get("ServiceID", "").strip()

    df["_service_id"] = df["Row"].map(row_to_service).fillna("Unknown")

    # Remap Row column and any row references in descriptions to output row numbers
    if row_num_map:
        df["Row"] = df["Row"].map(lambda r: row_num_map.get(int(r), r) if pd.notna(r) else r)
        df["Issue_Description"] = df["Issue_Description"].apply(
            lambda d: _remap_row_numbers_in_text(str(d), row_num_map) if pd.notna(d) else d
        )
    service_ids = sorted(df["_service_id"].unique(), key=lambda sid: _svc_sort_key(sid, service_map))
    report_cols = IssueRecorder.REPORT_FIELDNAMES

    # ── Build summary data ────────────────────────────────────────────────────
    summary_rows = []
    total_errors = total_warnings = total_fixed = 0
    for svc_id in service_ids:
        svc_df   = df[df["_service_id"] == svc_id]
        errors   = int((svc_df["Severity_Level"] == "ERROR").sum())
        warnings = int((svc_df["Severity_Level"] == "WARNING").sum())
        fixed    = int((svc_df["Severity_Level"] == "FIXED").sum())
        svc_name = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
        total_errors += errors; total_warnings += warnings; total_fixed += fixed
        summary_rows.append({
            "Service ID": svc_id, "Service Name": svc_name,
            "Errors": errors, "Warnings": warnings,
            "Fixed": fixed, "Total Issues": errors + warnings + fixed,
        })
    summary_rows.append({
        "Service ID": "TOTAL", "Service Name": "",
        "Errors": total_errors, "Warnings": total_warnings,
        "Fixed": total_fixed, "Total Issues": total_errors + total_warnings + total_fixed,
    })
    summary_df = pd.DataFrame(summary_rows)

    # ── Bulk write all sheets via pandas (fast path) ──────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        for svc_id in service_ids:
            svc_df     = df[df["_service_id"] == svc_id][report_cols].copy()
            svc_name   = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
            sheet_name = sanitise_filename(svc_name or svc_id or "Unknown")[:31]
            svc_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # ── Second pass: apply styling via openpyxl ───────────────────────────────
    # Re-open the workbook to add header styles and severity colours.
    # We only colour the Severity_Level cell (not the whole row) to keep the
    # file size small and writes fast on large datasets.
    wb = load_workbook(output_path)

    fill_map = {
        "ERROR":   PatternFill("solid", fgColor=COLOUR_ERROR),
        "WARNING": PatternFill("solid", fgColor=COLOUR_WARNING),
        "FIXED":   PatternFill("solid", fgColor=COLOUR_FIXED),
    }
    summary_fill  = PatternFill("solid", fgColor=COLOUR_SUMMARY)
    bold_font     = Font(bold=True)

    # Style Summary sheet
    ws_sum = wb["Summary"]
    _apply_header_style(ws_sum, 1, len(summary_df.columns))
    for row_idx in range(2, ws_sum.max_row):   # all rows except totals
        for col in range(1, len(summary_df.columns) + 1):
            ws_sum.cell(row=row_idx, column=col).fill = summary_fill
    for col in range(1, len(summary_df.columns) + 1):  # totals row bold
        ws_sum.cell(row=ws_sum.max_row, column=col).font = bold_font
    ws_sum.freeze_panes = "A2"
    _auto_size_columns(ws_sum)

    # Style per-service sheets
    sev_col_idx  = report_cols.index("Severity_Level") + 1   # 1-based
    desc_col_idx = report_cols.index("Issue_Description") + 1

    for svc_id in service_ids:
        svc_name   = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
        sheet_name = sanitise_filename(svc_name or svc_id or "Unknown")[:31]
        ws = wb[sheet_name]

        _apply_header_style(ws, 1, len(report_cols))
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            sev_cell = row[sev_col_idx - 1]   # 0-based index
            severity = str(sev_cell.value or "")
            row_fill = fill_map.get(severity)
            if row_fill:
                # Colour the entire row for readability
                for cell in row:
                    cell.fill = row_fill
            # Wrap long descriptions
            row[desc_col_idx - 1].alignment = Alignment(wrap_text=True)

        _auto_size_columns(ws)

    wb.save(output_path)


def write_client_excel_report(
    recorder:    "IssueRecorder",
    all_rows:    list,
    output_path: str,
    service_map: "ServiceMapping",
    row_num_map: dict | None = None,
) -> None:
    """
    Writes a client-facing Excel audit report containing only two issue types:

      • Duplicate Parent Emails   — Parent1_Email == Parent2_Email (auto-fixed)
      • Redundant Emergency Contacts — an EC email matches a parent email (auto-fixed)

    Structure mirrors the main report:
      - Alphabetical tabs by Service Name
      - Same column layout (Row, Child_Name, Field, Issue_Description,
        Severity_Level, Action_Taken)
      - Summary tab with per-service counts of each issue type
      - Output row numbers applied throughout
    """
    report_cols = IssueRecorder.REPORT_FIELDNAMES

    # ── Filter to client-facing issues only ───────────────────────────────────
    client_issues = [i for i in recorder.issues if i.get("_tag") in CLIENT_ISSUE_TAGS]

    if client_issues:
        df = pd.DataFrame(client_issues)
        # Ensure all report columns exist (guard against edge cases)
        for col in report_cols:
            if col not in df.columns:
                df[col] = ""
    else:
        df = pd.DataFrame(columns=report_cols + ["_tag"])

    # ── Tag service group ─────────────────────────────────────────────────────
    row_to_service: dict[int, str] = {
        entry["row_num"]: entry["row"].get("ServiceID", "").strip()
        for entry in all_rows
    }
    if not df.empty:
        df["_service_id"] = df["Row"].map(row_to_service).fillna("Unknown")
    else:
        df["_service_id"] = pd.Series(dtype=str)

    # ── Remap row numbers to output file row numbers ──────────────────────────
    if row_num_map and not df.empty:
        df["Row"] = df["Row"].map(lambda r: row_num_map.get(int(r), r) if pd.notna(r) else r)
        df["Issue_Description"] = df["Issue_Description"].apply(
            lambda d: _remap_row_numbers_in_text(str(d), row_num_map) if pd.notna(d) else d
        )

    # ── Alphabetical service order (Unknown last) ─────────────────────────────
    unique_ids   = df["_service_id"].unique() if not df.empty else []
    service_ids  = sorted(unique_ids, key=lambda sid: _svc_sort_key(sid, service_map))

    # ── Build client Summary tab ──────────────────────────────────────────────
    summary_rows = []
    total_dup = total_ec = 0
    for svc_id in service_ids:
        svc_df    = df[df["_service_id"] == svc_id]
        dup_count = int((svc_df["_tag"] == "duplicate_parent_email").sum())
        ec_count  = int((svc_df["_tag"] == "redundant_ec").sum())
        svc_name  = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
        total_dup += dup_count
        total_ec  += ec_count
        summary_rows.append({
            "Service Name":                  svc_name,
            "Duplicate Parent Emails":        dup_count,
            "Redundant Emergency Contacts":   ec_count,
            "Total Issues":                   dup_count + ec_count,
        })
    summary_rows.append({
        "Service Name":                  "TOTAL",
        "Duplicate Parent Emails":        total_dup,
        "Redundant Emergency Contacts":   total_ec,
        "Total Issues":                   total_dup + total_ec,
    })
    summary_df = pd.DataFrame(summary_rows)

    # ── Bulk write via pandas ─────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        for svc_id in service_ids:
            svc_df     = df[df["_service_id"] == svc_id][report_cols].copy()
            svc_name   = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
            sheet_name = sanitise_filename(svc_name or svc_id or "Unknown")[:31]
            svc_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # ── Styling pass ──────────────────────────────────────────────────────────
    wb = load_workbook(output_path)

    fill_map = {
        "ERROR":   PatternFill("solid", fgColor=COLOUR_ERROR),
        "WARNING": PatternFill("solid", fgColor=COLOUR_WARNING),
        "FIXED":   PatternFill("solid", fgColor=COLOUR_FIXED),
    }
    summary_fill = PatternFill("solid", fgColor=COLOUR_SUMMARY)
    bold_font    = Font(bold=True)

    # Summary sheet
    ws_sum = wb["Summary"]
    _apply_header_style(ws_sum, 1, len(summary_df.columns))
    for row_idx in range(2, ws_sum.max_row):
        for col in range(1, len(summary_df.columns) + 1):
            ws_sum.cell(row=row_idx, column=col).fill = summary_fill
    for col in range(1, len(summary_df.columns) + 1):
        ws_sum.cell(row=ws_sum.max_row, column=col).font = bold_font
    ws_sum.freeze_panes = "A2"
    _auto_size_columns(ws_sum)

    # Per-service sheets
    sev_col_idx  = report_cols.index("Severity_Level") + 1
    desc_col_idx = report_cols.index("Issue_Description") + 1

    for svc_id in service_ids:
        svc_name   = service_map.get_name_by_xplor(svc_id) if service_map.is_loaded else svc_id
        sheet_name = sanitise_filename(svc_name or svc_id or "Unknown")[:31]
        ws = wb[sheet_name]

        _apply_header_style(ws, 1, len(report_cols))
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            sev_cell = row[sev_col_idx - 1]
            severity = str(sev_cell.value or "")
            row_fill = fill_map.get(severity)
            if row_fill:
                for cell in row:
                    cell.fill = row_fill
            row[desc_col_idx - 1].alignment = Alignment(wrap_text=True)

        _auto_size_columns(ws)

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT — DUPLICATE PARENT REPORT
# ─────────────────────────────────────────────────────────────────────────────

DUPLICATE_PARENTS_REPORT_FIELDNAMES = [
    "Service_Name",
    "Parent_Legacy_ID",
    "Parent_Name",
    "Matched_On",
    "Parent_CRN",
    "Parent_Slot",
    "Linked_Child",
]


def _collect_intra_file_dup_groups(
    all_rows: list,
    service_map: "ServiceMapping",
) -> list[list[dict]]:
    """
    Uses Union-Find to cluster parent profiles that appear to be the same physical
    person (matching name + DOB + contact) but carry differing base legacy IDs.

    Returns a list of clusters; each cluster is a list of occurrence dicts
    with all fields needed for the duplicate-parent report.
    """
    def _base(lid: str) -> str:
        return re.sub(r'_\d+$', '', lid) if lid else lid

    profiles: list[dict] = []
    for entry in all_rows:
        row      = entry["row"]
        svc_id   = row.get("ServiceID", "").strip()
        svc_name = row.get("Service_Name", "").strip()
        if not svc_name and service_map and service_map.is_loaded:
            svc_name = service_map.get_name_by_xplor(svc_id) or svc_id

        child_name = (
            f"{row.get('Child_First_Name', '').strip()} "
            f"{row.get('Child_Last_Name',  '').strip()}".strip()
        )

        for prefix in ("Parent1", "Parent2"):
            first = row.get(f"{prefix}_First_Name", "").strip().lower()
            last  = row.get(f"{prefix}_Last_Name",  "").strip().lower()
            if not first and not last:
                continue

            dob = row.get(f"{prefix}_DOB", "").strip()

            contacts: set[str] = set()
            for cf in PARENT_PHONE_FIELDS_BY_PREFIX.get(prefix, []):
                v = row.get(cf, "").strip()
                if v:
                    contacts.add(re.sub(r"[\s\-\(\)]", "", v).lower())

            legacy_id   = row.get(f"{prefix}_Legacy_Account_ID", "").strip()
            crn         = row.get(f"{prefix}_CRN",                "").strip()
            parent_name = (
                f"{row.get(f'{prefix}_First_Name', '').strip()} "
                f"{row.get(f'{prefix}_Last_Name',  '').strip()}".strip()
            )

            profiles.append({
                "first_name":   first,
                "last_name":    last,
                "dob":          dob,
                "contacts":     contacts,
                "legacy_id":    legacy_id,
                "crn":          crn,
                "service_name": svc_name,
                "parent_name":  parent_name,
                "parent_slot":  prefix,
                "linked_child": child_name,
            })

    # Union-Find
    n  = len(profiles)
    uf = list(range(n))

    def _find(x: int) -> int:
        while uf[x] != x:
            uf[x] = uf[uf[x]]
            x = uf[x]
        return x

    def _union(x: int, y: int) -> None:
        uf[_find(x)] = _find(y)

    for i in range(n):
        pa = profiles[i]
        for j in range(i + 1, n):
            pb = profiles[j]
            if pa["first_name"] != pb["first_name"] or pa["last_name"] != pb["last_name"]:
                continue
            if not pa["dob"] or not pb["dob"] or pa["dob"] != pb["dob"]:
                continue
            if not (pa["contacts"] & pb["contacts"]):
                continue
            if pa["legacy_id"] and pb["legacy_id"] and _base(pa["legacy_id"]) == _base(pb["legacy_id"]):
                continue
            _union(i, j)

    # Build connected components
    components: defaultdict[int, list[dict]] = defaultdict(list)
    for i, prof in enumerate(profiles):
        components[_find(i)].append(prof)

    # Keep only groups that contain 2+ distinct base legacy IDs
    result: list[list[dict]] = []
    for members in components.values():
        if len(members) < 2:
            continue
        unique_bases = {_base(m["legacy_id"]) for m in members if m["legacy_id"]}
        if len(unique_bases) < 2:
            continue
        result.append(members)

    return result


def write_duplicate_parents_report(
    recorder:    "IssueRecorder",
    output_path: str,
    service_map: "ServiceMapping",
    all_rows:    list,
    row_num_map: dict | None = None,
) -> tuple[int, int]:
    """
    Writes the duplicate-parent Excel report (cross_service_duplicate_parents.xlsx).

    Sheet 'Duplicate Parents':
      One row per occurrence — every service/child combination where the flagged
      parent appears.  Rows are sorted by Parent_Name then Service_Name.

    Returns (cross_service_count, intra_file_group_count).
    """
    _CROSS_TAG = "cross_service_duplicate_parent"

    # ── Intra-file: occurrence rows derived from Union-Find clusters ───────────
    groups      = _collect_intra_file_dup_groups(all_rows, service_map)
    intra_count = len(groups)

    detail_rows: list[dict] = []

    for members in groups:
        # Compute a representative Matched_On for the whole group
        dob = members[0]["dob"]
        contact_freq: dict[str, int] = {}
        for m in members:
            for c in m["contacts"]:
                contact_freq[c] = contact_freq.get(c, 0) + 1
        shared = [c for c, cnt in contact_freq.items() if cnt >= 2]
        matched_on = f"DOB: {dob}, Contact: {shared[0]}" if shared else f"DOB: {dob}"

        for m in sorted(members, key=lambda x: (x["service_name"], x["linked_child"])):
            detail_rows.append({
                "Service_Name":     m["service_name"],
                "Parent_Legacy_ID": m["legacy_id"],
                "Parent_Name":      m["parent_name"],
                "Matched_On":       matched_on,
                "Parent_CRN":       m["crn"],
                "Parent_Slot":      m["parent_slot"],
                "Linked_Child":     m["linked_child"],
            })

    # ── Cross-service: one row per recorder issue ──────────────────────────────
    cross_issues = [i for i in recorder.issues if i.get("_tag") == _CROSS_TAG]
    cross_count  = len(cross_issues)

    row_lookup = {e["row_num"]: e for e in all_rows}

    for issue in cross_issues:
        prefix = issue.get("_parent_slot", "")
        entry  = row_lookup.get(issue.get("Row"))
        if entry:
            row      = entry["row"]
            svc_id   = row.get("ServiceID", "").strip()
            svc_name = row.get("Service_Name", "").strip()
            if not svc_name and service_map and service_map.is_loaded:
                svc_name = service_map.get_name_by_xplor(svc_id) or svc_id
            legacy_id = row.get(f"{prefix}_Legacy_Account_ID", "").strip()
            crn       = row.get(f"{prefix}_CRN",               "").strip()
            linked    = (
                f"{row.get('Child_First_Name', '').strip()} "
                f"{row.get('Child_Last_Name',  '').strip()}".strip()
            )
        else:
            svc_name  = ""
            legacy_id = ""
            crn       = ""
            linked    = issue.get("Child_Name", "")

        detail_rows.append({
            "Service_Name":     svc_name,
            "Parent_Legacy_ID": legacy_id,
            "Parent_Name":      issue.get("_parent_name", ""),
            "Matched_On":       issue.get("_matched_on", ""),
            "Parent_CRN":       crn,
            "Parent_Slot":      prefix,
            "Linked_Child":     linked,
        })

    # ── Sort and write ─────────────────────────────────────────────────────────
    detail_df = pd.DataFrame(detail_rows, columns=DUPLICATE_PARENTS_REPORT_FIELDNAMES)
    if not detail_df.empty:
        detail_df = detail_df.sort_values(
            ["Parent_Name", "Service_Name"], kind="stable"
        ).reset_index(drop=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail_df.to_excel(writer, sheet_name="Duplicate Parents", index=False)

    wb       = load_workbook(output_path)
    dup_fill = PatternFill("solid", fgColor="FFF0E6")   # Pastel peach
    ws       = wb["Duplicate Parents"]

    _apply_header_style(ws, 1, len(DUPLICATE_PARENTS_REPORT_FIELDNAMES))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    matched_col = DUPLICATE_PARENTS_REPORT_FIELDNAMES.index("Matched_On") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = dup_fill
        row[matched_col - 1].alignment = Alignment(wrap_text=True)

    _auto_size_columns(ws)
    wb.save(output_path)

    return cross_count, intra_count


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATION — run_v2
# ─────────────────────────────────────────────────────────────────────────────

def run_v2(
    input_path:                str | None = None,
    service_map:               "ServiceMapping | None" = None,
    output_dir:                str = "output",
    report_path:               str = "output/validation_audit_report_v2.xlsx",
    client_report_path:        str | None = None,
    existing_file_paths:       list[str] | None = None,
    cross_service_report_path: str | None = None,
    input_paths:               list[str] | None = None,
) -> IssueRecorder:
    """
    Full v2 pipeline:
      1. Load input CSV(s) — supports a list via input_paths for multi-file runs
      2. Apply transformations (service ID, legacy IDs, state, email dedup,
         phone leading-zero, consents default)
      3. Run all v1 validation checks against the transformed data
      4. Cross-check parents against existing-service files (if provided)
      5. Write split CSVs (one per service, named {Service_Name}_families_import.csv)
      6. Write Excel audit report (one tab per service + Summary)
      7. Write client-facing Excel report (duplicate parent emails + redundant ECs)
      8. Write cross-service/intra-file duplicate parent report

    input_paths               — list of CSV/XLSX paths to load and merge (preferred
                                for multi-file runs; overrides input_path when given).
    existing_file_paths       — optional list of CSV/XLSX paths from services already
                                in Xplor, used to detect duplicate parents.
    cross_service_report_path — output path for the duplicate-parents report.

    Returns the IssueRecorder containing all issues and transformation logs.
    """
    recorder    = IssueRecorder()
    all_rows:   list[dict] = []
    headers:    set        = set()
    fieldnames: list[str]  = []

    # Ensure we always have a ServiceMapping object (may be a no-op stub)
    svc_map: ServiceMapping = service_map if service_map is not None else ServiceMapping.__new__(ServiceMapping)
    if service_map is None:
        svc_map._loaded = False
        svc_map._qk_to_xplor   = {}
        svc_map._qk_to_name    = {}
        svc_map._xplor_to_name = {}
        svc_map._name_to_xplor = {}

    # ── Load input rows (single path or merged list) ──────────────────────────
    paths_to_load: list[str] = input_paths if input_paths else ([input_path] if input_path else [])
    if not paths_to_load:
        print("\nERROR: No input file(s) specified.")
        sys.exit(1)

    print(f"\n  Loading rows     ...", end="", flush=True)
    row_offset = 2   # row 1 = header; data rows start at 2 per file
    for path in paths_to_load:
        try:
            raw_rows, fns = load_input_file(path)
        except Exception as exc:
            print(f"\nERROR: Could not read '{path}': {exc}")
            sys.exit(1)
        if not fieldnames:
            fieldnames = fns
        for raw_row in raw_rows:
            all_rows.append({
                "row_num":    row_offset,
                "child_name": get_child_name(raw_row),
                "row":        raw_row,
            })
            row_offset += 1

    headers = set(fieldnames)
    print(f" done.  {len(all_rows)} rows loaded from {len(paths_to_load)} file(s).")
    print(f"  Columns detected : {len(headers)}")

    # ── Pass 1: Service ID transformation ────────────────────────────────────
    print("  Mapping service IDs          ...", end="", flush=True)
    for entry in all_rows:
        transform_service_id(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, svc_map,
        )
    print(" done.")

    # ── Filter: keep only rows whose service ID resolved to a known Xplor ID ──
    if svc_map.is_loaded:
        before = len(all_rows)
        all_rows = [
            e for e in all_rows
            if svc_map.is_valid_xplor_id(e["row"].get("ServiceID", "").strip())
        ]
        skipped = before - len(all_rows)
        if skipped:
            print(f"  Skipped {skipped} row(s) whose service ID is not in serviceIDs.csv.")

    # ── Pass 2: Build per-service state fallbacks then normalise states ───────
    print("  Normalising state fields     ...", end="", flush=True)
    service_fallbacks = build_service_state_fallbacks(all_rows)
    for entry in all_rows:
        transform_states(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, service_fallbacks,
        )
    print(" done.")

    # ── Pass 3: In-row email deduplication ────────────────────────────────────
    print("  Deduplicating emails         ...", end="", flush=True)
    for entry in all_rows:
        transform_email_dedup(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, headers,
        )
    print(" done.")

    # ── Pass 4: Child CRN = Parent CRN fix ───────────────────────────────────
    print("  Fixing duplicate CRNs        ...", end="", flush=True)
    for entry in all_rows:
        transform_crn_child_parent_equality(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, headers,
        )
    print(" done.")

    # ── Pass 5: Blank first name fix ─────────────────────────────────────────
    print("  Fixing blank first names     ...", end="", flush=True)
    for entry in all_rows:
        transform_blank_first_names(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, headers,
        )
    print(" done.")

    # ── Pass 6: Legacy ID prefixing (child + parents, P1==P2 dedup) ──────────
    print("  Prefixing legacy IDs         ...", end="", flush=True)
    for entry in all_rows:
        transform_legacy_ids(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, headers,
        )
    print(" done.")

    # ── Pass 7: Phone leading-zero fix ────────────────────────────────────────
    print("  Fixing phone leading zeros   ...", end="", flush=True)
    for entry in all_rows:
        transform_phone_leading_zero(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, headers,
        )
    print(" done.")

    # ── Pass 8: Consents_Photos blank → 'N' ──────────────────────────────────
    print("  Defaulting blank consents    ...", end="", flush=True)
    for entry in all_rows:
        transform_consents_photos(
            entry["row"], entry["row_num"], entry["child_name"],
            recorder, headers,
        )
    print(" done.")

    # ── Pass 9: Per-row validation ────────────────────────────────────────────
    print("  Running per-row validation   ...", end="", flush=True)
    for entry in all_rows:
        row        = entry["row"]
        row_num    = entry["row_num"]
        child_name = entry["child_name"]

        validate_mandatory_child_fields(row, row_num, child_name, recorder, headers)
        validate_mandatory_parent_fields(row, row_num, child_name, recorder, headers)
        validate_emergency_contact_legacy_ids(row, row_num, child_name, recorder, headers)
        validate_date_fields(row, row_num, child_name, recorder, headers)
        validate_crn_format(row, row_num, child_name, recorder, headers)
        validate_phone_fields(row, row_num, child_name, recorder, headers)
        validate_email_fields(row, row_num, child_name, recorder, headers)
        validate_status_field(row, row_num, child_name, recorder, headers)
        validate_gender_fields(row, row_num, child_name, recorder, headers)
        validate_boolean_fields(row, row_num, child_name, recorder, headers)
        validate_state_fields(row, row_num, child_name, recorder, headers)
        validate_postcode_fields(row, row_num, child_name, recorder, headers)
        validate_field_lengths(row, row_num, child_name, recorder, headers)
        validate_service_id(row, row_num, child_name, recorder, headers)
        validate_paired_name_fields(row, row_num, child_name, recorder, headers)
        validate_waitlist_logic(row, row_num, child_name, recorder, headers)
        validate_crn_child_parent_equality(row, row_num, child_name, recorder, headers)
        validate_future_dob(row, row_num, child_name, recorder, headers)
        validate_enrolment_date_not_before_dob(row, row_num, child_name, recorder, headers)
        validate_medicare_number(row, row_num, child_name, recorder, headers)
        validate_consents_photos_videos(row, row_num, child_name, recorder, headers)
        validate_ec_parent_email_redundancy(row, row_num, child_name, recorder, headers)

    print(" done.")

    # ── Pass 10: Cross-row validation ─────────────────────────────────────────
    print("  Running cross-row checks     ...", end="", flush=True)
    check_duplicates(all_rows, recorder, headers)
    check_duplicate_parent_emails(all_rows, recorder, headers)
    check_parent_crn_email_registry(all_rows, recorder, headers)
    check_enrolment_parent_crn_consistency(all_rows, recorder, headers)
    print(" done.")

    # ── Pass 10b: Intra-file duplicate parent check ───────────────────────────
    print("  Checking duplicate parents   ...", end="", flush=True)
    check_intra_file_parent_duplicates(all_rows, recorder)
    print(" done.")

    # ── Pass 10c: Cross-service duplicate parent check (vs existing files) ────
    if existing_file_paths:
        print("  Checking existing svc dupes  ...", end="", flush=True)
        existing_profiles = load_existing_parent_profiles_from_paths(existing_file_paths)
        check_cross_service_parent_duplicates(all_rows, existing_profiles, recorder)
        print(f" done.  ({len(existing_profiles)} existing parent profile(s) checked)")

    # ── Pass 11: Write split CSVs ─────────────────────────────────────────────
    print("\n  Writing import-ready CSV files:")
    _, row_num_map = write_split_csvs(all_rows, fieldnames, output_dir, svc_map)

    # ── Pass 12: Write Excel audit report ────────────────────────────────────
    print(f"\n  Writing audit report         ...", end="", flush=True)
    write_excel_report(recorder, all_rows, report_path, svc_map, row_num_map)
    print(" done.")

    # ── Pass 13: Write client-facing report ──────────────────────────────────
    if client_report_path:
        print(f"  Writing client report        ...", end="", flush=True)
        write_client_excel_report(recorder, all_rows, client_report_path, svc_map, row_num_map)
        print(" done.")

    # ── Pass 14: Write duplicate parents report ───────────────────────────────
    if cross_service_report_path:
        print(f"  Writing dup-parents report   ...", end="", flush=True)
        cross_n, intra_n = write_duplicate_parents_report(
            recorder, cross_service_report_path, svc_map, all_rows,
        )
        print(f" done.  (cross-service: {cross_n}, within-upload: {intra_n} group(s))")

    return recorder


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT-COMPATIBLE INTERFACE
# Allows app.py to call run_v2 with uploaded files (in-memory)
# ─────────────────────────────────────────────────────────────────────────────

def run_v2_from_bytes(
    file_bytes:       bytes,
    filename:         str,
    service_map:      ServiceMapping,
    existing_files:   list[tuple[bytes, str]] | None = None,
    include_waitlist: bool = True,
) -> tuple["IssueRecorder", list[dict], list[str]]:
    """
    In-memory variant of run_v2 — accepts raw file bytes (CSV or XLSX) from
    a Streamlit upload or any other in-memory source.
    Does NOT write any files; returns (recorder, all_rows, fieldnames) for the
    caller to handle output as needed.

    Args:
        file_bytes        — raw bytes of the uploaded file
        filename          — original filename including extension (used to detect format)
        service_map       — loaded ServiceMapping instance
        existing_files    — optional list of (bytes, filename) tuples from other services
                            already in Xplor, used to detect duplicate parents
        include_waitlist  — when False, only rows with Status == 'active' are processed;
                            Waitlist rows are dropped before validation and output
    """
    recorder:   IssueRecorder = IssueRecorder()
    all_rows:   list[dict]    = []

    raw_rows, fieldnames = load_input_bytes(file_bytes, filename)
    headers = set(fieldnames)

    for row_idx, raw_row in enumerate(raw_rows, start=2):
        all_rows.append({
            "row_num":    row_idx,
            "child_name": get_child_name(raw_row),
            "row":        raw_row,
        })

    # Import scope filter — drop Waitlist rows when the user selects Active only
    if not include_waitlist:
        all_rows = [
            e for e in all_rows
            if e["row"].get("Status", "").strip().lower() != "waitlist"
        ]

    # Transformations
    for entry in all_rows:
        transform_service_id(entry["row"], entry["row_num"], entry["child_name"], recorder, service_map)

    # Filter to only rows whose service mapped to a known Xplor ID
    if service_map.is_loaded:
        all_rows = [
            e for e in all_rows
            if service_map.is_valid_xplor_id(e["row"].get("ServiceID", "").strip())
        ]

    service_fallbacks = build_service_state_fallbacks(all_rows)
    for entry in all_rows:
        transform_states(entry["row"], entry["row_num"], entry["child_name"], recorder, service_fallbacks)

    for entry in all_rows:
        transform_email_dedup(entry["row"], entry["row_num"], entry["child_name"], recorder, headers)

    for entry in all_rows:
        transform_crn_child_parent_equality(entry["row"], entry["row_num"], entry["child_name"], recorder, headers)

    for entry in all_rows:
        transform_blank_first_names(entry["row"], entry["row_num"], entry["child_name"], recorder, headers)

    for entry in all_rows:
        transform_legacy_ids(entry["row"], entry["row_num"], entry["child_name"], recorder, headers)

    for entry in all_rows:
        transform_phone_leading_zero(entry["row"], entry["row_num"], entry["child_name"], recorder, headers)

    for entry in all_rows:
        transform_consents_photos(entry["row"], entry["row_num"], entry["child_name"], recorder, headers)

    # Per-row validation
    for entry in all_rows:
        row, row_num, child_name = entry["row"], entry["row_num"], entry["child_name"]
        validate_mandatory_child_fields(row, row_num, child_name, recorder, headers)
        validate_mandatory_parent_fields(row, row_num, child_name, recorder, headers)
        validate_emergency_contact_legacy_ids(row, row_num, child_name, recorder, headers)
        validate_date_fields(row, row_num, child_name, recorder, headers)
        validate_crn_format(row, row_num, child_name, recorder, headers)
        validate_phone_fields(row, row_num, child_name, recorder, headers)
        validate_email_fields(row, row_num, child_name, recorder, headers)
        validate_status_field(row, row_num, child_name, recorder, headers)
        validate_gender_fields(row, row_num, child_name, recorder, headers)
        validate_boolean_fields(row, row_num, child_name, recorder, headers)
        validate_state_fields(row, row_num, child_name, recorder, headers)
        validate_postcode_fields(row, row_num, child_name, recorder, headers)
        validate_field_lengths(row, row_num, child_name, recorder, headers)
        validate_service_id(row, row_num, child_name, recorder, headers)
        validate_paired_name_fields(row, row_num, child_name, recorder, headers)
        validate_waitlist_logic(row, row_num, child_name, recorder, headers)
        validate_crn_child_parent_equality(row, row_num, child_name, recorder, headers)
        validate_future_dob(row, row_num, child_name, recorder, headers)
        validate_enrolment_date_not_before_dob(row, row_num, child_name, recorder, headers)
        validate_medicare_number(row, row_num, child_name, recorder, headers)
        validate_consents_photos_videos(row, row_num, child_name, recorder, headers)
        validate_ec_parent_email_redundancy(row, row_num, child_name, recorder, headers)

    # Cross-row validation
    check_duplicates(all_rows, recorder, headers)
    check_duplicate_parent_emails(all_rows, recorder, headers)
    check_parent_crn_email_registry(all_rows, recorder, headers)
    check_enrolment_parent_crn_consistency(all_rows, recorder, headers)

    # Intra-file duplicate parent check (always runs)
    check_intra_file_parent_duplicates(all_rows, recorder)

    # Cross-service duplicate parent check (only when existing files provided)
    if existing_files:
        existing_profiles = load_existing_parent_profiles_from_bytes(existing_files)
        check_cross_service_parent_duplicates(all_rows, existing_profiles, recorder)

    return recorder, all_rows, fieldnames


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    script_dir = Path(__file__).parent

    # ── Folder layout ────────────────────────────────────────────────────────
    # input/              — place migration CSV file(s) AND serviceIDs.csv here
    # output/             — all generated files are written here
    input_folder  = script_dir / "input"
    output_folder = script_dir / "output"

    # Create input/output folders automatically if they do not yet exist
    input_folder.mkdir(exist_ok=True)
    output_folder.mkdir(exist_ok=True)

    # ── Locate input files ────────────────────────────────────────────────────
    if len(sys.argv) >= 2:
        # Explicit path(s) on the command line (space-separated)
        input_csvs = [sys.argv[1]]
    else:
        # Auto-discover all CSV/XLSX files in input/ (excluding serviceIDs.csv)
        candidates = sorted(
            f for f in input_folder.iterdir()
            if f.suffix.lower() in (".csv", ".xlsx", ".xls")
            and f.name.lower() != "serviceids.csv"
        )
        if not candidates:
            print()
            print("ERROR: No CSV or XLSX file found in the input folder:")
            print(f"  {input_folder}")
            print()
            print("Please place your QikKids export CSV(s) in the 'input' folder and re-run.")
            print()
            sys.exit(1)
        input_csvs = [str(f) for f in candidates]

    # ── Locate the service mapping file ──────────────────────────────────────
    if len(sys.argv) >= 3:
        service_map_csv = sys.argv[2]
    else:
        service_map_csv = str(input_folder / "serviceIDs.csv")

    # ── Output paths (all files written to output/) ───────────────────────────
    report_path               = str(output_folder / "validation_audit_report_v2.xlsx")
    client_report_path        = str(output_folder / "client_audit_report.xlsx")
    cross_service_report_path = str(output_folder / "cross_service_duplicate_parents.xlsx")

    print()
    print("=" * 65)
    print("  Xplor Data Migration — Validator & Transformer  v2.0")
    print("=" * 65)
    for p in input_csvs:
        print(f"  Input file       : {p}")
    print(f"  Service map      : {service_map_csv}")
    print(f"  Output folder    : {output_folder}")
    print(f"  Audit report     : {report_path}")
    print(f"  Client report    : {client_report_path}")
    print("=" * 65)

    svc_map  = ServiceMapping(service_map_csv)
    recorder = run_v2(
        input_paths               = input_csvs,
        service_map               = svc_map,
        output_dir                = str(output_folder),
        report_path               = report_path,
        client_report_path        = client_report_path,
        cross_service_report_path = cross_service_report_path,
    )

    errors   = recorder.error_count()
    warnings = recorder.warning_count()
    fixed    = recorder.fixed_count()
    total    = len(recorder.issues)

    print()
    print("=" * 65)
    print("  Pipeline complete.")
    print(f"  Total log entries : {total}")
    print(f"  Errors            : {errors}")
    print(f"  Warnings          : {warnings}")
    print(f"  Auto-fixed        : {fixed}")
    print("=" * 65)
    print()

    if errors > 0:
        print(f"  [!] {errors} error(s) found — please resolve before importing.")
    elif warnings > 0:
        print(f"  [!] {warnings} warning(s) found — please review before importing.")
    else:
        print("  [OK] No errors or warnings. Files are ready for import.")

    if fixed > 0:
        print(f"  [FIX] {fixed} item(s) were automatically corrected. "
              f"Review the FIXED entries in the audit report.")
    print()
